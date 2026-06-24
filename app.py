# =========================================================
# PYTHON STANDARD LIBRARY
# =========================================================

import os
import re
import time
import random
import smtplib
import subprocess
import math
import io
import urllib.request

from datetime import datetime, date, timedelta
from functools import wraps

from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# =========================================================
# THIRD PARTY PACKAGES
# =========================================================

import pandas as pd
import pdfkit
import razorpay

from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv

from flask import (
    Flask,
    flash,
    render_template,
    render_template_string,
    session,
    request,
    redirect,
    url_for,
    send_file,
    jsonify,
    make_response
)

from flask_bcrypt import Bcrypt
from flask_mail import Mail, Message
from flask_wtf.csrf import CSRFProtect

from werkzeug.security import (
    generate_password_hash,
    check_password_hash
)

from werkzeug.utils import secure_filename


# =========================================================
# LOCAL PROJECT IMPORTS
# =========================================================

from db import get_connection

 


# Load env variables from .env file
load_dotenv()

# Initialize Flask app
app = Flask(__name__)

# Secret key for sessions 
app.secret_key = os.getenv("SECRET_KEY")

if not app.secret_key:
    raise Exception("SECRET_KEY missing in .env")

# Enable CSRF protection
csrf = CSRFProtect(app)

# =========================================================
# PDF CONFIGURATION
# =========================================================

WKHTMLTOPDF_PATH = os.getenv("WKHTMLTOPDF_PATH")
if not WKHTMLTOPDF_PATH:
    raise Exception("WKHTMLTOPDF_PATH missing in .env")

pdf_config = pdfkit.configuration(
    wkhtmltopdf=WKHTMLTOPDF_PATH
)

# =========================================================
# SESSION SECURITY CONFIGURATION
# =========================================================

# Prevent JavaScript access to cookies
app.config["SESSION_COOKIE_HTTPONLY"] = True

# Use True in production with HTTPS
app.config["SESSION_COOKIE_SECURE"] = False  #LAter make it True

# Protect against CSRF-like cross-site behavior
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# Session auto-expiry
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)

# File Maximum Upload Size 10 MB (for Excel imports, etc.)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024



# Initialize Bcrypt for password hashing
bcrypt = Bcrypt(app)


# =========================================================
# MAIL CONFIGURATION
# =========================================================

app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USE_SSL"] = False

# your gmail
app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME")

# gmail app password (not normal password)
app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD")

app.config["MAIL_DEFAULT_SENDER"] = os.getenv("MAIL_USERNAME")

# initialize mail
mail = Mail(app)

# Razorpay Configuration

RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID")
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET")

if not RAZORPAY_KEY_ID or not RAZORPAY_KEY_SECRET:
    raise Exception("Razorpay keys missing in .env")
 

# =========================================================
# 🛠️ COMMON HELPER FUNCTIONS (USED ACROSS APP)
# =========================================================
 
 
# =========================================================
# GLOBAL ERROR HANDLERS
# =========================================================

@app.errorhandler(500)
def internal_server_error(e):
    """
    Handles unexpected server errors.
    """

    print("❌ INTERNAL SERVER ERROR:", e)

    return "Something went wrong ❌", 500


@app.errorhandler(404)
def page_not_found(e):
    """
    Handles invalid URLs.
    """

    return "Page not found ❌", 404

# =========================================================
# GLOBAL CURRENT YEAR
# Available in all templates
# =========================================================

@app.context_processor
def inject_year():

    return {
        "current_year": datetime.now().year
    }

 #========================================================
# 🌟 GLOBAL SYSTEM SETTING
# Available in all templates as 'global_settings'
# =========================================================

@app.context_processor
def inject_system_settings():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                system_name,
                system_logo,
                       
                support_email,
                support_phone,
                       
                favicon,
                       
                footer_text,
                powered_by,
                erp_version,
                website_url,
                       
                updated_at
                FROM system_settings
                LIMIT 1
        """)

        row = cursor.fetchone()

        if row:

            return {
                "global_settings": {
                    "system_name": row[0],
                    "system_logo": row[1],

                    "support_email": row[2],
                    "support_phone": row[3],

                    "favicon": row[4],

                    "footer_text": row[5],
                    "powered_by": row[6],
                    "erp_version": row[7],
                    "website_url": row[8],

                    "updated_at": row[9]
                }
            }

    except Exception as e:

        print("GLOBAL SETTINGS ERROR:", e)

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

    if not row:
        return {
            "global_settings": {}
        }

# =========================================================
# GLOBAL LEAD COUNT
# Used in Super Admin Dashboard
# =========================================================
    
@app.context_processor
def inject_lead_count():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT COUNT(*)
            FROM lead_requests
            WHERE status='New'
        """)

        result = cursor.fetchone()

        count = result[0] if result else 0

        return {
            "new_leads_count": count
        }

    except Exception as e:

        print(
            "LEAD COUNT ERROR:",
            e
        )

        return {
            "new_leads_count": 0
        }

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📧 GLOBAL EMAIL SENDER
# =========================================================

def send_email(

    to_email,
    subject,
    body,
    attachment_path=None,
    attachment_name=None

):

    conn = None
    cursor = None
    server = None

    try:

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # GET SMTP SETTINGS
        # =========================================

        cursor.execute("""

            SELECT

                smtp_email,
                smtp_password,
                smtp_server,
                smtp_port,
                smtp_tls

            FROM system_settings
            WHERE id = 1

        """)

        smtp = cursor.fetchone()

        if not smtp:
            return False

        smtp_email = smtp[0]
        smtp_password = smtp[1]
        smtp_server = smtp[2]
        smtp_port = int(smtp[3])
        smtp_tls = smtp[4]

        # =========================================
        # CREATE EMAIL
        # =========================================

        message = MIMEMultipart()

        message["From"] = smtp_email
        message["To"] = to_email

        message["Subject"] = Header(
            subject,
            "utf-8"
        )

        # =========================================
        # EMAIL BODY
        # =========================================

        message.attach(

            MIMEText(
                body,
                "html",
                "utf-8"
            )

        )

        # =========================================
        # PDF ATTACHMENT
        # =========================================

        if attachment_path and attachment_name:

            from email.mime.base import MIMEBase
            from email import encoders

            with open(attachment_path, "rb") as file:

                part = MIMEBase(
                    "application",
                    "octet-stream"
                )

                part.set_payload(
                    file.read()
                )

            encoders.encode_base64(part)

            part.add_header(

                "Content-Disposition",

                f'attachment; filename="{attachment_name}"'

            )

            message.attach(part)

        # =========================================
        # SMTP CONNECTION
        # =========================================

        server = smtplib.SMTP(

            smtp_server,
            smtp_port,
            timeout=20

        )

        # =========================================
        # TLS SECURITY
        # =========================================

        if str(smtp_tls).strip() == "Enabled":

            server.starttls()

        # =========================================
        # LOGIN
        # =========================================

        server.login(

            smtp_email,
            smtp_password

        )

        # =========================================
        # SEND EMAIL
        # =========================================

        server.sendmail(

            smtp_email,
            to_email,
            message.as_bytes()

        )

        print("✅ EMAIL SENT")

        return True

    except Exception as e:

        print(
            "❌ EMAIL SEND ERROR:",
            str(e)
        )

        return False

    finally:

        # =========================================
        # CLOSE SMTP
        # =========================================

        try:
            if server:
                server.quit()
        except Exception:
            pass

        # =========================================
        # CLOSE DB
        # =========================================

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# MAINTENANCE MODE CHECK
# Runs before every request
# =========================================================
 
@app.before_request
def check_maintenance_mode():

    conn = None
    cursor = None

    try:

        # Static files
        if request.path.startswith("/static/"):
            return

        # Public routes
        allowed_paths = [

            "/login",
            "/logout",

            "/superadmin/login",
            "/superadmin/logout"
        ]

        if request.path in allowed_paths:
            return

        # Super Admin always allowed
        if request.path.startswith("/superadmin"):

            if session.get("admin_logged_in") is True:
                return

        conn = get_connection()

        if not conn:
            return

        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                maintenance_mode,
                maintenance_message
            FROM system_settings
            WHERE id = 1
        """)

        row = cursor.fetchone()

        if not row:
            return

        maintenance_mode = (
            row[0] or ""
        ).strip().upper()

        maintenance_message = (
            row[1]
            or "ERP system is currently under maintenance."
        )

        if maintenance_mode == "ON":

            return (
                render_template(
                    "maintenance.html",
                    message=maintenance_message
                ),
                503
            )

    except Exception as e:

        print(
            "❌ MAINTENANCE CHECK ERROR:",
            e
        )

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
            
# =========================================================
# 🌟 GLOBAL BRANDING FUNCTION
# =========================================================

        # @app.context_processor
        # def inject_branding_settings():

        #     conn = get_connection()
        #     cursor = conn.cursor()

        #     try:

        #         cursor.execute("""
        #             SELECT
        #                 primary_color,
        #                 secondary_color,
        #                 button_color,
        #                 system_logo,
        #                 favicon,
        #                 system_name,
        #                 footer_text,
        #                 erp_version
        #             FROM system_settings
        #             WHERE id = 1
        #         """)

        #         branding = cursor.fetchone()

        #         return dict(
        #             global_settings=branding
        #         )

        #     except Exception as e:

        #         print("Branding Context Error:", e)

        #         return dict(
        #             global_settings=None
        #         )

        #     finally:

        #         cursor.close()
        #         conn.close()

# =========================================================
# 🌟 SCHOOL FEATURE MODULES
# =========================================================

@app.context_processor
def inject_feature_modules():

    conn = None
    cursor = None

    try:

        school_id = session.get(
            "clerk_school_id"
        )

        if not school_id:

            return {
                "feature_modules": {}
            }

        conn = get_connection()

        cursor = conn.cursor()

        cursor.execute("""

            SELECT

                enable_tc_management,
                enable_bonafide_management,
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board

            FROM schools

            WHERE school_id = %s

        """, (school_id,))

        row = cursor.fetchone()

        if not row:

            return {
                "feature_modules": {}
            }

        # =========================================
        # Convert Enabled/Disabled values
        # into True/False flags
        # =========================================

        feature_modules = {

            "tc":
                row[0] == "Enabled",

            "bonafide":
                row[1] == "Enabled",

            "import_export":
                row[2] == "Enabled",

            "attendance":
                row[3] == "Enabled",

            "fees":
                row[4] == "Enabled",

            "teachers":
                row[5] == "Enabled",

            "results":
                row[6] == "Enabled",

            "timetable":
                row[7] == "Enabled",

            "notice_board":
                row[8] == "Enabled"

        }

        return {
            "feature_modules": feature_modules
        }

    except Exception as e:

        print(
            "FEATURE MODULE ERROR:",
            e
        )

        return {
            "feature_modules": {}
        }

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# GET SCHOOL DETAILS (GLOBAL FIX)
# =========================================================
def get_school_details(school_id):

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                school_id,
                name,
                udise_no
            FROM schools
            WHERE school_id = %s
        """, (school_id,))

        school = cursor.fetchone()

        if not school:
            return None

        return {
            "school_id": school[0],
            "school_name": school[1],
            "school_udise": school[2]
        }
    except Exception as e:

        print(
            "GET SCHOOL DETAILS ERROR:",
            e
        )

        return None

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

 # =========================================================
# APPLY PLAN FEATURES TO SCHOOL
# =========================================================

def apply_plan_features(
    school_id,
    plan_id
):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # GET PLAN FEATURES
        # =========================================

        cursor.execute("""

            SELECT

                enable_tc_management,
                enable_bonafide_management,
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board

            FROM subscription_plans

            WHERE id = %s

        """, (plan_id,))

        plan = cursor.fetchone()

        if not plan:

            return False

        # =========================================
        # UPDATE SCHOOL FEATURES
        # =========================================

        cursor.execute("""

            UPDATE schools

            SET

                enable_tc_management = %s,
                enable_bonafide_management = %s,
                enable_import_export = %s,
                enable_attendance = %s,
                enable_fee_management = %s,
                enable_teacher_management = %s,
                enable_results = %s,
                enable_timetable = %s,
                enable_notice_board = %s

            WHERE school_id = %s

        """, (

            plan[0],
            plan[1],
            plan[2],
            plan[3],
            plan[4],
            plan[5],
            plan[6],
            plan[7],
            plan[8],

            school_id

        ))

        conn.commit()

        print(
            f"✅ Features Applied | School={school_id} Plan={plan_id}"
        )

        return True

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ APPLY PLAN FEATURES ERROR:",
            e
        )

        return False

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# ---------------------------------------------------------
# 🔹 HANDLE NULL / EMPTY VALUES (MAINLY FOR EXCEL IMPORT)
# ---------------------------------------------------------
def safe_value(val):
    """
    Converts Excel/DB value safely:
    - If NaN → returns None
    - Else → returns cleaned string
    """
 

    if pd.isna(val):
        return None

    return str(val).strip()


 # ---------------------------------------------------------
# 🔹 HANDLE DATE FROM EXCEL (IMPORT SIDE)
# Converts Excel date into YYYY-MM-DD
# ---------------------------------------------------------
def safe_date(val):
    """
    Converts Excel date to DB format (YYYY-MM-DD)

    Example:
    Excel → 2008-09-23 → '2008-09-23'

    Returns:
    - Proper string date OR
    - None if invalid
    """
    import pandas as pd

    if pd.isna(val):
        return None

    try:
        return pd.to_datetime(val).strftime("%Y-%m-%d")
    except Exception:
        return None


# ---------------------------------------------------------
# 🔹 VALIDATION HELPER 
#---------------------------------------------------------

def is_valid_email(email):
    """
    Validate email format.
    """

    pattern = r"^[^\s@]+@[^\s@]+\.[^\s@]+$"

    return bool(re.match(pattern, email))

def is_valid_phone(phone):
    """
    Validate phone number format.
    """
    return phone.isdigit() and len(phone) == 10


def is_valid_aadhaar(aadhaar):
    """
    Validate Aadhaar number format.
    """
    return aadhaar.isdigit() and len(aadhaar) == 12


# ---------------------------------------------------------
# PARSE FORM DATE
# Converts string into datetime object
# ---------------------------------------------------------
def parse_date(date_str):
    """
    Converts form string → datetime object

    Example:
    '2026-04-17' → datetime object
    '17-04-2026' → datetime object

    Used in:
    - Add Student
    - TC Form
    - Bonafide Form
    """
    from datetime import datetime

    if not date_str:
        return None

    try:
        # FORMAT: YYYY-MM-DD
        return datetime.strptime(
            date_str,
            "%Y-%m-%d"
        )

    except Exception:
        try:
            # FORMAT: DD-MM-YYYY
            return datetime.strptime(
                date_str,
                "%d-%m-%Y"
            )

        except Exception:
            return None


# ---------------------------------------------------------
# 🔹 FORMAT DATE FOR UI / PDF DISPLAY
# ---------------------------------------------------------
def format_date(d):
    """
    Converts DB date → readable format

    Example:
    2026-04-17 → 17-04-2026

    Used in:
    - TC view
    - Bonafide view
    - Print / PDF
    """
    if not d:
        return ""

    try:

        if hasattr(d, "strftime"):
            return d.strftime("%d-%m-%Y")

        return str(d)

    except Exception:
        return str(d)

 # ---------------------------------------------------------
# SAFE STRING
# Returns cleaned string or empty string
# ---------------------------------------------------------
def safe_str(value):

    if value is None:
        return ""

    return str(value).strip()

# ---------------------------------------------------------
# 🔹 GET SCHOOL CODE
# ---------------------------------------------------------
def get_school_code(school_id):

    conn = None
    cursor = None

    """
    Fetch school_code from DB

    Used in:
    - Admission No → ABC-ADM-0001
    - TC No        → ABC-TC-0001
    - Bonafide No  → ABC-BON-0001
    """

    try:

        conn = get_connection()

        cursor = conn.cursor()

        cursor.execute("""

            SELECT school_code

            FROM schools

            WHERE school_id = %s

        """, (school_id,))

        row = cursor.fetchone()

        return row[0] if row else "SCH"

    except Exception as e:

        print(
            "GET SCHOOL CODE ERROR:",
            e
        )

        return "SCH"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 # ---------------------------------------------------------
# SAFE DATETIME FOR SORTING
# Converts date and datetime into same format
# ---------------------------------------------------------
def normalize_datetime(value):

    from datetime import datetime, date

    if not value:
        return datetime.min

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(
            value,
            datetime.min.time()
        )

    return datetime.min

# ---------------------------------------------------------
# 🔹 SUBSCRIPTION CHECK DECORATOR
# Ensures school has active subscription before accessing certain routes
# ---------------------------------------------------------            

def subscription_required(f):

    @wraps(f)
    def decorated_function(*args, **kwargs):

        school_id = session.get("clerk_school_id")

        if not school_id:
            return redirect(url_for("login"))

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                status,
                end_date
            FROM subscriptions
            WHERE school_id = %s
            ORDER BY id DESC
            LIMIT 1
        """, (school_id,))

        sub = cursor.fetchone()

        cursor.close()
        conn.close()

        if not sub:

            flash(
                "No active subscription found.",
                "danger"
            )

            return redirect(
                url_for("renew_subscription")
            )

        status = sub[0]
        end_date = sub[1]

        if end_date:

            if hasattr(end_date, "date"):
                end_date = end_date.date()

            remaining_days = (
                end_date - datetime.now().date()
            ).days

        else:

            remaining_days = -1

        if (
            status == "expired"
            or remaining_days < 0
        ):

            flash(
                "Your subscription has expired. Please renew.",
                "danger"
            )

            return redirect(
                url_for("renew_subscription")
            )

        return f(*args, **kwargs)

    return decorated_function
# ---------------------------------------------------------
# SECUIRTY SETTING GLOBAL
# Allows only logged-in Super Admin users
# ---
 
def get_security_setting(key, default_value):

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                password_length,
                login_attempt_limit,
                session_timeout
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return default_value

        value = settings.get(key)

        if value is None:
            return default_value

        return int(value)

    except Exception as e:
        print("❌ SECURITY SETTING FETCH ERROR:", e)
        return default_value

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
# ---------------------------------------------------------
# ADMIN AUTHORIZATION
# Allows only logged-in Super Admin users
# ---
    
def admin_required(f):

    @wraps(f)
    def wrapper(*args, **kwargs):

        if (
            not session.get("admin_logged_in")
            or session.get("admin_role") != "admin"
        ):

            # AJAX request
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":

                return jsonify({

                    "success": False,
                    "message": "Admin login required"

                }), 401

            # Normal page request
            return redirect(
                url_for("superadmin_login")
            )

        return f(*args, **kwargs)

    return wrapper

# ---------------------------------------------------------
# LOGIN AUTHORIZATION
# Allows access to authenticated Clerk OR Admin users
# ---------------------------------------------------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):

        clerk_logged_in = session.get("clerk_logged_in")
        clerk_role = session.get("clerk_role")

        admin_logged_in = session.get("admin_logged_in")
        admin_role = session.get("admin_role")

        clerk_ok = (
            clerk_logged_in is True
            and clerk_role == "clerk"
        )

        admin_ok = (
            admin_logged_in is True
            and admin_role == "admin"
        )

        if not clerk_ok and not admin_ok:
            return redirect(url_for("login"))

        return f(*args, **kwargs)

    return wrapper

 # =========================================================
# SCHOOL FEATURE COLUMNS
# =========================================================

FEATURE_COLUMNS = (

    "enable_tc_management",
    "enable_bonafide_management",
    "enable_import_export",
    "enable_attendance",
    "enable_fee_management",
    "enable_teacher_management",
    "enable_results",
    "enable_timetable",
    "enable_notice_board"

)

# =========================================================
# 🔒 FEATURE ACCESS CONTROL
# =========================================================

def feature_required(feature_column):

    def decorator(f):

        @wraps(f)
        def decorated_function(*args, **kwargs):

            conn = None
            cursor = None

            try:
                # =========================================
                # BLOCK INVALID COLUMN ACCESS
                # =========================================
                
                if feature_column not in FEATURE_COLUMNS:

                    return "Invalid Feature ❌", 400

                # =========================================
                #  ADMIN BYPASS
                # =========================================

                if (
                    session.get("admin_logged_in") is True
                    and session.get("admin_role") == "admin"
                ):

                    return f(*args, **kwargs)

                # =========================================
                # GET SCHOOL
                # =========================================

                school_id = session.get(
                    "clerk_school_id"
                )

                if not school_id:

                    return redirect(
                        url_for("login")
                    )

                # =========================================
                # DB CONNECTION
                # =========================================

                conn = get_connection()

                cursor = conn.cursor()

                # =========================================
                # SAFE QUERY
                # =========================================

                query = f"""

                    SELECT {feature_column}

                    FROM schools

                    WHERE school_id = %s

                """

                cursor.execute(
                    query,
                    (school_id,)
                )

                result = cursor.fetchone()

                # =========================================
                # FEATURE DISABLED
                # =========================================

                if (
                    not result
                    or result[0] != "Enabled"
                ):

                    return """

                    <h2 style='font-family:sans-serif;
                               padding:40px;
                               color:red;'>

                        Feature Disabled By Admin ❌

                    </h2>

                    """

                # =========================================
                # ALLOW ACCESS
                # =========================================

                return f(*args, **kwargs)

            except Exception as e:

                print(
                    "FEATURE ACCESS ERROR:",
                    e
                )

                return "Something went wrong ❌",500

            finally:

                if cursor:
                    cursor.close()

                if conn:
                    conn.close()

        return decorated_function

    return decorator

# =========================================================
# 🛡 SCHOOL FEATURE ACCESS CHECK
# PURPOSE:
# Check module access school-wise
# Supports SaaS + subscription based ERP
# =========================================================

def is_module_enabled(module_name):

    conn = None
    cursor = None

    try:

        # =========================================
        # GET CURRENT SCHOOL
        # =========================================

        school_id = session.get(
            "clerk_school_id"
        )

        if not school_id:
            return False

 
                
        # =========================================
        # BLOCK INVALID MODULE ACCESS
        # =========================================

        if module_name not in FEATURE_COLUMNS:
            return False

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # SCHOOL SPECIFIC CHECK
        # =========================================

        query = f"""
            SELECT {module_name}
            FROM schools
            WHERE school_id = %s
        """

        cursor.execute(
            query,
            (school_id,)
        )

        row = cursor.fetchone()

        # =========================================
        # ENABLED CHECK
        # =========================================

        if row and str(row[0]).strip() == "Enabled":
            return True

        return False

    except Exception as e:

        print(
            "❌ MODULE CHECK ERROR:",
            e
        )

        return False

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# ---------------------------------------------------------
# CREATE SCHOOL SUBSCRIPTION
# Used during:
# - New School Creation
# - Subscription Renewal
# - Plan Upgrade
# ---------------------------------------------------------

def create_subscription(
    school_id,
    plan_id,
    plan_name,
    days,
    amount
):
    """
    Create a new subscription record.

    Returns:
        True  -> Success
        False -> Failure
    """

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        if days <= 0:
            return False

        start_date = date.today()

        end_date = (
            start_date
            + timedelta(days=days)
        )

        if not school_id:
            return False

        amount = amount or 0   

        cursor.execute("""

            INSERT INTO subscriptions
            (
                school_id,
                plan_id,
                plan_name,
                start_date,
                end_date,
                status,
                amount
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)

        """, (

            school_id,
            plan_id,
            plan_name,
            start_date,
            end_date,
            "active",
            amount

        ))

        conn.commit()

        return True

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ CREATE SUBSCRIPTION ERROR:",
            e
        )

        return False

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 📲 SMS TEST SENDER
# PURPOSE:
# Temporary function to test SMS gateway only
# =========================================================

def send_test_birthday_sms(mobile):

    try:

        mobile = "".join(
            filter(str.isdigit, mobile)
        )

        if len(mobile) != 10:
            return False, "Invalid mobile number"

        sms_api_url = os.getenv("SMS_API_URL")
        sms_user_id = os.getenv("SMS_USER_ID")
        sms_password = os.getenv("SMS_PASSWORD")

        if not sms_password:
            return False, "SMS password missing in .env"
        
        template_id = os.getenv("SMS_BIRTHDAY_TEMPLATE_ID")

        if not sms_api_url or not sms_user_id or not template_id:
            return False, "SMS settings missing in .env"

        message = (
            "Dear Prajwal , Wish you many many happy returns of the day, Regards, Amol Arun Thakre, Shri Prabhu Softlink Pvt. "
            "Ltd Amravati. Whatsapp: 07212670525, Phone: 07212568331, 07212970310, 8888244176. Email adv.thakre@gmail.com"
        )

        url = (
        f"{sms_api_url}"
        f"?ID={quote_plus(sms_user_id)}"
        f"&Pwd={quote_plus(sms_password)}"
        f"&PhNo={quote_plus(mobile)}"
        f"&Text={quote_plus(message)}"
        f"&TemplateID={quote_plus(template_id)}"
    )

        with urllib.request.urlopen(url, timeout=15) as response:

            result = response.read().decode("utf-8", errors="ignore")

        print("✅ SMS API RESPONSE:", result)
        print("USER:", sms_user_id)
        print("PASS:", sms_password)
        print("TEMPLATE:", template_id)
        print("URL:", url)

        return True, result

    except Exception as e:

        print("❌ SMS TEST ERROR:", e)

        return False, str(e)


# =========================================================
# 📲 TEMP SMS TEST ROUTE
# PURPOSE:
# Test SMS gateway only
# REMOVE AFTER TESTING
# =========================================================

@app.route("/superadmin/test-sms")
@admin_required
def test_sms():

    mobile = (
        request.args.get("mobile") or ""
    ).strip()

    if not mobile:
        return "Mobile number required ❌ Example: /superadmin/test-sms?mobile=9876543210"

    success, response = send_test_birthday_sms(mobile)

    if success:
        return f"SMS request sent ✅<br>Gateway Response: {response}"

    return f"SMS failed ❌<br>Error: {response}"


# =========================================================
# PUBLIC PAGES
# =========================================================

# -----------------------------------------
# Privacy Policy
# -----------------------------------------

@app.route("/privacy-policy")
def privacy_policy():
    """
    Privacy Policy Page
    """
    return render_template(
        "privacy_policy.html"
    )


# -----------------------------------------
# Terms & Conditions
# -----------------------------------------

@app.route("/terms-and-conditions")
def terms_conditions():
    """
    Terms & Conditions Page
    """
    return render_template(
        "terms_conditions.html"
    )


# -----------------------------------------
# Home Page
# -----------------------------------------

@app.route("/")
def home():
    """
    Landing Page
    """
    return render_template(
        "index.html"
    )


# -----------------------------------------
# Demo Page
# -----------------------------------------

@app.route("/demo")
def demo():
    """
    Product Demo Page
    """
    return render_template(
        "demo.html"
    )



# =========================================================
# 🔐 CLERK LOGIN
# Handles clerk login, school status, subscription status,
# failed attempts, and secure session creation.
# =========================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    # =====================================================
    # SHOW LOGIN PAGE
    # =====================================================
    if request.method == "GET":
        return render_template("auth/login.html")

    # =====================================================
    # GET FORM DATA
    # =====================================================
    email = (request.form.get("email") or "").strip().lower()
    password = (request.form.get("password") or "").strip()

    # =====================================================
    # BASIC VALIDATION
    # =====================================================
    if not email or not password:
        return "Email and password required ❌"

    if not is_valid_email(email):
        return "Invalid email format ❌"

    # =====================================================
    # SESSION BASED FAILED LOGIN LIMIT
    # Protects same browser/session from repeated attempts
    # =====================================================
    failed_attempts = session.get("clerk_failed_attempts", 0)

    login_attempt_limit = get_security_setting(
        "login_attempt_limit",
        5
    )

    if failed_attempts >= login_attempt_limit:
        return "Too many failed login attempts ❌"

    conn = None
    cursor = None

    try:

        # =================================================
        # DATABASE CONNECTION
        # =================================================
        conn = get_connection()
        cursor = conn.cursor()

        # =================================================
        # FETCH CLERK USER WITH SCHOOL DETAILS
        # Also fetch failed_login_attempts from DB
        # =================================================
        cursor.execute("""
            SELECT
                u.id,
                u.email,
                u.password,
                u.school_id,
                u.status,
                s.name,
                s.is_active,
                COALESCE(u.failed_login_attempts, 0)
            FROM users u
            JOIN schools s
                ON u.school_id = s.school_id
            WHERE u.email = %s
            AND u.role = 'clerk'
            LIMIT 1
        """, (email,))

        user = cursor.fetchone()

        # =================================================
        # INVALID EMAIL / USER NOT FOUND
        # =================================================
        if not user:

            session["clerk_failed_attempts"] = failed_attempts + 1
            session.modified = True

            return "Invalid Credentials ❌"

        # =================================================
        # MAP DATABASE VALUES
        # =================================================
        user_id = user[0]
        user_email = user[1]
        db_password = user[2]
        school_id = user[3]
        user_status = user[4]
        school_status = user[6]
        db_failed_attempts = int(user[7] or 0)

        # =================================================
        # DATABASE BASED FAILED LOGIN LIMIT
        # Protects account across different browsers/devices
        # =================================================
        if db_failed_attempts >= login_attempt_limit:
            return "Account temporarily locked due to failed attempts ❌"

        # =================================================
        # PASSWORD CHECK
        # If password is wrong, increase failed attempts
        # =================================================
        if not bcrypt.check_password_hash(db_password, password):

            cursor.execute("""
                UPDATE users
                SET
                    failed_login_attempts = COALESCE(failed_login_attempts, 0) + 1,
                    updated_at = NOW()
                WHERE id = %s
            """, (user_id,))

            conn.commit()

            session["clerk_failed_attempts"] = failed_attempts + 1
            session.modified = True

            return "Invalid Credentials ❌"

        # =================================================
        # CHECK SCHOOL STATUS
        # Clerk cannot login if school is disabled
        # =================================================
        if int(school_status or 0) != 1:
            return "School disabled ❌"

        # =================================================
        # CHECK USER STATUS
        # Clerk cannot login if user account is inactive
        # =================================================
        if user_status != "active":
            return "User inactive ❌"

        # =================================================
        # CHECK LATEST SUBSCRIPTION
        # School must have at least one subscription record
        # =================================================
        cursor.execute("""
            SELECT
                status,
                end_date
            FROM subscriptions
            WHERE school_id = %s
            ORDER BY id DESC
            LIMIT 1
        """, (school_id,))

        sub = cursor.fetchone()

        if not sub:

            flash(
                "No subscription found. Please contact administrator.",
                "danger"
            )

            return redirect(
                url_for("renew_subscription")
            )

        # =================================================
        # RESET FAILED ATTEMPTS AFTER SUCCESSFUL PASSWORD
        # Also update last login timestamp
        # =================================================
        cursor.execute("""
            UPDATE users
            SET
                last_login = NOW(),
                updated_at = NOW(),
                failed_login_attempts = 0
            WHERE id = %s
        """, (user_id,))

        conn.commit()

        # =================================================
        # CLEAR OLD CLERK SESSION DATA
        # Keep this because admin and clerk can use same browser
        # =================================================
        session.pop("clerk_logged_in", None)
        session.pop("clerk_user_id", None)
        session.pop("clerk_email", None)
        session.pop("clerk_school_id", None)
        session.pop("clerk_role", None)
        session.pop("clerk_failed_attempts", None)

        # =================================================
        # APPLY GLOBAL SESSION TIMEOUT
        # Value comes from System Settings > Security
        # =================================================
        session_timeout = get_security_setting(
            "session_timeout",
            60
        )

        app.permanent_session_lifetime = timedelta(
            minutes=session_timeout
        )

        session.permanent = True

        # =================================================
        # CREATE NEW CLERK SESSION
        # These values are used across clerk dashboard routes
        # =================================================
        session["clerk_logged_in"] = True
        session["clerk_user_id"] = user_id
        session["clerk_email"] = user_email
        session["clerk_school_id"] = school_id
        session["clerk_role"] = "clerk"
        session.modified = True

        # =================================================
        # CHECK SUBSCRIPTION EXPIRY AFTER SESSION CREATION
        # So clerk can be redirected to renewal page
        # =================================================
        status = sub[0]
        end_date = sub[1]

        if status == "expired" or (end_date and end_date < date.today()):

            flash(
                "Your subscription has expired. Please renew.",
                "danger"
            )

            return redirect(
                url_for("renew_subscription")
            )

        # =================================================
        # LOGIN SUCCESS
        # =================================================
        print("✅ Clerk Login Success:", email)

        return redirect(
            url_for("clerk_dashboard")
        )

    except Exception as e:

        # =================================================
        # ERROR HANDLING
        # =================================================
        print("❌ LOGIN ERROR:", e)

        return "Login failed ❌", 500

    finally:

        # =================================================
        # CLOSE DATABASE RESOURCES
        # =================================================
        if cursor:
            cursor.close()

        if conn:
            conn.close()





# =========================================================
# 🔐 SUPER ADMIN LOGIN
# Handles admin login, failed attempts, password verification,
# account status check, and secure admin session creation.
# =========================================================

@app.route("/superadmin/login", methods=["GET", "POST"])
def superadmin_login():

    # =====================================================
    # SHOW SUPER ADMIN LOGIN PAGE
    # =====================================================
    if request.method == "GET":
        return render_template("auth/superadmin_login.html")

    # =====================================================
    # GET FORM DATA
    # =====================================================
    email = (request.form.get("email") or "").strip().lower()
    password = (request.form.get("password") or "").strip()

    # =====================================================
    # BASIC VALIDATION
    # =====================================================
    if not email or not password:
        print("❌ Admin Login Error: Email and password required")
        return "Email and password required ❌"

    if not is_valid_email(email):
        return "Invalid email format ❌"

    # =====================================================
    # SESSION BASED FAILED LOGIN LIMIT
    # Protects same browser/session from repeated attempts
    # =====================================================
    failed_attempts = session.get("admin_failed_attempts", 0)

    login_attempt_limit = get_security_setting(
        "login_attempt_limit",
        5
    )

    if failed_attempts >= login_attempt_limit:
        return "Too many failed admin login attempts ❌"

    conn = None
    cursor = None

    try:

        # =================================================
        # DATABASE CONNECTION
        # =================================================
        conn = get_connection()
        cursor = conn.cursor()

        # =================================================
        # FETCH ADMIN USER
        # Also fetch failed_login_attempts from DB
        # =================================================
        cursor.execute("""
            SELECT
                id,
                email,
                password,
                role,
                status,
                COALESCE(failed_login_attempts, 0)
            FROM users
            WHERE email = %s
            AND role = 'admin'
            LIMIT 1
        """, (email,))

        user = cursor.fetchone()

        # =================================================
        # INVALID EMAIL / ADMIN NOT FOUND
        # =================================================
        if not user:

            session["admin_failed_attempts"] = failed_attempts + 1
            session.modified = True

            return "Invalid Admin Credentials ❌"

        # =================================================
        # MAP DATABASE VALUES
        # =================================================
        admin_id = user[0]
        admin_email = user[1]
        db_password = user[2]
        role = user[3]
        status = user[4]
        db_failed_attempts = int(user[5] or 0)

        # =================================================
        # DATABASE BASED FAILED LOGIN LIMIT
        # Protects admin account across browsers/devices
        # =================================================
        if db_failed_attempts >= login_attempt_limit:
            return "Admin account temporarily locked due to failed attempts ❌"

        # =================================================
        # ROLE SAFETY CHECK
        # Ensures only admin role can login here
        # =================================================
        if role != "admin":

            session["admin_failed_attempts"] = failed_attempts + 1
            session.modified = True

            return "Invalid Admin Credentials ❌"

        # =================================================
        # ACCOUNT STATUS CHECK
        # Admin cannot login if account is inactive
        # =================================================
        if status != "active":
            return "Admin account inactive ❌"

        # =================================================
        # PASSWORD CHECK
        # If password is wrong, increase failed attempts
        # =================================================
        if not bcrypt.check_password_hash(db_password, password):

            cursor.execute("""
                UPDATE users
                SET
                    failed_login_attempts = COALESCE(failed_login_attempts, 0) + 1,
                    updated_at = NOW()
                WHERE id = %s
            """, (admin_id,))

            conn.commit()

            session["admin_failed_attempts"] = failed_attempts + 1
            session.modified = True

            return "Invalid Admin Credentials ❌"

        # =================================================
        # RESET FAILED ATTEMPTS AFTER SUCCESSFUL LOGIN
        # Also update last login timestamp
        # =================================================
        cursor.execute("""
            UPDATE users
            SET
                last_login = NOW(),
                updated_at = NOW(),
                failed_login_attempts = 0
            WHERE id = %s
        """, (admin_id,))

        conn.commit()

        # =================================================
        # CLEAR ONLY ADMIN SESSION DATA
        # Keep this because admin and clerk can use same browser
        # =================================================
        session.pop("admin_logged_in", None)
        session.pop("admin_user_id", None)
        session.pop("admin_email", None)
        session.pop("admin_role", None)
        session.pop("admin_failed_attempts", None)

        # =================================================
        # APPLY GLOBAL SESSION TIMEOUT
        # Value comes from System Settings > Security
        # =================================================
        session_timeout = get_security_setting(
            "session_timeout",
            60
        )

        app.permanent_session_lifetime = timedelta(
            minutes=session_timeout
        )

        session.permanent = True

        # =================================================
        # CREATE NEW ADMIN SESSION
        # These values are used across superadmin routes
        # =================================================
        session["admin_logged_in"] = True
        session["admin_user_id"] = admin_id
        session["admin_email"] = admin_email
        session["admin_role"] = "admin"
        session.modified = True

        # =================================================
        # LOGIN SUCCESS
        # =================================================
        print("✅ Admin Login Success:", email)

        return redirect(
            url_for("superadmin_dashboard")
        )

    except Exception as e:

        # =================================================
        # ERROR HANDLING
        # =================================================
        print("❌ ADMIN LOGIN ERROR:", e)

        return "Admin login failed ❌", 500

    finally:

        # =================================================
        # CLOSE DATABASE RESOURCES
        # =================================================
        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🚪 LOGOUT ROUTE
# =========================================================
@app.route("/logout", methods=["POST"])
@login_required
def logout():

    role = (
            request.form.get("role", "")
            .strip()
            .lower()
        )

    if role not in ["clerk", "admin"]:
        return "Invalid logout request ❌"

   # ================= CLERK LOGOUT =================
    if role == "clerk":

        session.pop("clerk_logged_in", None)
        session.pop("clerk_user_id", None)
        session.pop("clerk_email", None)
        session.pop("clerk_school_id", None)
        session.pop("clerk_role", None)
        session.pop("clerk_failed_attempts", None)

        return redirect(url_for("login"))


    # ================= ADMIN LOGOUT =================
    elif role == "admin":

        session.pop("admin_logged_in", None)
        session.pop("admin_user_id", None)
        session.pop("admin_email", None)
        session.pop("admin_role", None)
        session.pop("admin_failed_attempts", None)

        return redirect(url_for("superadmin_login"))
 
# =========================================================
# 📊 SUPER ADMIN DASHBOARD
# =========================================================

@app.route("/superadmin/dashboard")
@admin_required
def superadmin_dashboard():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # TOTAL SCHOOLS
        cursor.execute("""
            SELECT COUNT(*)
            FROM schools
        """)
        total_schools = cursor.fetchone()[0] or 0

        # ACTIVE SCHOOLS BASED ON SUBSCRIPTION
        cursor.execute("""
            SELECT COUNT(DISTINCT school_id)
            FROM subscriptions
            WHERE LOWER(status) = 'active'
            AND end_date >= CURDATE()
        """)
        active_schools = cursor.fetchone()[0] or 0

        # EXPIRED SCHOOLS
        cursor.execute("""
            SELECT COUNT(DISTINCT school_id)
            FROM subscriptions
            WHERE LOWER(status) = 'expired'
            OR end_date < CURDATE()
        """)
        expired_schools = cursor.fetchone()[0] or 0

        # TOTAL STUDENTS
        cursor.execute("""
            SELECT COUNT(*)
            FROM students
        """)
        total_students = cursor.fetchone()[0] or 0

        # TOTAL TC
        cursor.execute("""
            SELECT COUNT(*)
            FROM tc
        """)
        total_tc = cursor.fetchone()[0] or 0

        # TOTAL BONAFIDE
        cursor.execute("""
            SELECT COUNT(*)
            FROM bonafide
        """)
        total_bonafide = cursor.fetchone()[0] or 0

        # NEW LEADS
        cursor.execute("""
            SELECT COUNT(*)
            FROM lead_requests
            WHERE status='New'
        """)
        new_leads_count = cursor.fetchone()[0] or 0

        # REVENUE THIS MONTH
        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0)
            FROM payment_logs
            WHERE payment_status='success'
            AND MONTH(created_at) = MONTH(CURDATE())
            AND YEAR(created_at) = YEAR(CURDATE())
        """)
        revenue_this_month = cursor.fetchone()[0] or 0

        # TOTAL REVENUE
        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0)
            FROM payment_logs
            WHERE payment_status='success'
        """)
        total_revenue = cursor.fetchone()[0] or 0

        # UPCOMING RENEWALS NEXT 30 DAYS
        cursor.execute("""
            SELECT COUNT(DISTINCT school_id)
            FROM subscriptions
            WHERE end_date BETWEEN CURDATE()
            AND DATE_ADD(CURDATE(), INTERVAL 30 DAY)
        """)
        upcoming_renewals = cursor.fetchone()[0] or 0

        # SCHOOL WISE STUDENTS CHART
        cursor.execute("""
            SELECT
                s.name,
                COUNT(st.id)
            FROM schools s
            LEFT JOIN students st
                ON st.school_id = s.school_id
            GROUP BY
                s.school_id,
                s.name
            ORDER BY COUNT(st.id) DESC
        """)
        school_chart = cursor.fetchall()

        chart_labels = []
        chart_values = []

        for row in school_chart:
            chart_labels.append(row[0])
            chart_values.append(row[1])

        # REAL REVENUE ANALYTICS - LAST 6 MONTHS
        cursor.execute("""
            SELECT
                DATE_FORMAT(MIN(created_at), '%b') AS month_name,
                YEAR(created_at) AS report_year,
                MONTH(created_at) AS report_month,
                COALESCE(SUM(amount), 0) AS revenue
            FROM payment_logs
            WHERE payment_status='success'
            AND created_at >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
            GROUP BY
                YEAR(created_at),
                MONTH(created_at)
            ORDER BY
                report_year,
                report_month
        """)
        revenue_rows = cursor.fetchall()

        revenue_labels = []
        revenue_values = []

        for row in revenue_rows:
            revenue_labels.append(row[0])
            revenue_values.append(float(row[3] or 0))

        # SUBSCRIPTION ANALYTICS
        subscription_labels = [
            "Active",
            "Expired"
        ]

        subscription_values = [
            active_schools,
            expired_schools
        ]

        # RECENT TC
        cursor.execute("""
            SELECT
                st.name,
                tc.tc_number,
                tc.tc_date
            FROM tc tc
            JOIN students st
                ON tc.student_id = st.id
            ORDER BY tc.id DESC
            LIMIT 5
        """)
        recent_tc = cursor.fetchall()

        # RECENT BONAFIDE
        cursor.execute("""
            SELECT
                st.name,
                b.bonafide_number,
                b.created_at
            FROM bonafide b
            JOIN students st
                ON b.student_id = st.id
            ORDER BY b.id DESC
            LIMIT 5
        """)
        recent_bonafide = cursor.fetchall()

        return render_template(
            "dashboard/superadmin.html",

            total_schools=total_schools,
            active_schools=active_schools,
            expired_schools=expired_schools,

            total_students=total_students,
            total_tc=total_tc,
            total_bonafide=total_bonafide,

            revenue_this_month=revenue_this_month,
            total_revenue=total_revenue,
            upcoming_renewals=upcoming_renewals,
            new_leads_count=new_leads_count,

            chart_labels=chart_labels,
            chart_values=chart_values,

            revenue_labels=revenue_labels,
            revenue_values=revenue_values,

            subscription_labels=subscription_labels,
            subscription_values=subscription_values,

            recent_tc=recent_tc,
            recent_bonafide=recent_bonafide,

            role="admin",
            active_page="dashboard",
            school_name="SchoolSphere Admin"
        )

    except Exception as e:

        print("❌ ADMIN DASHBOARD ERROR:", e)

        return f"Dashboard Error: {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


 # =========================================================
# 🏫 SUPER ADMIN - ALL SCHOOLS
# =========================================================

@app.route("/superadmin/schools")
@admin_required
def superadmin_schools():

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor(
            dictionary=True
        )

        # =====================================
        # FILTERS
        # =====================================

        search = request.args.get(
            "search",
            ""
        ).strip()

        # =====================================
        # PAGINATION
        # =====================================

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 8

        if page < 1:
            page = 1

        offset = (
            page - 1
        ) * per_page

        # =====================================
        # BASE WHERE
        # =====================================

        where_query = """
            WHERE 1=1
        """

        params = []

        if search:

            where_query += """

                AND (

                    s.name LIKE %s
                    OR s.school_code LIKE %s
                    OR s.udise_no LIKE %s
                    OR s.email LIKE %s
                    OR s.phone LIKE %s

                )

            """

            keyword = f"%{search}%"

            params.extend([

                keyword,
                keyword,
                keyword,
                keyword,
                keyword

            ])

        # =====================================
        # TOTAL FILTERED SCHOOLS
        # =====================================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM schools s

            {where_query}

        """, params)

        total_filtered = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            math.ceil(
                total_filtered / per_page
            )
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # =====================================
        # SCHOOLS DATA
        # =====================================

        query = f"""

            SELECT

                s.school_id,
                s.name,
                s.school_code,
                s.udise_no,
                s.address,
                s.phone,
                s.email,
                s.principal_name,
                s.is_active,

                -- CERTIFICATE SETTINGS
                s.tc_prefix,
                s.bonafide_prefix,

                s.auto_numbering,
                s.enable_certificate_labels,

                s.show_tc_logo,
                s.show_tc_watermark,

                s.show_bonafide_logo,
                s.show_bonafide_watermark,

                -- FEATURE MODULES
                s.enable_tc_management,
                s.enable_bonafide_management,
                s.enable_import_export,
                s.enable_attendance,
                s.enable_fee_management,
                s.enable_teacher_management,
                s.enable_results,
                s.enable_timetable,
                s.enable_notice_board,

                COUNT(st.id) AS total_students

            FROM schools s

            LEFT JOIN students st
                ON s.school_id = st.school_id

            {where_query}

            GROUP BY

                s.school_id,
                s.name,
                s.school_code,
                s.udise_no,
                s.address,
                s.phone,
                s.email,
                s.principal_name,
                s.is_active,

                -- CERTIFICATE SETTINGS
                s.tc_prefix,
                s.bonafide_prefix,

                s.auto_numbering,
                s.enable_certificate_labels,

                s.show_tc_logo,
                s.show_tc_watermark,

                s.show_bonafide_logo,
                s.show_bonafide_watermark,

                -- FEATURE MODULES
                s.enable_tc_management,
                s.enable_bonafide_management,
                s.enable_import_export,
                s.enable_attendance,
                s.enable_fee_management,
                s.enable_teacher_management,
                s.enable_results,
                s.enable_timetable,
                s.enable_notice_board

            ORDER BY s.school_id DESC

            LIMIT %s OFFSET %s

        """

        query_params = params.copy()

        query_params.extend([

            per_page,
            offset

        ])

        cursor.execute(
            query,
            query_params
        )

        schools = cursor.fetchall()

        # =====================================
        # TOTAL SCHOOLS
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM schools
        """)

        total_schools = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # ACTIVE SCHOOLS
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM schools
            WHERE is_active = 1
        """)

        active_schools = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # SIDEBAR NEW LEADS COUNT
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # RENDER
        # =====================================

        return render_template(

            "superadmin/schools.html",

            schools=schools,

            total_schools=total_schools,

            active_schools=active_schools,

            search=search,

            page=page,

            total_pages=total_pages,

            total_filtered=total_filtered,

            new_leads_count=new_leads_count,

            role="admin",

            school_name="Admin",

            active_page="schools"

        )

    except Exception as e:

        print(
            "❌ SUPERADMIN SCHOOLS ERROR:",
            e
        )

        return f"Error loading schools ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 💾 SAVE/ADD SCHOOL
# =========================================================
@app.route("/superadmin/save-school", methods=["POST"])
@admin_required
def save_school():

    conn = None
    cursor = None

    try:

        # =====================================================
        # GET FORM DATA
        # =====================================================

        name = (
            request.form.get("name") or ""
        ).strip()

        udise_no = (
            request.form.get("udise_no") or ""
        ).strip()

        address = (
            request.form.get("address") or ""
        ).strip()

        phone = (
            request.form.get("phone") or ""
        ).strip()

        email = (
            request.form.get("email") or ""
        ).strip().lower()

        school_code = (
            request.form.get("school_code") or ""
        ).strip()

        principal_name = (
            request.form.get("principal_name") or ""
        ).strip()

        recognition_no = (
            request.form.get("recognition_no") or ""
        ).strip()

        medium = (
            request.form.get("medium") or ""
        ).strip()

        school_index_no = (
            request.form.get("school_index_no") or ""
        ).strip()

        board_name = (
            request.form.get("board_name") or ""
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if (
            not name
            or not udise_no
            or not phone
            or not email
            or not school_code
        ):
            return "Required fields missing ❌"


        # =====================================================
        # FORMAT VALIDATION
        # =====================================================

        if not re.match(r"^[\w\.-]+@[\w\.-]+\.\w+$", email):
            return "Invalid email format ❌"

        clean_phone = "".join(
                filter(str.isdigit, phone)
            )

        if len(clean_phone) < 10 or len(clean_phone) > 15:
            return "Invalid phone number ❌"

        phone = clean_phone
       
        # =====================================================
        # LENGTH VALIDATION
        # =====================================================

        if len(name) > 150:
            return "School name too long ❌"

        if len(address) > 500:
            return "Address too long ❌"

        if len(phone) > 15:
            return "Invalid phone ❌"

        if len(email) > 150:
            return "Invalid email ❌"

        if len(principal_name) > 150:
            return "Principal name too long ❌"

        if len(school_code) > 20:
            return "School code too long ❌"

        # =====================================================
        # DB CONNECTION
        # =====================================================

        conn = get_connection()
        cursor = conn.cursor()

        # =====================================================
        # DUPLICATE CHECK
        # =====================================================

        cursor.execute("""
            SELECT school_id
            FROM schools
            WHERE
                udise_no = %s
                OR email = %s
                OR school_code = %s
        """, (

            udise_no,
            email,
            school_code

        ))

        existing_school = cursor.fetchone()

        if existing_school:
            return (
                "School with same UDISE, "
                "Email or School Code already exists ❌"
            )

        # =====================================================
        # GET SYSTEM SETTINGS
        # =====================================================

        cursor.execute("""
            SELECT

                -- SUBSCRIPTION
                default_plan_id,
                trial_days,

                -- CERTIFICATE SETTINGS
                tc_prefix,
                bonafide_prefix,
                auto_numbering,
                enable_certificate_labels,

                show_tc_logo,
                show_tc_watermark,

                show_bonafide_logo,
                show_bonafide_watermark,

                -- FEATURE MODULES
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board,

                enable_tc_management,
                enable_bonafide_management

            FROM system_settings
            WHERE id = 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        # =====================================================
        # SUBSCRIPTION
        # =====================================================

        default_plan_id = settings[0]
        try:
            trial_days = int(settings[1] or 7)
        except:
            trial_days = 7

        if trial_days <= 0:
            trial_days = 7

        # =====================================================
        # CERTIFICATE SETTINGS
        # =====================================================

        tc_prefix = settings[2]
        bonafide_prefix = settings[3]

        auto_numbering = settings[4]
        enable_certificate_labels = settings[5]

        show_tc_logo = settings[6]
        show_tc_watermark = settings[7]

        show_bonafide_logo = settings[8]
        show_bonafide_watermark = settings[9]

        # =====================================================
        # FEATURE MODULES
        # =====================================================

        enable_import_export = settings[10]
        enable_attendance = settings[11]
        enable_fee_management = settings[12]
        enable_teacher_management = settings[13]
        enable_results = settings[14]
        enable_timetable = settings[15]
        enable_notice_board = settings[16]

        enable_tc_management = settings[17]
        enable_bonafide_management = settings[18]

        # =====================================================
        # INSERT SCHOOL
        # =====================================================

        cursor.execute("""
            INSERT INTO schools (

                -- BASIC INFO
                name,
                udise_no,
                address,
                phone,
                email,
                school_code,
                principal_name,
                recognition_no,
                medium,
                school_index_no,
                board_name,

                -- CERTIFICATE SETTINGS
                tc_prefix,
                bonafide_prefix,
                auto_numbering,
                enable_certificate_labels,

                show_tc_logo,
                show_tc_watermark,

                show_bonafide_logo,
                show_bonafide_watermark,

                -- FEATURE MODULES
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board,

                enable_tc_management,
                enable_bonafide_management

            )

            VALUES (

                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,

                %s, %s, %s, %s,

                %s, %s,

                %s, %s,

                %s, %s, %s, %s, %s, %s, %s,

                %s, %s

            )
        """, (

            # BASIC INFO
            name,
            udise_no,
            address,
            phone,
            email,
            school_code,
            principal_name,
            recognition_no,
            medium,
            school_index_no,
            board_name,

            # CERTIFICATE SETTINGS
            tc_prefix,
            bonafide_prefix,
            auto_numbering,
            enable_certificate_labels,

            show_tc_logo,
            show_tc_watermark,

            show_bonafide_logo,
            show_bonafide_watermark,

            # FEATURE MODULES
            enable_import_export,
            enable_attendance,
            enable_fee_management,
            enable_teacher_management,
            enable_results,
            enable_timetable,
            enable_notice_board,

            enable_tc_management,
            enable_bonafide_management

        ))

        # =====================================================
        # GET NEW SCHOOL ID
        # =====================================================

        new_school_id = cursor.lastrowid

        print("✅ New School ID:", new_school_id)


        # =====================================================
        # CREATE SCHOOL SEQUENCE
        # =====================================================

        cursor.execute("""
            INSERT INTO school_sequences
            (
                school_id,
                tc_last_number,
                bonafide_last_number,
                admission_last_number
            )
            VALUES
            (
                %s,
                0,
                0,
                0
            )
        """, (new_school_id,))

        # =====================================================
        # GET PLAN INFO
        # =====================================================

        cursor.execute("""
            SELECT
                plan_name,
                monthly_price
            FROM subscription_plans
            WHERE id = %s
        """, (default_plan_id,))

        plan = cursor.fetchone()

        if not plan:
             if conn:
                 conn.rollback()
             return "Default plan not found ❌"

        plan_name = plan[0]
        amount = plan[1]

        start_date = date.today()

        end_date = (
            start_date
            + timedelta(days=trial_days)
        )

        # =====================================================
        # CREATE SUBSCRIPTION
        # =====================================================

        cursor.execute("""
            INSERT INTO subscriptions
            (
                school_id,
                plan_id,
                plan_name,
                start_date,
                end_date,
                status,
                amount
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (

            new_school_id,
            default_plan_id,
            plan_name,
            start_date,
            end_date,
            "active",
            amount

        ))

        # =====================================================
        # SAVE ALL CHANGES
        # =====================================================

        conn.commit()

        # =====================================================
        # SEND WELCOME EMAIL
        # =====================================================

        try:

            subject = (
                "Welcome to SPL ShalaSarthi ERP"
            )

            body = f"""

            <div style="font-family:Arial;padding:20px;">

                <h2 style="color:#10b981;">
                    Welcome to SPL ShalaSarthi ERP            </h2>

                <p>
                    Dear {name},
                </p>

                <p>
                    Your school has been successfully
                    registered in SPL ShalaSarthi ERP.
                </p>

                <hr>

                <p>
                    <b>School Name:</b>
                    {name}
                </p>

                <p>
                    <b>School Code:</b>
                    {school_code}
                </p>

                <p>
                    <b>UDISE:</b>
                    {udise_no}
                </p>

                <p>
                    <b>Plan:</b>
                    {plan_name}
                </p>

                <p>
                    <b>Trial Valid Till:</b>
                    {end_date}
                </p>

                <hr>

                <p>
                    Thank you for choosing
                    <b>SPL ShalaSarthi ERP</b>.
                </p>

            </div>

            """

            send_email(
                email,
                subject,
                body
            )

            print(
                "✅ SCHOOL WELCOME EMAIL SENT"
            )

        except Exception as email_error:

            print(
                "❌ WELCOME EMAIL ERROR:",
                email_error
            )

        # =====================================================
        # REDIRECT
        # =====================================================

        return redirect(
            url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ SAVE SCHOOL ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 💾 UPDATE SCHOOL
# =========================================================

@app.route("/superadmin/update-school", methods=["POST"])
@admin_required
def update_school():

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor()

        # ================= SAFE INPUT =================

        school_id = (
            request.form.get("school_id")
            or ""
        ).strip()

        try:

            school_id = int(school_id)

        except ValueError:

            return "Invalid School ID ❌"

        name = (
            request.form.get("name")
            or ""
        ).strip()

        udise_no = (
            request.form.get("udise_no")
            or ""
        ).strip()

        address = (
            request.form.get("address")
            or ""
        ).strip()

        phone = (
            request.form.get("phone")
            or ""
        ).strip()

        email = (
            request.form.get("email")
            or ""
        ).strip().lower()

        school_code = (
            request.form.get("school_code")
            or ""
        ).strip()

        principal_name = (
            request.form.get("principal_name")
            or ""
        ).strip()

        recognition_no = (
            request.form.get("recognition_no")
            or ""
        ).strip()

        medium = (
            request.form.get("medium")
            or ""
        ).strip()

        school_index_no = (
            request.form.get("school_index_no")
            or ""
        ).strip()

        board_name = (
            request.form.get("board_name")
            or ""
        ).strip()

        # ================= REQUIRED VALIDATION =================

        if (
            not school_id
            or not name
            or not udise_no
            or not phone
            or not email
            or not school_code
        ):

            return "Required fields missing ❌"

        # ================= FORMAT VALIDATION =================

        if not re.match(
            r"^[\w\.-]+@[\w\.-]+\.\w+$",
            email
        ):

            return "Invalid email format ❌"

        clean_phone = "".join(
            filter(str.isdigit, phone)
        )

        if (
            len(clean_phone) < 10
            or len(clean_phone) > 15
        ):

            return "Invalid phone number ❌"

        phone = clean_phone

        # =====================================================
        # LENGTH VALIDATION
        # =====================================================

        if len(name) > 150:
            return "School name too long ❌"

        if len(address) > 500:
            return "Address too long ❌"

        if len(email) > 150:
            return "Invalid email ❌"

        if len(principal_name) > 150:
            return "Principal name too long ❌"

        if len(school_code) > 20:
            return "School code too long ❌"

        if len(udise_no) > 30:
            return "UDISE number too long ❌"

        if len(recognition_no) > 100:
            return "Recognition number too long ❌"

        if len(medium) > 50:
            return "Medium value too long ❌"

        if len(school_index_no) > 100:
            return "School index number too long ❌"

        if len(board_name) > 100:
            return "Board name too long ❌"

        # ================= SCHOOL EXISTS CHECK =================

        cursor.execute("""

            SELECT school_id

            FROM schools

            WHERE school_id = %s

            LIMIT 1

        """, (

            school_id,

        ))

        existing = cursor.fetchone()

        if not existing:

            return "School not found ❌"

        # ================= DUPLICATE CHECK =================

        cursor.execute("""

            SELECT school_id

            FROM schools

            WHERE (

                udise_no = %s

                OR email = %s

                OR school_code = %s

            )

            AND school_id != %s

            LIMIT 1

        """, (

            udise_no,
            email,
            school_code,
            school_id

        ))

        duplicate_school = cursor.fetchone()

        if duplicate_school:

            return (
                "School with same UDISE, "
                "Email or School Code already exists ❌"
            )

        # ================= UPDATE SCHOOL =================

        cursor.execute("""

            UPDATE schools

            SET

                name = %s,

                udise_no = %s,

                address = %s,

                phone = %s,

                email = %s,

                school_code = %s,

                principal_name = %s,

                recognition_no = %s,

                medium = %s,

                school_index_no = %s,

                board_name = %s

            WHERE school_id = %s

        """, (

            name,
            udise_no,
            address,
            phone,
            email,
            school_code,
            principal_name,
            recognition_no,
            medium,
            school_index_no,
            board_name,
            school_id

        ))

        conn.commit()

        return redirect(
            url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ UPDATE SCHOOL ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
            
# =========================================================
# ⚙ UPDATE SCHOOL FEATURES
# =========================================================

@app.route(
    "/superadmin/update-school-features",
    methods=["POST"]
)
@admin_required
def update_school_features():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # SCHOOL ID
        # =========================================

        school_id = (
            request.form.get("school_id")
            or ""
        ).strip()

        if not school_id:
            return "School ID missing ❌"

        try:

            school_id = int(school_id)

        except ValueError:

            return "Invalid School ID ❌"

        # =========================================
        # VERIFY SCHOOL EXISTS
        # =========================================

        cursor.execute("""

            SELECT school_id

            FROM schools

            WHERE school_id = %s

            LIMIT 1

        """, (

            school_id,

        ))

        school = cursor.fetchone()

        if not school:

            return "School not found ❌"

        # =========================================
        # FEATURE VALUES
        # =========================================

        enable_tc_management = (
            request.form.get("enable_tc_management")
            or "Disabled"
        ).strip()

        enable_bonafide_management = (
            request.form.get("enable_bonafide_management")
            or "Disabled"
        ).strip()

        enable_import_export = (
            request.form.get("enable_import_export")
            or "Disabled"
        ).strip()

        enable_attendance = (
            request.form.get("enable_attendance")
            or "Disabled"
        ).strip()

        enable_fee_management = (
            request.form.get("enable_fee_management")
            or "Disabled"
        ).strip()

        enable_teacher_management = (
            request.form.get("enable_teacher_management")
            or "Disabled"
        ).strip()

        enable_results = (
            request.form.get("enable_results")
            or "Disabled"
        ).strip()

        enable_timetable = (
            request.form.get("enable_timetable")
            or "Disabled"
        ).strip()

        enable_notice_board = (
            request.form.get("enable_notice_board")
            or "Disabled"
        ).strip()

        # =========================================
        # VALIDATION
        # =========================================

        valid_values = [
            "Enabled",
            "Disabled"
        ]

        feature_values = [

            enable_tc_management,
            enable_bonafide_management,
            enable_import_export,
            enable_attendance,
            enable_fee_management,
            enable_teacher_management,
            enable_results,
            enable_timetable,
            enable_notice_board

        ]

        for value in feature_values:

            if value not in valid_values:

                return "Invalid feature value ❌"

        # =========================================
        # UPDATE SCHOOL FEATURES
        # =========================================

        cursor.execute("""

            UPDATE schools

            SET

                enable_tc_management = %s,

                enable_bonafide_management = %s,

                enable_import_export = %s,

                enable_attendance = %s,

                enable_fee_management = %s,

                enable_teacher_management = %s,

                enable_results = %s,

                enable_timetable = %s,

                enable_notice_board = %s

            WHERE school_id = %s

        """, (

            enable_tc_management,
            enable_bonafide_management,
            enable_import_export,
            enable_attendance,
            enable_fee_management,
            enable_teacher_management,
            enable_results,
            enable_timetable,
            enable_notice_board,
            school_id

        ))

        conn.commit()

        return redirect(
            url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ UPDATE SCHOOL FEATURES ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📜 UPDATE SCHOOL CERTIFICATE SETTINGS
# =========================================================

@app.route(
    "/superadmin/update-school-certificates",
    methods=["POST"]
)
@admin_required
def update_school_certificates():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # GET SCHOOL ID
        # =========================================

        school_id = (
            request.form.get("school_id")
            or ""
        ).strip()

        if not school_id:
            return "School ID missing ❌"

        try:

            school_id = int(school_id)

        except ValueError:

            return "Invalid School ID ❌"

        # =========================================
        # VERIFY SCHOOL EXISTS
        # =========================================

        cursor.execute("""

            SELECT school_id

            FROM schools

            WHERE school_id = %s

            LIMIT 1

        """, (

            school_id,

        ))

        school = cursor.fetchone()

        if not school:

            return "School not found ❌"

        # =========================================
        # CERTIFICATE SETTINGS
        # =========================================

        tc_prefix = (
            request.form.get("tc_prefix")
            or "TC"
        ).strip().upper()

        bonafide_prefix = (
            request.form.get("bonafide_prefix")
            or "BON"
        ).strip().upper()

        auto_numbering = (
            request.form.get("auto_numbering")
            or "Enabled"
        ).strip()

        enable_certificate_labels = (
            request.form.get("enable_certificate_labels")
            or "Enabled"
        ).strip()

        show_tc_logo = (
            request.form.get("show_tc_logo")
            or "Disabled"
        ).strip()

        show_tc_watermark = (
            request.form.get("show_tc_watermark")
            or "Disabled"
        ).strip()

        show_bonafide_logo = (
            request.form.get("show_bonafide_logo")
            or "Disabled"
        ).strip()

        show_bonafide_watermark = (
            request.form.get("show_bonafide_watermark")
            or "Disabled"
        ).strip()

        # =========================================
        # PREFIX VALIDATION
        # =========================================

        if not tc_prefix:
            return "TC Prefix required ❌"

        if not bonafide_prefix:
            return "Bonafide Prefix required ❌"

        if len(tc_prefix) > 20:
            return "TC Prefix too long ❌"

        if len(bonafide_prefix) > 20:
            return "Bonafide Prefix too long ❌"

        if not re.match(
            r"^[A-Z0-9_-]+$",
            tc_prefix
        ):

            return "Invalid TC Prefix ❌"

        if not re.match(
            r"^[A-Z0-9_-]+$",
            bonafide_prefix
        ):

            return "Invalid Bonafide Prefix ❌"

        # =========================================
        # VALUE VALIDATION
        # =========================================

        valid_values = [
            "Enabled",
            "Disabled"
        ]

        settings_values = [

            auto_numbering,
            enable_certificate_labels,
            show_tc_logo,
            show_tc_watermark,
            show_bonafide_logo,
            show_bonafide_watermark

        ]

        for value in settings_values:

            if value not in valid_values:

                return "Invalid certificate setting ❌"

        # =========================================
        # UPDATE SCHOOL CERTIFICATE SETTINGS
        # =========================================

        cursor.execute("""

            UPDATE schools

            SET

                tc_prefix = %s,

                bonafide_prefix = %s,

                auto_numbering = %s,

                enable_certificate_labels = %s,

                show_tc_logo = %s,

                show_tc_watermark = %s,

                show_bonafide_logo = %s,

                show_bonafide_watermark = %s

            WHERE school_id = %s

        """, (

            tc_prefix,
            bonafide_prefix,
            auto_numbering,
            enable_certificate_labels,
            show_tc_logo,
            show_tc_watermark,
            show_bonafide_logo,
            show_bonafide_watermark,
            school_id

        ))

        conn.commit()

        return redirect(
            url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ SCHOOL CERTIFICATE SETTINGS ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🔄 TOGGLE SCHOOL STATUS (ACTIVE/INACTIVE)
# =========================================================

@app.route("/superadmin/toggle-school-status", methods=["POST"])
@admin_required
def toggle_school_status():

    conn = None
    cursor = None

    try:

        # ================= SAFE INPUT =================

        school_id = (
            request.form.get("school_id")
            or ""
        ).strip()

        if not school_id:
            return "School ID missing ❌"

        try:

            school_id = int(school_id)

        except ValueError:

            return "Invalid School ID ❌"

        # ================= DB =================

        conn = get_connection()

        cursor = conn.cursor()

        # ================= CHECK SCHOOL EXISTS =================

        cursor.execute("""

            SELECT
                school_id,
                is_active

            FROM schools

            WHERE school_id = %s

            LIMIT 1

        """, (

            school_id,

        ))

        school = cursor.fetchone()

        if not school:

            return "School not found ❌"

        # ================= TOGGLE STATUS =================

        cursor.execute("""

            UPDATE schools

            SET is_active =

                CASE

                    WHEN is_active = 1 THEN 0

                    ELSE 1

                END

            WHERE school_id = %s

        """, (

            school_id,

        ))

        if cursor.rowcount != 1:

            conn.rollback()

            return "Status update failed ❌"

        conn.commit()

        return redirect(
            url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ STATUS TOGGLE ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 👨‍🎓 SUPER ADMIN - STUDENTS BY SCHOOL
# =========================================================

@app.route("/superadmin/superadmin_students")
@admin_required
def superadmin_students():

    conn = None
    cursor = None

    try:

        school_id = (
            request.args.get("school_id") or ""
        ).strip()

        search = (
            request.args.get("search") or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 8

        if page < 1:
            page = 1

        offset = (
            page - 1
        ) * per_page

        if not school_id:
            return "School ID missing ❌"

        try:
            school_id = int(school_id)
        except ValueError:
            return "Invalid School ID ❌"

        conn = get_connection()
        cursor = conn.cursor()

        # ================= GET SCHOOL =================
        cursor.execute("""
            SELECT
                school_id,
                name,
                is_active
            FROM schools
            WHERE school_id = %s
            LIMIT 1
        """, (school_id,))

        school = cursor.fetchone()

        if not school:
            return "School not found ❌"

        if str(school[2]) != "1":
            return "School is inactive ❌"

        # ================= WHERE QUERY =================

        where_query = """
            WHERE school_id = %s
        """

        params = [
            school_id
        ]

        if search:

            where_query += """
                AND (
                    name LIKE %s
                    OR admission_no LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword
            ])

        # ================= TOTAL FILTERED STUDENTS =================

        cursor.execute(f"""
            SELECT COUNT(*)
            FROM students
            {where_query}
        """, params)

        total_filtered = cursor.fetchone()[0] or 0

        total_pages = max(
            1,
            math.ceil(total_filtered / per_page)
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # ================= PAGINATED STUDENTS =================

        student_params = params.copy()

        student_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""
            SELECT
                id,
                name,
                `class`,
                admission_no,
                primary_mobile
            FROM students
            {where_query}
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """, student_params)

        students = cursor.fetchall()

        # ================= TOTAL STATS FOR SCHOOL =================

        cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE school_id = %s
        """, (school_id,))

        total_students = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(DISTINCT `class`)
            FROM students
            WHERE school_id = %s
            AND `class` IS NOT NULL
            AND `class` <> ''
        """, (school_id,))

        total_classes = cursor.fetchone()[0] or 0

        return render_template(
            "superadmin/superadmin_students.html",
            students=students,
            total_students=total_students,
            total_classes=total_classes,
            total_filtered=total_filtered,
            page=page,
            total_pages=total_pages,
            school_id=school_id,
            school_name=school[1],
            search=search,
            role="admin",
            active_page="schools"
        )

    except Exception as e:

        print(
            "❌ SUPERADMIN STUDENTS ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📦 GET STUDENT DATA (API FOR MODAL) FOR ADMIN
# =========================================================

@app.route("/get-student/<int:id>")
@admin_required
def get_student(id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""

            SELECT

                id,
                school_id,
                name,
                `class`,
                admission_no,
                primary_mobile,
                father_name,
                mother_name,
                dob,
                nationality,
                previous_school,
                admission_date,
                progress,
                conduct,
                aadhaar

            FROM students

            WHERE id = %s

            LIMIT 1

        """, (id,))

        student = cursor.fetchone()

        if not student:

            return jsonify({
                "error": "Student not found"
            }), 404

        response = {
            "id": student.get("id"),
            "school_id": student.get("school_id"),
            "name": student.get("name") or "",
            "class": student.get("class") or "",
            "admission_no": student.get("admission_no") or "",
            "primary_mobile": student.get("primary_mobile") or "",
            "father_name": student.get("father_name") or "",
            "mother_name": student.get("mother_name") or "",
            "dob": str(student.get("dob") or ""),
            "nationality": student.get("nationality") or "",
            "previous_school": student.get("previous_school") or "",
            "admission_date": str(student.get("admission_date") or ""),
            "progress": student.get("progress") or "",
            "conduct": student.get("conduct") or "",
            "aadhaar": student.get("aadhaar") or ""
        }

        return jsonify(response)

    except Exception as e:

        print(
            "❌ GET STUDENT ERROR:",
            e
        )

        return jsonify({
            "error": "Something went wrong"
        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# ❌ DELETE STUDENT (ADMIN SAFE)
# =========================================================

@app.route("/superadmin/delete_student", methods=["POST"])
@admin_required
def delete_student():

    conn = None
    cursor = None

    try:

        student_id = (
            request.form.get("student_id")
            or ""
        ).strip()

        if not student_id:
            return "Invalid student ID ❌"

        try:

            student_id = int(student_id)

        except ValueError:

            return "Invalid student ID ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ================= CHECK STUDENT =================

        cursor.execute("""

            SELECT
                id,
                school_id,
                name

            FROM students

            WHERE id = %s

            LIMIT 1

        """, (student_id,))

        student = cursor.fetchone()

        if not student:

            return "Student not found ❌"

        # ================= CHECK TC RECORDS =================

        cursor.execute("""

            SELECT 1

            FROM tc

            WHERE student_id = %s

            LIMIT 1

        """, (student_id,))

        if cursor.fetchone():

            return "Cannot delete: TC records exist ❌"

        # ================= CHECK BONAFIDE RECORDS =================

        cursor.execute("""

            SELECT 1

            FROM bonafide

            WHERE student_id = %s

            LIMIT 1

        """, (student_id,))

        if cursor.fetchone():

            return "Cannot delete: Bonafide records exist ❌"

        # ================= DELETE STUDENT =================

        cursor.execute("""

            DELETE FROM students

            WHERE id = %s

        """, (student_id,))

        if cursor.rowcount != 1:

            conn.rollback()

            return "Student deletion failed ❌"

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_schools")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ DELETE STUDENT ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 👨‍🎓 SUPER ADMIN - ALL STUDENTS
# =========================================================

@app.route("/superadmin/all-students")
@admin_required
def superadmin_all_students():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ================= GET PARAMS =================

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 10

        search = (
            request.args.get("search")
            or ""
        ).strip()

        school_id = request.args.get(
            "school_id",
            type=int
        )

        if page < 1:
            page = 1

        offset = (
            page - 1
        ) * per_page

        # ================= WHERE =================

        where = """
            WHERE 1=1
        """

        params = []

        if school_id:

            where += """
                AND st.school_id = %s
            """

            params.append(school_id)

        if search:

            where += """
                AND (
                    st.name LIKE %s
                    OR st.admission_no LIKE %s
                    OR st.primary_mobile LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword
            ])

        # ================= TOTAL FILTERED COUNT =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM students st

            {where}

        """, params)

        total = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # ================= GET STUDENTS =================

        student_params = params.copy()

        student_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                st.id,
                st.name,
                st.`class`,
                st.admission_no,
                st.primary_mobile,

                sc.name AS school_name,
                sc.school_id

            FROM students st

            JOIN schools sc
                ON st.school_id = sc.school_id

            {where}

            ORDER BY st.id DESC

            LIMIT %s OFFSET %s

        """, student_params)

        students = cursor.fetchall()

        # ================= TOTAL SCHOOLS =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM schools
        """)

        total_schools = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= SCHOOL DROPDOWN =================

        cursor.execute("""

            SELECT
                school_id,
                name

            FROM schools

            ORDER BY name

        """)

        schools = cursor.fetchall()

        # ================= SIDEBAR LEAD COUNT =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        return render_template(

            "superadmin/superadmin_all_students.html",

            students=students,
            schools=schools,

            total=total,
            total_schools=total_schools,

            page=page,
            total_pages=total_pages,

            search=search,
            selected_school=school_id,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="all-students"

        )

    except Exception as e:

        print(
            "❌ ALL STUDENTS ERROR:",
            e
        )

        return f"Something went wrong ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🧾 SUPER ADMIN - TC MANAGEMENT
# =========================================================

@app.route("/superadmin/tc-management")
@admin_required
def superadmin_tc_management():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ================= PARAMS =================

        search = (
            request.args.get("search")
            or ""
        ).strip()

        school_filter = (
            request.args.get("school_id")
            or ""
        ).strip()

        class_filter = (
            request.args.get("class")
            or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 5

        if page < 1:
            page = 1

        if school_filter and not school_filter.isdigit():
            school_filter = ""

        if class_filter and not class_filter.isdigit():
            class_filter = ""

        offset = (
            page - 1
        ) * per_page

        # ================= WHERE =================

        where_query = """
            WHERE 1=1
        """

        params = []

        if search:

            where_query += """
                AND (
                    st.name LIKE %s
                    OR st.admission_no LIKE %s
                    OR tc.tc_number LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword
            ])

        if school_filter:

            where_query += """
                AND sc.school_id = %s
            """

            params.append(
                int(school_filter)
            )

        if class_filter:

            where_query += """
                AND st.`class` = %s
            """

            params.append(
                class_filter
            )

        # ================= TOTAL FILTERED COUNT =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM tc

            JOIN students st
                ON tc.student_id = st.id

            JOIN schools sc
                ON tc.school_id = sc.school_id

            {where_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # ================= TC RECORDS =================

        query_params = params.copy()

        query_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                tc.id,
                st.name,
                st.admission_no,
                st.`class`,
                sc.name AS school_name,
                sc.school_id,
                tc.tc_number,
                tc.leaving_date,
                tc.leaving_reason,
                tc.tc_date

            FROM tc

            JOIN students st
                ON tc.student_id = st.id

            JOIN schools sc
                ON tc.school_id = sc.school_id

            {where_query}

            ORDER BY tc.id DESC

            LIMIT %s OFFSET %s

        """, query_params)

        tc_records = cursor.fetchall()

        # ================= SCHOOLS =================

        cursor.execute("""

            SELECT
                school_id,
                name

            FROM schools

            ORDER BY name

        """)

        schools = cursor.fetchall()

        # ================= TOTAL TC =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM tc
        """)

        total_tc = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= TODAY TC =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM tc

            WHERE DATE(tc_date) = CURDATE()

        """)

        today_tc = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= THIS MONTH TC =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM tc

            WHERE MONTH(tc_date) = MONTH(CURDATE())
            AND YEAR(tc_date) = YEAR(CURDATE())

        """)

        month_tc = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= SCHOOLS WITH TC =================

        cursor.execute("""

            SELECT COUNT(DISTINCT school_id) AS total

            FROM tc

        """)

        school_tc_count = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= SIDEBAR LEADS COUNT =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM lead_requests

            WHERE status = 'New'

        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        return render_template(

            "superadmin/superadmin_tc.html",

            active_page="tc-management",

            tc_records=tc_records,
            schools=schools,

            total_tc=total_tc,
            today_tc=today_tc,
            month_tc=month_tc,
            school_tc_count=school_tc_count,

            total_records=total_records,
            page=page,
            total_pages=total_pages,

            search=search,
            school_filter=school_filter,
            class_filter=class_filter,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel"

        )

    except Exception as e:

        print(
            "❌ TC MANAGEMENT ERROR:",
            e
        )

        return f"Something went wrong ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# ❌ DELETE TC (ADMIN SAFE)
# =========================================================

@app.route("/superadmin/delete-tc", methods=["POST"])
@admin_required
def delete_tc():

    conn = None
    cursor = None

    try:

        tc_id = (
            request.form.get("tc_id")
            or ""
        ).strip()

        if not tc_id:
            return "Invalid TC ID ❌"

        try:

            tc_id = int(tc_id)

        except ValueError:

            return "Invalid TC ID ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ================= CHECK TC EXISTS =================

        cursor.execute("""

            SELECT
                id,
                student_id,
                school_id,
                tc_number

            FROM tc

            WHERE id = %s

            LIMIT 1

        """, (tc_id,))

        tc_record = cursor.fetchone()

        if not tc_record:

            return "TC record not found ❌"

        # ================= DELETE TC =================

        cursor.execute("""

            DELETE FROM tc

            WHERE id = %s

        """, (tc_id,))

        if cursor.rowcount != 1:

            conn.rollback()

            return "TC deletion failed ❌"

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_tc_management")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ DELETE TC ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📜 SUPER ADMIN - BONAFIDE MANAGEMENT
# =========================================================

@app.route("/superadmin/bonafide-management")
@admin_required
def superadmin_bonafide_management():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        search = (
            request.args.get("search")
            or ""
        ).strip()

        school_filter = (
            request.args.get("school_id")
            or ""
        ).strip()

        class_filter = (
            request.args.get("class")
            or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 5

        if page < 1:
            page = 1

        if school_filter and not school_filter.isdigit():
            school_filter = ""

        if class_filter and not class_filter.isdigit():
            class_filter = ""

        offset = (
            page - 1
        ) * per_page

        # ================= WHERE =================

        where_query = """
            WHERE 1=1
        """

        params = []

        if search:

            where_query += """
                AND (
                    s.name LIKE %s
                    OR s.admission_no LIKE %s
                    OR b.bonafide_number LIKE %s
                    OR b.purpose LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword,
                keyword
            ])

        if school_filter:

            where_query += """
                AND sc.school_id = %s
            """

            params.append(
                int(school_filter)
            )

        if class_filter:

            where_query += """
                AND s.`class` = %s
            """

            params.append(
                class_filter
            )

        # ================= TOTAL FILTERED COUNT =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM bonafide b

            JOIN students s
                ON b.student_id = s.id

            JOIN schools sc
                ON b.school_id = sc.school_id

            {where_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # ================= BONAFIDE RECORDS =================

        query_params = params.copy()

        query_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                b.id,
                s.name,
                s.admission_no,
                s.`class`,
                sc.name AS school_name,
                sc.school_id,
                b.bonafide_number,
                b.purpose,
                b.date

            FROM bonafide b

            JOIN students s
                ON b.student_id = s.id

            JOIN schools sc
                ON b.school_id = sc.school_id

            {where_query}

            ORDER BY b.id DESC

            LIMIT %s OFFSET %s

        """, query_params)

        bonafides = cursor.fetchall()

        # ================= SCHOOLS =================

        cursor.execute("""

            SELECT
                school_id,
                name

            FROM schools

            ORDER BY name

        """)

        schools = cursor.fetchall()

        # ================= TOTAL BONAFIDE =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM bonafide
        """)

        total_bonafide = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= THIS MONTH =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM bonafide

            WHERE MONTH(date) = MONTH(CURDATE())
            AND YEAR(date) = YEAR(CURDATE())

        """)

        month_bonafide = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= LAST 7 DAYS =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM bonafide

            WHERE date >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)

        """)

        week_bonafide = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= SCHOOLS USING BONAFIDE =================

        cursor.execute("""

            SELECT COUNT(DISTINCT school_id) AS total

            FROM bonafide

        """)

        school_count = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= SIDEBAR LEADS COUNT =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM lead_requests

            WHERE status = 'New'

        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        return render_template(

            "superadmin/superadmin_bonafide.html",

            active_page="bonafide-management",

            bonafides=bonafides,
            schools=schools,

            total_bonafide=total_bonafide,
            month_bonafide=month_bonafide,
            week_bonafide=week_bonafide,
            school_count=school_count,

            total_records=total_records,
            page=page,
            total_pages=total_pages,

            search=search,
            school_filter=school_filter,
            class_filter=class_filter,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel"

        )

    except Exception as e:

        print(
            "❌ BONAFIDE MANAGEMENT ERROR:",
            e
        )

        return f"Something went wrong ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# ❌ DELETE BONAFIDE (ADMIN SAFE)
# =========================================================
@app.route("/superadmin/delete-bonafide", methods=["POST"])
@admin_required
def delete_bonafide():

    conn = None
    cursor = None

    try:

        # ================= GET BONAFIDE ID =================
        bonafide_id = (
            request.form.get("bonafide_id") or ""
        ).strip()

        if not bonafide_id:
            return "Invalid bonafide ID ❌"

        try:
            bonafide_id = int(bonafide_id)
        except ValueError:
            return "Invalid bonafide ID ❌"

        conn = get_connection()
        cursor = conn.cursor()

        # ================= CHECK EXISTS =================
        cursor.execute("""
            SELECT id
            FROM bonafide
            WHERE id = %s
        """, (bonafide_id,))

        bonafide = cursor.fetchone()

        if not bonafide:
            return "Bonafide record not found ❌"

        # ================= DELETE =================
        cursor.execute("""
            DELETE FROM bonafide
            WHERE id = %s
        """, (bonafide_id,))

        if cursor.rowcount == 0:
            return "Bonafide record not found ❌"

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_bonafide_management")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ DELETE BONAFIDE ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 👥 SUPER ADMIN - USERS MANAGEMENT
# =========================================================

@app.route("/superadmin/users")
@admin_required
def superadmin_users():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        search = (
            request.args.get("search") or ""
        ).strip()

        role_filter = (
            request.args.get("role") or ""
        ).strip()

        status_filter = (
            request.args.get("status") or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 5

        if page < 1:
            page = 1

        offset = (page - 1) * per_page

        valid_roles = ["admin", "clerk"]
        valid_status = ["active", "blocked"]

        if role_filter not in valid_roles:
            role_filter = ""

        if status_filter not in valid_status:
            status_filter = ""

        where_query = """
            WHERE 1=1
        """

        params = []

        if search:

            where_query += """
                AND (
                    u.name LIKE %s
                    OR u.email LIKE %s
                    OR u.phone LIKE %s
                    OR s.name LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword,
                keyword
            ])

        if role_filter:

            where_query += """
                AND u.role = %s
            """

            params.append(role_filter)

        if status_filter:

            where_query += """
                AND u.status = %s
            """

            params.append(status_filter)

        # ================= TOTAL FILTERED COUNT =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM users u

            LEFT JOIN schools s
                ON u.school_id = s.school_id

            {where_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (page - 1) * per_page

        # ================= USERS =================

        user_params = params.copy()

        user_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT
                u.id,
                u.name,
                u.email,
                u.phone,
                u.role,
                u.status,
                u.last_login,
                u.created_at,
                u.school_id,
                u.designation,
                u.address,

                COALESCE(
                    s.name,
                    'System Admin'
                ) AS school_name,

                COALESCE(
                    sub.plan_name,
                    'No Plan'
                ) AS subscription_plan

            FROM users u

            LEFT JOIN schools s
                ON u.school_id = s.school_id

            LEFT JOIN subscriptions sub
                ON u.school_id = sub.school_id
                AND sub.status = 'active'

            {where_query}

            ORDER BY u.id DESC

            LIMIT %s OFFSET %s

        """, user_params)

        users = cursor.fetchall()

        # ================= KPI COUNTS =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM users
        """)
        total_users = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM users
            WHERE status = 'active'
        """)
        active_users = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM users
            WHERE role = 'clerk'
        """)
        clerk_users = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM users
            WHERE role = 'admin'
        """)
        admin_users = cursor.fetchone()["total"] or 0

        # ================= SCHOOLS =================

        cursor.execute("""
            SELECT
                school_id,
                name
            FROM schools
            ORDER BY name
        """)

        schools = cursor.fetchall()

        # ================= SIDEBAR LEADS COUNT =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = cursor.fetchone()["total"] or 0

        return render_template(
            "superadmin/superadmin_users.html",

            users=users,
            schools=schools,

            total_users=total_users,
            active_users=active_users,
            clerk_users=clerk_users,
            admin_users=admin_users,

            total_records=total_records,
            page=page,
            total_pages=total_pages,

            search=search,
            role_filter=role_filter,
            status_filter=status_filter,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="users"
        )

    except Exception as e:

        print("❌ USERS MANAGEMENT ERROR:", e)

        return f"Something went wrong ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# ➕ ADD NEW USER
# =========================================================

@app.route(
    "/superadmin/add-user",
    methods=["POST"]
)
@admin_required
def add_user():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # GET FORM DATA
        # =========================================

        name = (
            request.form.get("name") or ""
        ).strip()

        email = (
            request.form.get("email") or ""
        ).strip().lower()

        phone = (
            request.form.get("phone") or ""
        ).strip()

        password = (
            request.form.get("password") or ""
        ).strip()

        role = (
            request.form.get("role") or ""
        ).strip().lower()

        school_id = (
            request.form.get("school_id") or ""
        ).strip()

        designation = (
            request.form.get("designation") or ""
        ).strip()

        address = (
            request.form.get("address") or ""
        ).strip()

        # =========================================
        # REQUIRED VALIDATION
        # =========================================

        if not name:
            return "Name is required ❌"

        if not email:
            return "Email is required ❌"

        if not password:
            return "Password is required ❌"

        if not role:
            return "Role is required ❌"

        # =========================================
        # ROLE VALIDATION
        # =========================================

        valid_roles = [
            "admin",
            "clerk"
        ]

        if role not in valid_roles:
            return "Invalid role ❌"

        if role == "clerk" and not school_id:
            return "School required for clerk ❌"

        # =========================================
        # NAME VALIDATION
        # =========================================

        if len(name) < 3:
            return "Name too short ❌"

        if len(name) > 120:
            return "Name too long ❌"

        # =========================================
        # EMAIL VALIDATION
        # =========================================

        if len(email) > 150:
            return "Email too long ❌"

        if not is_valid_email(email):
            return "Invalid email format ❌"

        # =========================================
        # PHONE VALIDATION
        # =========================================

        if phone:

            if not phone.isdigit():
                return "Phone must contain digits only ❌"

            if len(phone) < 10 or len(phone) > 15:
                return "Invalid phone number ❌"

        # =========================================
        # PASSWORD VALIDATION
        # =========================================

        min_password_length = get_security_setting("password_length", 8)

        if len(password) < min_password_length:
            return f"Password must be at least {min_password_length} characters ❌"

        if len(password) > 128:
            return "Password too long ❌"

        if not re.search(r"[A-Z]", password):
            return "Password must contain uppercase letter ❌"

        if not re.search(r"[a-z]", password):
            return "Password must contain lowercase letter ❌"

        if not re.search(r"\d", password):
            return "Password must contain number ❌"

        # =========================================
        # OPTIONAL FIELD VALIDATION
        # =========================================

        if designation and len(designation) > 100:
            return "Designation too long ❌"

        if address and len(address) > 500:
            return "Address too long ❌"

        # =========================================
        # SCHOOL VALIDATION
        # =========================================

        if school_id:

            try:

                school_id = int(school_id)

            except ValueError:

                return "Invalid school ❌"

            cursor.execute("""

                SELECT
                    school_id,
                    name,
                    is_active

                FROM schools

                WHERE school_id = %s

                LIMIT 1

            """, (school_id,))

            school = cursor.fetchone()

            if not school:
                return "Invalid school ❌"

            if int(school["is_active"] or 0) != 1:
                return "Cannot assign user to inactive school ❌"

        else:

            school_id = None

        # =========================================
        # ADMIN SCHOOL RULE
        # =========================================

        if role == "admin":
            school_id = None

        # =========================================
        # DUPLICATE EMAIL CHECK
        # =========================================

        cursor.execute("""

            SELECT id

            FROM users

            WHERE email = %s

            LIMIT 1

        """, (email,))

        if cursor.fetchone():
            return "Email already exists ❌"

        # =========================================
        # DUPLICATE PHONE CHECK
        # =========================================

        if phone:

            cursor.execute("""

                SELECT id

                FROM users

                WHERE phone = %s

                LIMIT 1

            """, (phone,))

            if cursor.fetchone():
                return "Phone already exists ❌"

        # =========================================
        # DUPLICATE SAME USER IN SAME SCHOOL CHECK
        # =========================================

        cursor.execute("""

            SELECT id

            FROM users

            WHERE name = %s
            AND role = %s
            AND (
                school_id = %s
                OR (
                    school_id IS NULL
                    AND %s IS NULL
                )
            )

            LIMIT 1

        """, (
            name,
            role,
            school_id,
            school_id
        ))

        if cursor.fetchone():
            return "User already exists for this role/school ❌"

        # =========================================
        # HASH PASSWORD
        # =========================================

        hashed_password = bcrypt.generate_password_hash(
            password
        ).decode("utf-8")

        # =========================================
        # INSERT USER
        # =========================================

        cursor.execute("""

            INSERT INTO users
            (
                name,
                email,
                password,
                role,
                school_id,
                created_at,
                status,
                phone,
                address,
                designation,
                last_login
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                NOW(),
                %s,
                %s,
                %s,
                %s,
                NULL
            )

        """, (
            name,
            email,
            hashed_password,
            role,
            school_id,
            "active",
            phone,
            address,
            designation
        ))

        conn.commit()

        print(
            "✅ USER CREATED:",
            email
        )

        return redirect(
            url_for("superadmin_users")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ ADD USER ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
# =========================================================
# ✏️ SUPER ADMIN - EDIT USER
# =========================================================

@app.route("/superadmin/user/edit/<int:user_id>", methods=["POST"])
@admin_required
def superadmin_edit_user(user_id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # GET FORM DATA
        # =========================================

        name = (
            request.form.get("name") or ""
        ).strip()

        email = (
            request.form.get("email") or ""
        ).strip().lower()

        phone = (
            request.form.get("phone") or ""
        ).strip()

        role = (
            request.form.get("role") or ""
        ).strip().lower()

        school_id = (
            request.form.get("school_id") or ""
        ).strip()

        designation = (
            request.form.get("designation") or ""
        ).strip()

        address = (
            request.form.get("address") or ""
        ).strip()

        # =========================================
        # REQUIRED VALIDATION
        # =========================================

        if not name:
            return "Name is required ❌"

        if not email:
            return "Email is required ❌"

        if not role:
            return "Role is required ❌"

        # =========================================
        # ROLE VALIDATION
        # =========================================

        valid_roles = [
            "admin",
            "clerk"
        ]

        if role not in valid_roles:
            return "Invalid role ❌"

        if role == "clerk" and not school_id:
            return "School required for clerk ❌"

        # =========================================
        # NAME VALIDATION
        # =========================================

        if len(name) < 3:
            return "Name too short ❌"

        if len(name) > 120:
            return "Name too long ❌"

        # =========================================
        # EMAIL VALIDATION
        # =========================================

        if len(email) > 150:
            return "Email too long ❌"

        if not is_valid_email(email):
            return "Invalid email format ❌"

        # =========================================
        # PHONE VALIDATION
        # =========================================

        if phone:

            if not phone.isdigit():
                return "Phone must contain digits only ❌"

            if len(phone) < 10 or len(phone) > 15:
                return "Invalid phone number ❌"

        # =========================================
        # OPTIONAL FIELD VALIDATION
        # =========================================

        if designation and len(designation) > 100:
            return "Designation too long ❌"

        if address and len(address) > 500:
            return "Address too long ❌"

        # =========================================
        # USER EXISTS
        # =========================================

        cursor.execute("""

            SELECT
                id,
                role,
                school_id,
                email

            FROM users

            WHERE id = %s

            LIMIT 1

        """, (user_id,))

        user = cursor.fetchone()

        if not user:
            return "User not found ❌"

        # =========================================
        # SCHOOL VALIDATION
        # =========================================

        if school_id:

            try:

                school_id = int(school_id)

            except ValueError:

                return "Invalid school ❌"

            cursor.execute("""

                SELECT
                    school_id,
                    is_active

                FROM schools

                WHERE school_id = %s

                LIMIT 1

            """, (school_id,))

            school = cursor.fetchone()

            if not school:
                return "Invalid school ❌"

            if int(school["is_active"] or 0) != 1:
                return "Cannot assign user to inactive school ❌"

        else:

            school_id = None

        # =========================================
        # ADMIN SCHOOL RULE
        # =========================================

        if role == "admin":
            school_id = None

        # =========================================
        # DUPLICATE EMAIL CHECK
        # =========================================

        cursor.execute("""

            SELECT id

            FROM users

            WHERE email = %s
            AND id != %s

            LIMIT 1

        """, (
            email,
            user_id
        ))

        if cursor.fetchone():
            return "Email already exists ❌"

        # =========================================
        # DUPLICATE PHONE CHECK
        # =========================================

        if phone:

            cursor.execute("""

                SELECT id

                FROM users

                WHERE phone = %s
                AND id != %s

                LIMIT 1

            """, (
                phone,
                user_id
            ))

            if cursor.fetchone():
                return "Phone already exists ❌"

        # =========================================
        # DUPLICATE SAME USER IN SAME SCHOOL CHECK
        # =========================================

        cursor.execute("""

            SELECT id

            FROM users

            WHERE name = %s
            AND role = %s
            AND (
                school_id = %s
                OR (
                    school_id IS NULL
                    AND %s IS NULL
                )
            )
            AND id != %s

            LIMIT 1

        """, (
            name,
            role,
            school_id,
            school_id,
            user_id
        ))

        if cursor.fetchone():
            return "User already exists for this role/school ❌"

        # =========================================
        # UPDATE USER
        # =========================================

        cursor.execute("""

            UPDATE users

            SET
                name = %s,
                email = %s,
                phone = %s,
                role = %s,
                school_id = %s,
                designation = %s,
                address = %s,
                updated_at = NOW()

            WHERE id = %s

        """, (
            name,
            email,
            phone,
            role,
            school_id,
            designation,
            address,
            user_id
        ))

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_users")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ EDIT USER ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
 
# =========================================================
# 🔄 SUPER ADMIN - TOGGLE USER STATUS
# =========================================================

@app.route("/superadmin/user/status", methods=["POST"])
@admin_required
def superadmin_toggle_user_status():

    conn = None
    cursor = None

    try:

        user_id = (
            request.form.get("user_id")
            or ""
        ).strip()

        if not user_id:
            return "Invalid user ID ❌"

        try:

            user_id = int(user_id)

        except ValueError:

            return "Invalid user ID ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # CHECK USER
        # =========================================

        cursor.execute("""

            SELECT
                id,
                role,
                status,
                email

            FROM users

            WHERE id = %s

            LIMIT 1

        """, (user_id,))

        user = cursor.fetchone()

        if not user:
            return "User not found ❌"

        role = user["role"]
        current_status = user["status"]

        if current_status not in [
            "active",
            "blocked"
        ]:
            return "Invalid current user status ❌"

        # =========================================
        # PREVENT SELF BLOCK
        # =========================================

        current_admin_id = session.get(
            "admin_user_id"
        )

        if current_admin_id and int(current_admin_id) == user_id:
            return "You cannot change your own status ❌"

        # =========================================
        # PREVENT BLOCKING LAST ACTIVE ADMIN
        # =========================================

        if (
            role == "admin"
            and current_status == "active"
        ):

            cursor.execute("""

                SELECT COUNT(*) AS total

                FROM users

                WHERE role = 'admin'
                AND status = 'active'

            """)

            active_admins = (
                cursor.fetchone()["total"]
                or 0
            )

            if active_admins <= 1:
                return "Cannot block last active admin ❌"

        # =========================================
        # TOGGLE STATUS
        # =========================================

        new_status = (
            "blocked"
            if current_status == "active"
            else "active"
        )

        cursor.execute("""

            UPDATE users

            SET
                status = %s,
                updated_at = NOW()

            WHERE id = %s

        """, (
            new_status,
            user_id
        ))

        if cursor.rowcount != 1:

            conn.rollback()

            return "Status update failed ❌"

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_users")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ TOGGLE STATUS ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
 
# =========================================================
# ❌ SUPER ADMIN - DELETE USER
# =========================================================

@app.route("/superadmin/user/delete", methods=["POST"])
@admin_required
def superadmin_delete_user():

    conn = None
    cursor = None

    try:

        user_id = (
            request.form.get("user_id")
            or ""
        ).strip()

        if not user_id:
            return "Invalid user ID ❌"

        try:

            user_id = int(user_id)

        except ValueError:

            return "Invalid user ID ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # CHECK USER
        # =========================================

        cursor.execute("""

            SELECT
                id,
                role,
                school_id,
                status,
                email

            FROM users

            WHERE id = %s

            LIMIT 1

        """, (user_id,))

        user = cursor.fetchone()

        if not user:
            return "User not found ❌"

        role = user["role"]

        # =========================================
        # PREVENT SELF DELETE
        # =========================================

        current_admin_id = session.get(
            "admin_user_id"
        )

        if current_admin_id and int(current_admin_id) == user_id:
            return "You cannot delete your own account ❌"

        # =========================================
        # PREVENT ADMIN DELETE
        # =========================================

        if role == "admin":
            return "Admin account cannot be deleted ❌"

        # =========================================
        # DELETE USER
        # =========================================

        cursor.execute("""

            DELETE FROM users

            WHERE id = %s

        """, (user_id,))

        if cursor.rowcount != 1:

            conn.rollback()

            return "User deletion failed ❌"

        conn.commit()

        return redirect(
            request.referrer
            or url_for("superadmin_users")
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ DELETE USER ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
 

# =========================================================
# 📞 SUPER ADMIN - LEADS MANAGEMENT
# =========================================================

@app.route("/superadmin/leads")
@admin_required
def superadmin_leads():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        search = (request.args.get("search") or "").strip()
        status_filter = (request.args.get("status") or "").strip()
        plan_filter = (request.args.get("plan") or "").strip()
        source_filter = (request.args.get("source") or "").strip()

        page = request.args.get("page", 1, type=int)
        per_page = 5

        if page < 1:
            page = 1

        valid_status = [
            "New",
            "Contacted",
            "Demo Scheduled",
            "Converted"
        ]

        if status_filter and status_filter not in valid_status:
            status_filter = ""

        where_query = """
            WHERE 1=1
        """

        params = []

        if search:

            where_query += """
                AND (
                    school_name LIKE %s
                    OR contact_person LIKE %s
                    OR mobile LIKE %s
                    OR email LIKE %s
                    OR selected_plan LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword,
                keyword,
                keyword
            ])

        if status_filter:

            where_query += """
                AND status = %s
            """

            params.append(status_filter)

        if plan_filter:

            where_query += """
                AND selected_plan = %s
            """

            params.append(plan_filter)

        if source_filter:

            where_query += """
                AND lead_source = %s
            """

            params.append(source_filter)

        # ================= TOTAL FILTERED =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM lead_requests

            {where_query}

        """, params)

        total_records = cursor.fetchone()["total"] or 0

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages

        offset = (page - 1) * per_page

        # ================= LEADS =================

        lead_params = params.copy()

        lead_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT
                id,
                lead_source,
                school_name,
                contact_person,
                mobile,
                email,
                student_strength,
                selected_plan,
                message,
                status,
                created_at

            FROM lead_requests

            {where_query}

            ORDER BY id DESC

            LIMIT %s OFFSET %s

        """, lead_params)

        leads = cursor.fetchall()

        # ================= KPI COUNTS =================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
        """)
        total_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)
        new_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Contacted'
        """)
        contacted_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Demo Scheduled'
        """)
        demo_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Converted'
        """)
        converted_leads = cursor.fetchone()["total"] or 0

        # ================= DROPDOWNS =================

        cursor.execute("""
            SELECT DISTINCT selected_plan
            FROM lead_requests
            WHERE selected_plan IS NOT NULL
            AND selected_plan <> ''
            ORDER BY selected_plan
        """)

        available_plans = [
            row["selected_plan"]
            for row in cursor.fetchall()
        ]

        cursor.execute("""
            SELECT DISTINCT lead_source
            FROM lead_requests
            WHERE lead_source IS NOT NULL
            AND lead_source <> ''
            ORDER BY lead_source
        """)

        available_sources = [
            row["lead_source"]
            for row in cursor.fetchall()
        ]

        return render_template(
            "superadmin/leads.html",

            leads=leads,

            total_leads=total_leads,
            new_leads=new_leads,
            contacted_leads=contacted_leads,
            demo_leads=demo_leads,
            converted_leads=converted_leads,

            total_records=total_records,
            page=page,
            total_pages=total_pages,

            search=search,
            status_filter=status_filter,
            plan_filter=plan_filter,
            source_filter=source_filter,

            available_plans=available_plans,
            available_sources=available_sources,

            new_leads_count=new_leads,

            role="admin",
            school_name="Admin Panel",
            active_page="leads"
        )

    except Exception as e:

        print("❌ LEADS PAGE ERROR:", e)

        return f"Unable to load leads ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📞 PUBLIC LEAD SUBMISSION
# USED BY:
# 1. Landing Page Modal
# 2. Demo Page Form
#
# FLOW:
# User Submit
#      ↓
# Save in lead_requests
#      ↓
# Status = New
#      ↓
# Show in Super Admin Lead Panel
# =========================================================

@csrf.exempt
@app.route("/submit-lead", methods=["POST"])
def submit_lead():

    conn = None
    cursor = None

    try:

        data = request.get_json(silent=True)

        if not data:
            return jsonify({
                "success": False,
                "message": "Invalid request"
            }), 400

        lead_source = (
            data.get("lead_source")
            or "Website"
        ).strip()

        school_name = (
            data.get("school_name")
            or ""
        ).strip()

        contact_person = (
            data.get("contact_person")
            or ""
        ).strip()

        mobile = (
            data.get("mobile")
            or ""
        ).strip()

        email = (
            data.get("email")
            or ""
        ).strip().lower()

        student_strength = (
            str(data.get("student_strength") or "")
        ).strip()

        selected_plan = (
            data.get("selected_plan")
            or ""
        ).strip()

        message = (
            data.get("message")
            or ""
        ).strip()

        # ================= VALIDATION =================

        if not school_name:
            return jsonify({
                "success": False,
                "message": "School name required"
            }), 400

        if len(school_name) > 150:
            return jsonify({
                "success": False,
                "message": "School name too long"
            }), 400

        if not contact_person:
            return jsonify({
                "success": False,
                "message": "Contact person required"
            }), 400

        if len(contact_person) > 120:
            return jsonify({
                "success": False,
                "message": "Contact person name too long"
            }), 400

        if not re.fullmatch(r"\d{10}", mobile):
            return jsonify({
                "success": False,
                "message": "Enter valid 10 digit mobile number"
            }), 400

        if email and not is_valid_email(email):
            return jsonify({
                "success": False,
                "message": "Invalid email address"
            }), 400

        if student_strength and not student_strength.isdigit():
            return jsonify({
                "success": False,
                "message": "Student strength must be numeric"
            }), 400

        if message and len(message) > 1000:
            return jsonify({
                "success": False,
                "message": "Message too long"
            }), 400

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # ================= DUPLICATE CHECK =================

        cursor.execute("""
            SELECT id
            FROM lead_requests
            WHERE mobile = %s
            LIMIT 1
        """, (mobile,))

        if cursor.fetchone():
            return jsonify({
                "success": False,
                "message": "Lead already exists"
            }), 409

        # ================= INSERT =================

        cursor.execute("""
            INSERT INTO lead_requests
            (
                lead_source,
                school_name,
                contact_person,
                mobile,
                email,
                student_strength,
                selected_plan,
                message,
                status
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                'New'
            )
        """, (
            lead_source,
            school_name,
            contact_person,
            mobile,
            email,
            student_strength,
            selected_plan,
            message
        ))

        conn.commit()

        return jsonify({
            "success": True,
            "message": "Lead saved successfully"
        }), 201

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ LEAD SAVE ERROR:", e)

        return jsonify({
            "success": False,
            "message": "Something went wrong"
        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()



# =========================================================
# 📞 SUPER ADMIN - MARK LEAD AS CONTACTED
#
# PURPOSE:
# Lead has been contacted by phone / WhatsApp.
#
# FLOW:
# New
#   ↓
# Contacted
#
# USED BY:
# Super Admin Lead Management
# =========================================================

@csrf.exempt
@app.route(
    "/superadmin/lead/contacted/<int:lead_id>",
    methods=["POST"]
)
@admin_required
def mark_lead_contacted(lead_id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""

            UPDATE lead_requests

            SET status = 'Contacted'

            WHERE id = %s

        """, (lead_id,))

        conn.commit()

        return jsonify({

            "success": True,

            "message":
            "Lead marked as Contacted"

        })

    except Exception as e:

        print(
            "❌ CONTACTED STATUS ERROR:",
            e
        )

        return jsonify({

            "success": False,

            "message":
            "Failed to update lead"

        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📅 SUPER ADMIN - SCHEDULE DEMO
#
# PURPOSE:
# Demo date/time has been confirmed with school.
#
# FLOW:
# Contacted
#    ↓
# Demo Scheduled
#
# USED BY:
# Super Admin Lead Management
# =========================================================

@csrf.exempt
@app.route(
    "/superadmin/lead/demo/<int:lead_id>",
    methods=["POST"]
)
@admin_required
def schedule_demo(lead_id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""

            UPDATE lead_requests

            SET status = 'Demo Scheduled'

            WHERE id = %s

        """, (lead_id,))

        conn.commit()

        return jsonify({

            "success": True,

            "message":
            "Demo Scheduled Successfully"

        })

    except Exception as e:

        print(
            "❌ DEMO STATUS ERROR:",
            e
        )

        return jsonify({

            "success": False,

            "message":
            "Failed to schedule demo"

        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 # =========================================================
# 🏫 SUPER ADMIN - CONVERT LEAD
#
# PURPOSE:
# Lead purchased ERP
#
# FLOW:
# Demo Scheduled
#      ↓
# Converted
#
# USED BY:
# Lead Management
# =========================================================

@csrf.exempt
@app.route(
    "/superadmin/lead/converted/<int:lead_id>",
    methods=["POST"]
)
@admin_required
def convert_lead(lead_id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""

            UPDATE lead_requests

            SET status='Converted'

            WHERE id=%s

        """, (lead_id,))

        conn.commit()

        return jsonify({

            "success": True,

            "message":
            "Lead Converted Successfully"

        })

    except Exception as e:

        print(
            "❌ CONVERT ERROR:",
            e
        )

        return jsonify({

            "success": False,

            "message":
            "Failed to convert lead"

        }), 500

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()



# =========================================================
# 💳 SUPER ADMIN - SUBSCRIPTION MANAGEMENT
# =========================================================

@app.route("/superadmin/subscriptions")
@admin_required
def superadmin_subscriptions():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =====================================
        # FILTERS
        # =====================================

        search = (
            request.args.get("search")
            or ""
        ).strip()

        status_filter = (
            request.args.get("status")
            or ""
        ).strip()

        plan_filter = (
            request.args.get("plan")
            or ""
        ).strip()

        expiry_filter = (
            request.args.get("expiry")
            or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 5

        if page < 1:
            page = 1

        # =====================================
        # VALID FILTERS
        # =====================================

        valid_status = [
            "active",
            "expired",
            "Active",
            "Expired"
        ]

        valid_expiry = [
            "",
            "expiring_soon",
            "expired"
        ]

        if status_filter and status_filter not in valid_status:
            status_filter = ""

        if expiry_filter not in valid_expiry:
            expiry_filter = ""

        # =====================================
        # WHERE QUERY
        # =====================================

        where_query = """
            WHERE 1=1
        """

        params = []

        # SEARCH SCHOOL / PLAN

        if search:

            where_query += """
                AND (
                    s.name LIKE %s
                    OR sub.plan_name LIKE %s
                )
            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword
            ])

        # STATUS FILTER

        if status_filter:

            where_query += """
                AND LOWER(sub.status) = LOWER(%s)
            """

            params.append(status_filter)

        # PLAN FILTER

        if plan_filter:

            where_query += """
                AND sub.plan_name = %s
            """

            params.append(plan_filter)

        # EXPIRY FILTER

        if expiry_filter == "expiring_soon":

            where_query += """
                AND sub.end_date BETWEEN
                CURDATE()
                AND DATE_ADD(
                    CURDATE(),
                    INTERVAL 30 DAY
                )
            """

        elif expiry_filter == "expired":

            where_query += """
                AND sub.end_date < CURDATE()
            """

        # =====================================
        # TOTAL FILTERED COUNT
        # =====================================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            FROM subscriptions sub

            LEFT JOIN schools s
                ON sub.school_id = s.school_id

            {where_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages

        offset = (
            page - 1
        ) * per_page

        # =====================================
        # MAIN SUBSCRIPTION QUERY
        # =====================================

        query_params = params.copy()

        query_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                sub.id,
                sub.school_id,

                COALESCE(
                    s.name,
                    'Unknown School'
                ) AS school_name,

                sub.plan_name,
                sub.start_date,
                sub.end_date,
                sub.amount,
                sub.status,
                sub.created_at,

                DATEDIFF(
                    sub.end_date,
                    CURDATE()
                ) AS days_remaining

            FROM subscriptions sub

            LEFT JOIN schools s
                ON sub.school_id = s.school_id

            {where_query}

            ORDER BY sub.id DESC

            LIMIT %s OFFSET %s

        """, query_params)

        subscriptions = cursor.fetchall()

        # =====================================
        # KPI COUNTS
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM subscriptions
        """)

        total_subscriptions = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM subscriptions
            WHERE LOWER(status) = 'active'
        """)

        active_subscriptions = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM subscriptions
            WHERE LOWER(status) = 'expired'
            OR end_date < CURDATE()
        """)

        expired_subscriptions = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM subscriptions

            WHERE end_date BETWEEN
            CURDATE()
            AND DATE_ADD(
                CURDATE(),
                INTERVAL 30 DAY
            )

        """)

        renewal_due = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # TOTAL REVENUE
        # =====================================

        cursor.execute("""

            SELECT

                COALESCE(
                    SUM(amount),
                    0
                ) AS revenue

            FROM payment_logs

            WHERE payment_status = 'success'

        """)

        total_revenue = (
            cursor.fetchone()["revenue"]
            or 0
        )

        # =====================================
        # PLAN DROPDOWN
        # =====================================

        cursor.execute("""

            SELECT DISTINCT
                plan_name

            FROM subscriptions

            WHERE plan_name IS NOT NULL
            AND plan_name <> ''

            ORDER BY plan_name

        """)

        plans = cursor.fetchall()

        # =====================================
        # SIDEBAR LEADS COUNT
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # RENDER
        # =====================================

        return render_template(

            "superadmin/subscriptions_admin.html",

            subscriptions=subscriptions,

            total_subscriptions=total_subscriptions,
            active_subscriptions=active_subscriptions,
            expired_subscriptions=expired_subscriptions,
            renewal_due=renewal_due,
            total_revenue=total_revenue,

            plans=plans,

            total_records=total_records,
            page=page,
            total_pages=total_pages,

            search=search,
            status_filter=status_filter,
            plan_filter=plan_filter,
            expiry_filter=expiry_filter,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="subscriptions"

        )

    except Exception as e:

        print(
            "SUBSCRIPTION ERROR:",
            e
        )

        return f"Unable to load subscriptions ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🔔 SUPER ADMIN - RENEWAL CENTER
# =========================================================

@app.route("/superadmin/renewals")
@admin_required
def superadmin_renewals():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        search = (
            request.args.get("search") or ""
        ).strip()

        status_filter = (
            request.args.get("status") or ""
        ).strip()

        plan_filter = (
            request.args.get("plan") or ""
        ).strip()

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 5

        if page < 1:
            page = 1

        valid_status = [
            "",
            "expired",
            "today",
            "7",
            "30"
        ]

        if status_filter not in valid_status:
            status_filter = ""

        # ================= BASE QUERY =================

        base_query = """

            FROM subscriptions sub

            LEFT JOIN schools s
                ON sub.school_id = s.school_id

            WHERE 1=1

        """

        params = []

        # ================= SEARCH =================

        if search:

            base_query += """

                AND (
                    s.name LIKE %s
                    OR COALESCE(s.email, '') LIKE %s
                    OR COALESCE(s.phone, '') LIKE %s
                    OR sub.plan_name LIKE %s
                )

            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword,
                keyword
            ])

        # ================= PLAN FILTER =================

        if plan_filter:

            base_query += """
                AND sub.plan_name = %s
            """

            params.append(plan_filter)

        # ================= STATUS FILTER =================

        if status_filter == "expired":

            base_query += """
                AND DATEDIFF(sub.end_date, CURDATE()) < 0
            """

        elif status_filter == "today":

            base_query += """
                AND DATEDIFF(sub.end_date, CURDATE()) = 0
            """

        elif status_filter == "7":

            base_query += """
                AND DATEDIFF(sub.end_date, CURDATE()) BETWEEN 1 AND 7
            """

        elif status_filter == "30":

            base_query += """
                AND DATEDIFF(sub.end_date, CURDATE()) BETWEEN 1 AND 30
            """

        # ================= TOTAL FILTERED RECORDS =================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            {base_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages

        offset = (
            page - 1
        ) * per_page

        # ================= MAIN DATA =================

        query_params = params.copy()

        query_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                sub.id,
                sub.school_id,
                sub.plan_name,
                sub.start_date,
                sub.end_date,
                sub.amount,
                sub.status,

                COALESCE(s.name, 'Unknown School') AS school_name,
                s.email AS school_email,
                s.phone AS school_phone,

                DATEDIFF(
                    sub.end_date,
                    CURDATE()
                ) AS days_left,

                CASE

                    WHEN DATEDIFF(sub.end_date, CURDATE()) < 0
                    THEN 1

                    WHEN DATEDIFF(sub.end_date, CURDATE()) = 0
                    THEN 2

                    WHEN DATEDIFF(sub.end_date, CURDATE()) BETWEEN 1 AND 7
                    THEN 3

                    WHEN DATEDIFF(sub.end_date, CURDATE()) BETWEEN 8 AND 30
                    THEN 4

                    ELSE 5

                END AS priority_order

            {base_query}

            ORDER BY
                priority_order ASC,
                sub.end_date ASC,
                sub.id DESC

            LIMIT %s OFFSET %s

        """, query_params)

        renewals = cursor.fetchall()

        # ================= FORMAT RENEWALS =================

        for item in renewals:

            days_left = item.get("days_left")

            if days_left is None:

                item["days_left"] = -999
                item["alert_type"] = "expired"
                item["alert_label"] = "No Expiry Date"

            elif days_left < 0:

                item["alert_type"] = "expired"
                item["alert_label"] = "Expired"

            elif days_left == 0:

                item["alert_type"] = "today"
                item["alert_label"] = "Expires Today"

            elif days_left <= 7:

                item["alert_type"] = "warning"
                item["alert_label"] = "Urgent Renewal"

            elif days_left <= 30:

                item["alert_type"] = "safe"
                item["alert_label"] = "Renewal Due"

            else:

                item["alert_type"] = "safe"
                item["alert_label"] = "Active"

            phone = (
                item.get("school_phone")
                or ""
            )

            phone = "".join(
                filter(str.isdigit, phone)
            )

            if len(phone) == 10:
                phone = "91" + phone

            item["school_phone"] = phone

            message = (
                f"Hello {item.get('school_name')}, "
                f"your {item.get('plan_name')} subscription "
                f"expires on {item.get('end_date')}. "
                f"Please renew your SchoolSphere ERP subscription."
            )

            item["whatsapp_message"] = quote_plus(message)

        # ================= KPI COUNTS =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM subscriptions

            WHERE DATEDIFF(end_date, CURDATE()) < 0

        """)

        expired_count = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM subscriptions

            WHERE DATEDIFF(end_date, CURDATE()) = 0

        """)

        today_count = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM subscriptions

            WHERE DATEDIFF(end_date, CURDATE()) BETWEEN 1 AND 7

        """)

        week_count = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM subscriptions

            WHERE DATEDIFF(end_date, CURDATE()) BETWEEN 1 AND 30

        """)

        month_count = (
            cursor.fetchone()["total"]
            or 0
        )

        # ================= PLAN DROPDOWN =================

        cursor.execute("""

            SELECT DISTINCT
                plan_name

            FROM subscriptions

            WHERE plan_name IS NOT NULL
            AND plan_name <> ''

            ORDER BY plan_name

        """)

        plans = cursor.fetchall()

        # ================= SIDEBAR LEADS COUNT =================

        cursor.execute("""

            SELECT COUNT(*) AS total

            FROM lead_requests

            WHERE status = 'New'

        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        return render_template(

            "superadmin/renewals_admin.html",

            renewals=renewals,

            expired_count=expired_count,
            today_count=today_count,
            week_count=week_count,
            month_count=month_count,

            plans=plans,

            search=search,
            status_filter=status_filter,
            plan_filter=plan_filter,

            page=page,
            total_pages=total_pages,
            total_records=total_records,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="renewals"

        )

    except Exception as e:

        print(
            "❌ RENEWAL CENTER ERROR:",
            e
        )

        return f"Unable to load renewal center ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 📧 SUPER ADMIN - SEND RENEWAL EMAIL
# =========================================================

@app.route(
    "/superadmin/renewals/email/<int:subscription_id>",
    methods=["POST"]
)
@admin_required
def send_renewal_email(subscription_id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =====================================
        # KEEP CURRENT FILTER URL
        # =====================================

        back_url = (
            request.referrer
            or url_for("superadmin_renewals")
        )

        # =====================================
        # GET SUBSCRIPTION + SCHOOL
        # =====================================

        cursor.execute("""

            SELECT

                sub.id,
                sub.school_id,
                sub.plan_name,
                sub.start_date,
                sub.end_date,
                sub.amount,
                sub.status,

                COALESCE(
                    s.name,
                    'School'
                ) AS school_name,

                s.email AS school_email

            FROM subscriptions sub

            LEFT JOIN schools s
                ON sub.school_id = s.school_id

            WHERE sub.id = %s

            LIMIT 1

        """, (
            subscription_id,
        ))

        sub = cursor.fetchone()

        if not sub:

            flash(
                "Subscription not found ❌",
                "danger"
            )

            return redirect(back_url)

        school_email = (
            sub.get("school_email")
            or ""
        ).strip().lower()

        if not school_email:

            flash(
                "School email not found ❌",
                "danger"
            )

            return redirect(back_url)

        if not is_valid_email(school_email):

            flash(
                "Invalid school email ❌",
                "danger"
            )

            return redirect(back_url)

        # =====================================
        # EMAIL CONTENT
        # =====================================

        subject = (
            "Subscription Renewal Reminder - "
            "SchoolSphere ERP"
        )

        body = f"""

        <div style="
            font-family:Arial, sans-serif;
            padding:24px;
            line-height:1.7;
            color:#0f172a;
        ">

            <h2 style="
                color:#0EA5A4;
                margin-bottom:16px;
            ">
                Subscription Renewal Reminder
            </h2>

            <p>
                Dear <b>{sub["school_name"]}</b>,
            </p>

            <p>
                This is a reminder that your
                <b>{sub["plan_name"]}</b>
                subscription is due for renewal.
            </p>

            <div style="
                background:#f8fafc;
                border:1px solid #e2e8f0;
                border-radius:12px;
                padding:16px;
                margin:18px 0;
            ">

                <p style="margin:0;">
                    <b>Start Date:</b> {sub["start_date"]}<br>
                    <b>Expiry Date:</b> {sub["end_date"]}<br>
                    <b>Amount:</b> ₹{sub["amount"]}
                </p>

            </div>

            <p>
                Please renew your SchoolSphere ERP subscription
                to avoid service interruption.
            </p>

            <p>
                For assistance, contact the SchoolSphere ERP admin team.
            </p>

            <hr style="
                border:none;
                border-top:1px solid #e2e8f0;
                margin:24px 0;
            ">

            <p>
                Regards,<br>
                <b>SchoolSphere ERP Team</b>
            </p>

        </div>

        """

        # =====================================
        # SEND EMAIL
        # =====================================

        email_sent = send_email(
            school_email,
            subject,
            body
        )

        if email_sent is True:

            # =====================================
            # LOG EMAIL
            # =====================================

            cursor.execute("""

                INSERT INTO subscription_email_logs
                (
                    school_id,
                    subscription_id,
                    email_type,
                    sent_at
                )
                VALUES
                (
                    %s,
                    %s,
                    %s,
                    NOW()
                )

            """, (
                sub["school_id"],
                sub["id"],
                "admin_renewal_reminder"
            ))

            conn.commit()

            flash(
                "Renewal reminder email sent ✅",
                "success"
            )

        else:

            conn.rollback()

            flash(
                "Email sending failed ❌",
                "danger"
            )

        return redirect(back_url)

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ RENEWAL EMAIL ERROR:",
            e
        )

        flash(
            "Renewal email failed ❌",
            "danger"
        )

        return redirect(
            request.referrer
            or url_for("superadmin_renewals")
        )

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 💰 SUPER ADMIN - PAYMENT MANAGEMENT
# =========================================================

@app.route("/superadmin/payments")
@admin_required
def superadmin_payments():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =====================================
        # PAGINATION
        # =====================================

        page = request.args.get(
            "page",
            1,
            type=int
        )

        per_page = 10

        if page < 1:
            page = 1

        # =====================================
        # FILTERS
        # =====================================

        search = (
            request.args.get("search")
            or ""
        ).strip()

        status_filter = (
            request.args.get("status")
            or ""
        ).strip().lower()

        gateway_filter = (
            request.args.get("gateway")
            or ""
        ).strip()

        valid_status = [
            "",
            "success",
            "failed",
            "pending"
        ]

        if status_filter not in valid_status:
            status_filter = ""

        # =====================================
        # BASE QUERY
        # =====================================

        base_query = """

            FROM payment_logs pl

            LEFT JOIN subscription_plans sp
                ON pl.plan_id = sp.id

            LEFT JOIN schools s
                ON pl.school_id = s.school_id

            WHERE 1=1

        """

        params = []

        # =====================================
        # SEARCH
        # =====================================

        if search:

            base_query += """

                AND (

                    COALESCE(s.name, '') LIKE %s
                    OR COALESCE(pl.invoice_number, '') LIKE %s
                    OR COALESCE(pl.payment_id, '') LIKE %s
                    OR COALESCE(pl.order_id, '') LIKE %s
                    OR COALESCE(sp.plan_name, '') LIKE %s

                )

            """

            keyword = f"%{search}%"

            params.extend([
                keyword,
                keyword,
                keyword,
                keyword,
                keyword
            ])

        # =====================================
        # STATUS FILTER
        # =====================================

        if status_filter:

            base_query += """

                AND LOWER(pl.payment_status) = %s

            """

            params.append(status_filter)

        # =====================================
        # GATEWAY FILTER
        # =====================================

        if gateway_filter:

            base_query += """

                AND pl.payment_gateway = %s

            """

            params.append(gateway_filter)

        # =====================================
        # TOTAL RECORDS
        # =====================================

        cursor.execute(f"""

            SELECT COUNT(*) AS total

            {base_query}

        """, params)

        total_records = (
            cursor.fetchone()["total"]
            or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages

        offset = (
            page - 1
        ) * per_page

        # =====================================
        # MAIN QUERY
        # =====================================

        query_params = params.copy()

        query_params.extend([
            per_page,
            offset
        ])

        cursor.execute(f"""

            SELECT

                pl.id,
                pl.invoice_number,
                pl.amount,
                pl.payment_status,
                pl.payment_id,
                pl.order_id,
                pl.payment_gateway,
                pl.transaction_type,
                pl.created_at,

                COALESCE(
                    sp.plan_name,
                    'No Plan'
                ) AS plan_name,

                COALESCE(
                    s.name,
                    'Unknown School'
                ) AS school_name

            {base_query}

            ORDER BY pl.id DESC

            LIMIT %s OFFSET %s

        """, query_params)

        payments = cursor.fetchall()

        # =====================================
        # KPI CARDS
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM payment_logs
        """)
        total_transactions = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM payment_logs
            WHERE LOWER(payment_status) = 'success'
        """)
        successful_payments = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM payment_logs
            WHERE LOWER(payment_status) = 'failed'
        """)
        failed_payments = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM payment_logs
            WHERE LOWER(payment_status) = 'pending'
        """)
        pending_payments = cursor.fetchone()["total"] or 0

        cursor.execute("""

            SELECT
                COALESCE(
                    SUM(amount),
                    0
                ) AS revenue

            FROM payment_logs

            WHERE LOWER(payment_status) = 'success'

        """)
        total_revenue = cursor.fetchone()["revenue"] or 0

        # =====================================
        # GATEWAYS DROPDOWN
        # =====================================

        cursor.execute("""

            SELECT DISTINCT
                payment_gateway

            FROM payment_logs

            WHERE payment_gateway IS NOT NULL
            AND payment_gateway <> ''

            ORDER BY payment_gateway

        """)

        gateways = cursor.fetchall()

        # =====================================
        # SIDEBAR LEADS COUNT
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = cursor.fetchone()["total"] or 0

        # =====================================
        # RENDER
        # =====================================

        return render_template(

            "superadmin/payments_admin.html",

            payments=payments,

            total_transactions=total_transactions,
            successful_payments=successful_payments,
            failed_payments=failed_payments,
            pending_payments=pending_payments,
            total_revenue=total_revenue,

            gateways=gateways,

            search=search,
            status_filter=status_filter,
            gateway_filter=gateway_filter,

            page=page,
            total_pages=total_pages,
            total_records=total_records,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="payments"

        )

    except Exception as e:

        print(
            "PAYMENTS ERROR:",
            e
        )

        return f"Unable to load payments ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

            # =========================================================
# 📤 SUPER ADMIN - REUSABLE EXCEL EXPORT
# =========================================================

@app.route("/superadmin/export/excel/<export_type>")
@admin_required
def superadmin_export_excel(export_type):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        search = (request.args.get("search") or "").strip()
        status = (request.args.get("status") or "").strip()
        plan = (request.args.get("plan") or "").strip()
        school_id = (request.args.get("school_id") or "").strip()
        gateway = (request.args.get("gateway") or "").strip()
        from_date = (request.args.get("from_date") or "").strip()
        to_date = (request.args.get("to_date") or "").strip()

        query = ""
        params = []
        filename = f"{export_type}_export.xlsx"

        
        # ================= PAYMENTS =================
        if export_type == "payments":

            query = """
                SELECT
                    pl.invoice_number,
                    COALESCE(s.name, 'Unknown School') AS school_name,
                    COALESCE(sp.plan_name, 'No Plan') AS plan_name,
                    pl.amount,
                    pl.payment_gateway,
                    pl.payment_id,
                    pl.order_id,
                    pl.payment_status,
                    pl.transaction_type,
                    pl.created_at
                FROM payment_logs pl
                LEFT JOIN subscription_plans sp
                    ON pl.plan_id = sp.id
                LEFT JOIN schools s
                    ON pl.school_id = s.school_id
                WHERE 1=1
            """

            if search:
                query += """
                    AND (
                        s.name LIKE %s
                        OR pl.invoice_number LIKE %s
                        OR pl.payment_id LIKE %s
                        OR pl.order_id LIKE %s
                    )
                """
                keyword = f"%{search}%"
                params.extend([keyword, keyword, keyword, keyword])

            if status:
                query += " AND LOWER(pl.payment_status) = %s"
                params.append(status.lower())

            if gateway:
                query += " AND pl.payment_gateway = %s"
                params.append(gateway)

            query += " ORDER BY pl.id DESC"
            filename = "admin_payments.xlsx"

        # ================= SUBSCRIPTIONS =================
        elif export_type == "subscriptions":

            query = """
                SELECT
                    s.name AS school_name,
                    sub.plan_name,
                    sub.start_date,
                    sub.end_date,
                    sub.amount,
                    sub.status,
                    sub.created_at
                FROM subscriptions sub
                LEFT JOIN schools s
                    ON sub.school_id = s.school_id
                WHERE 1=1
            """

            if search:
                query += " AND s.name LIKE %s"
                params.append(f"%{search}%")

            if status:
                query += " AND LOWER(sub.status) = %s"
                params.append(status.lower())

            if plan:
                query += " AND sub.plan_name = %s"
                params.append(plan)

            query += " ORDER BY sub.id DESC"
            filename = "admin_subscriptions.xlsx"

        # ================= RENEWALS =================
        elif export_type == "renewals":

            query = """
                SELECT
                    s.name AS school_name,
                    s.email AS school_email,
                    s.phone AS school_phone,
                    sub.plan_name,
                    sub.start_date,
                    sub.end_date,
                    sub.amount,
                    sub.status,
                    DATEDIFF(sub.end_date, CURDATE()) AS days_left
                FROM subscriptions sub
                LEFT JOIN schools s
                    ON sub.school_id = s.school_id
                WHERE 1=1
            """

            if search:
                query += """
                    AND (
                        s.name LIKE %s
                        OR s.email LIKE %s
                        OR s.phone LIKE %s
                    )
                """
                keyword = f"%{search}%"
                params.extend([keyword, keyword, keyword])

            if plan:
                query += " AND sub.plan_name = %s"
                params.append(plan)

            if status == "expired":
                query += " AND DATEDIFF(sub.end_date, CURDATE()) < 0"
            elif status == "today":
                query += " AND DATEDIFF(sub.end_date, CURDATE()) = 0"
            elif status == "7":
                query += " AND DATEDIFF(sub.end_date, CURDATE()) BETWEEN 1 AND 7"
            elif status == "30":
                query += " AND DATEDIFF(sub.end_date, CURDATE()) BETWEEN 1 AND 30"

            query += " ORDER BY sub.end_date ASC"
            filename = "admin_renewals.xlsx"

        # ================= LEADS =================
        elif export_type == "leads":

            query = """
                SELECT
                    lead_source,
                    school_name,
                    contact_person,
                    mobile,
                    email,
                    student_strength,
                    selected_plan,
                    status,
                    message,
                    created_at
                FROM lead_requests
                WHERE 1=1
            """

            if search:
                query += """
                    AND (
                        school_name LIKE %s
                        OR contact_person LIKE %s
                        OR mobile LIKE %s
                        OR email LIKE %s
                    )
                """
                keyword = f"%{search}%"
                params.extend([keyword, keyword, keyword, keyword])

            if status:
                query += " AND status = %s"
                params.append(status)

            if plan:
                query += " AND selected_plan = %s"
                params.append(plan)

            query += " ORDER BY id DESC"
            filename = "admin_leads.xlsx"

        # ================= USERS =================
        elif export_type == "users":

            query = """
                SELECT
                    u.name,
                    u.email,
                    u.phone,
                    u.role,
                    u.status,
                    COALESCE(s.name, 'System Admin') AS school_name,
                    u.designation,
                    u.last_login,
                    u.created_at
                FROM users u
                LEFT JOIN schools s
                    ON u.school_id = s.school_id
                WHERE 1=1
            """

            if search:
                query += """
                    AND (
                        u.name LIKE %s
                        OR u.email LIKE %s
                        OR u.phone LIKE %s
                        OR s.name LIKE %s
                    )
                """
                keyword = f"%{search}%"
                params.extend([keyword, keyword, keyword, keyword])

            if status:
                query += " AND u.status = %s"
                params.append(status)

            query += " ORDER BY u.id DESC"
            filename = "admin_users.xlsx"

        else:
            return "Invalid export type ❌"

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            return "No data found for export ❌"

        def sanitize_excel(value):
            if isinstance(value, str):
                value = value.strip()
                if value.startswith(("=", "+", "-", "@")):
                    return "'" + value
            return value

        for row in rows:
            for key in row:
                row[key] = sanitize_excel(row[key])

        df = pd.DataFrame(rows)

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            df.to_excel(
                writer,
                index=False,
                sheet_name="Export",
                startrow=2
            )

            workbook = writer.book
            sheet = writer.sheets["Export"]

            sheet["A1"] = f"SchoolSphere ERP - {export_type.title()} Export"
            sheet["A2"] = f"Generated On: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            title_font = Font(
                bold=True,
                size=16,
                color="0F172A"
            )

            header_font = Font(
                bold=True,
                color="FFFFFF"
            )

            header_fill = PatternFill(
                "solid",
                fgColor="0EA5A4"
            )

            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0")
            )

            sheet["A1"].font = title_font
            sheet["A2"].font = Font(size=11, color="64748B")

            header_row = 3

            for cell in sheet[header_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row in sheet.iter_rows(min_row=4):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            for column_cells in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column_cells[0].column)

                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                sheet.column_dimensions[column_letter].width = min(max_length + 4, 35)

            sheet.freeze_panes = "A4"
            sheet.auto_filter.ref = sheet.dimensions

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        print("❌ ADMIN EXCEL EXPORT ERROR:", e)
        return f"Export failed ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 
# =========================================================
# 📊 SUPER ADMIN - REPORTS & ANALYTICS
# =========================================================

@app.route("/superadmin/reports")
@admin_required
def superadmin_reports():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =====================================
        # FILTERS
        # =====================================

        from_date = (
            request.args.get("from_date")
            or ""
        ).strip()

        to_date = (
            request.args.get("to_date")
            or ""
        ).strip()

        report_type = (
            request.args.get("report_type")
            or ""
        ).strip()

        valid_report_types = [
            "",
            "revenue",
            "subscriptions",
            "leads",
            "certificates"
        ]

        if report_type not in valid_report_types:
            report_type = ""

        # =====================================
        # DATE VALIDATION
        # =====================================

        def valid_date(value):
            try:
                datetime.strptime(value, "%Y-%m-%d")
                return True
            except:
                return False

        if from_date and not valid_date(from_date):
            return "Invalid from date ❌"

        if to_date and not valid_date(to_date):
            return "Invalid to date ❌"

        if from_date and to_date:
            if from_date > to_date:
                return "From date cannot be greater than To date ❌"

        # =====================================
        # DATE FILTERS
        # =====================================

        date_filter_payment = ""
        date_filter_sub = ""
        date_filter_leads = ""
        date_filter_tc = ""
        date_filter_bonafide = ""
        date_filter_school = ""

        payment_params = []
        sub_params = []
        lead_params = []
        tc_params = []
        bonafide_params = []
        school_params = []

        if from_date:

            date_filter_payment += " AND DATE(created_at) >= %s"
            payment_params.append(from_date)

            date_filter_sub += " AND DATE(created_at) >= %s"
            sub_params.append(from_date)

            date_filter_leads += " AND DATE(created_at) >= %s"
            lead_params.append(from_date)

            date_filter_tc += " AND DATE(tc_date) >= %s"
            tc_params.append(from_date)

            date_filter_bonafide += " AND DATE(date) >= %s"
            bonafide_params.append(from_date)

            date_filter_school += " AND DATE(created_at) >= %s"
            school_params.append(from_date)

        if to_date:

            date_filter_payment += " AND DATE(created_at) <= %s"
            payment_params.append(to_date)

            date_filter_sub += " AND DATE(created_at) <= %s"
            sub_params.append(to_date)

            date_filter_leads += " AND DATE(created_at) <= %s"
            lead_params.append(to_date)

            date_filter_tc += " AND DATE(tc_date) <= %s"
            tc_params.append(to_date)

            date_filter_bonafide += " AND DATE(date) <= %s"
            bonafide_params.append(to_date)

            date_filter_school += " AND DATE(created_at) <= %s"
            school_params.append(to_date)

        # =====================================
        # KPI CARDS
        # =====================================

        cursor.execute(f"""
            SELECT COALESCE(SUM(amount), 0) AS revenue
            FROM payment_logs
            WHERE LOWER(payment_status) = 'success'
            {date_filter_payment}
        """, payment_params)

        total_revenue = (
            cursor.fetchone()["revenue"]
            or 0
        )

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM schools
            WHERE 1=1
            {date_filter_school}
        """, school_params)

        new_schools = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM subscriptions
            WHERE LOWER(status) = 'active'
            {date_filter_sub}
        """, sub_params)

        active_subscriptions = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Converted'
            {date_filter_leads}
        """, lead_params)

        converted_leads = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute(f"""
            SELECT
                (
                    SELECT COUNT(*)
                    FROM tc
                    WHERE 1=1
                    {date_filter_tc}
                )
                +
                (
                    SELECT COUNT(*)
                    FROM bonafide
                    WHERE 1=1
                    {date_filter_bonafide}
                ) AS total
        """, tc_params + bonafide_params)

        total_certificates = (
            cursor.fetchone()["total"]
            or 0
        )

        # =====================================
        # REVENUE TREND CHART
        # =====================================

        cursor.execute(f"""
            SELECT
                DATE_FORMAT(MIN(created_at), '%b %Y') AS month,
                SUM(amount) AS revenue,
                YEAR(created_at) AS report_year,
                MONTH(created_at) AS report_month
            FROM payment_logs
            WHERE LOWER(payment_status) = 'success'
            {date_filter_payment}
            GROUP BY
                YEAR(created_at),
                MONTH(created_at)
            ORDER BY
                report_year,
                report_month
            LIMIT 12
        """, payment_params)

        revenue_rows = cursor.fetchall()

        revenue_labels = [
            row["month"]
            for row in revenue_rows
        ]

        revenue_values = [
            float(row["revenue"] or 0)
            for row in revenue_rows
        ]

        # =====================================
        # PLAN DISTRIBUTION
        # =====================================

        cursor.execute(f"""
            SELECT
                COALESCE(plan_name, 'Unknown') AS plan_name,
                COUNT(*) AS total
            FROM subscriptions
            WHERE 1=1
            {date_filter_sub}
            GROUP BY plan_name
            ORDER BY total DESC
        """, sub_params)

        plan_rows = cursor.fetchall()

        plan_labels = [
            row["plan_name"]
            for row in plan_rows
        ]

        plan_values = [
            row["total"]
            for row in plan_rows
        ]

        # =====================================
        # LEAD FUNNEL
        # =====================================

        cursor.execute(f"""
            SELECT
                COALESCE(status, 'Unknown') AS status,
                COUNT(*) AS total
            FROM lead_requests
            WHERE 1=1
            {date_filter_leads}
            GROUP BY status
        """, lead_params)

        lead_rows = cursor.fetchall()

        lead_labels = [
            row["status"]
            for row in lead_rows
        ]

        lead_values = [
            row["total"]
            for row in lead_rows
        ]

        # =====================================
        # CERTIFICATE CHART
        # =====================================

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM tc
            WHERE 1=1
            {date_filter_tc}
        """, tc_params)

        tc_count = (
            cursor.fetchone()["total"]
            or 0
        )

        cursor.execute(f"""
            SELECT COUNT(*) AS total
            FROM bonafide
            WHERE 1=1
            {date_filter_bonafide}
        """, bonafide_params)

        bonafide_count = (
            cursor.fetchone()["total"]
            or 0
        )

        certificate_labels = [
            "TC",
            "Bonafide"
        ]

        certificate_values = [
            tc_count,
            bonafide_count
        ]

        # =====================================
        # TOP SCHOOLS TABLE
        # =====================================

        cursor.execute(f"""
            SELECT
                s.school_id,
                s.name AS school_name,
                s.email,

                COALESCE(st_count.total_students, 0) AS total_students,

                COALESCE(sub_latest.plan_name, '-') AS plan_name,
                COALESCE(sub_latest.status, '-') AS status,

                COALESCE(rev.total_revenue, 0) AS revenue,

                (
                    COALESCE(tc_count.total_tc, 0)
                    +
                    COALESCE(bon_count.total_bonafide, 0)
                ) AS certificates

            FROM schools s

            LEFT JOIN (
                SELECT
                    school_id,
                    COUNT(*) AS total_students
                FROM students
                GROUP BY school_id
            ) st_count
                ON s.school_id = st_count.school_id

            LEFT JOIN (
                SELECT
                    school_id,
                    MAX(id) AS latest_id
                FROM subscriptions
                GROUP BY school_id
            ) latest_sub
                ON s.school_id = latest_sub.school_id

            LEFT JOIN subscriptions sub_latest
                ON latest_sub.latest_id = sub_latest.id

            LEFT JOIN (
                SELECT
                    school_id,
                    SUM(amount) AS total_revenue
                FROM payment_logs
                WHERE LOWER(payment_status) = 'success'
                {date_filter_payment}
                GROUP BY school_id
            ) rev
                ON s.school_id = rev.school_id

            LEFT JOIN (
                SELECT
                    school_id,
                    COUNT(*) AS total_tc
                FROM tc
                WHERE 1=1
                {date_filter_tc}
                GROUP BY school_id
            ) tc_count
                ON s.school_id = tc_count.school_id

            LEFT JOIN (
                SELECT
                    school_id,
                    COUNT(*) AS total_bonafide
                FROM bonafide
                WHERE 1=1
                {date_filter_bonafide}
                GROUP BY school_id
            ) bon_count
                ON s.school_id = bon_count.school_id

            ORDER BY
                revenue DESC,
                certificates DESC,
                total_students DESC

            LIMIT 10
        """, payment_params + tc_params + bonafide_params)

        top_schools = cursor.fetchall()

        # =====================================
        # SIDEBAR LEADS COUNT
        # =====================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'New'
        """)

        new_leads_count = (
            cursor.fetchone()["total"]
            or 0
        )

        if report_type == "revenue":
            plan_labels = []
            plan_values = []
            lead_labels = []
            lead_values = []
            certificate_labels = []
            certificate_values = []

        elif report_type == "subscriptions":
            revenue_labels = []
            revenue_values = []
            lead_labels = []
            lead_values = []
            certificate_labels = []
            certificate_values = []

        elif report_type == "leads":
            revenue_labels = []
            revenue_values = []
            plan_labels = []
            plan_values = []
            certificate_labels = []
            certificate_values = []

        elif report_type == "certificates":
            revenue_labels = []
            revenue_values = []
            plan_labels = []
            plan_values = []
            lead_labels = []
            lead_values = []

        # =====================================
        # RENDER
        # =====================================

        return render_template(
            "superadmin/reports_admin.html",

            total_revenue=total_revenue,
            new_schools=new_schools,
            active_subscriptions=active_subscriptions,
            converted_leads=converted_leads,
            total_certificates=total_certificates,

            revenue_labels=revenue_labels,
            revenue_values=revenue_values,

            plan_labels=plan_labels,
            plan_values=plan_values,

            lead_labels=lead_labels,
            lead_values=lead_values,

            certificate_labels=certificate_labels,
            certificate_values=certificate_values,

            top_schools=top_schools,

            from_date=from_date,
            to_date=to_date,
            report_type=report_type,

            new_leads_count=new_leads_count,

            role="admin",
            school_name="Admin Panel",
            active_page="reports"
        )

    except Exception as e:

        print("❌ REPORTS ERROR:", e)

        return f"Reports Error: {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 # =========================================================
# 📊 SUPER ADMIN - REPORTS EXCEL EXPORT
# =========================================================
@app.route("/superadmin/reports/export/excel")
@admin_required
def export_admin_reports_excel():

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT COALESCE(SUM(amount), 0) AS total_revenue
            FROM payment_logs
            WHERE LOWER(payment_status) = 'success'
        """)
        total_revenue = cursor.fetchone()["total_revenue"] or 0

        cursor.execute("SELECT COUNT(*) AS total FROM schools")
        total_schools = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM subscriptions
            WHERE LOWER(status) = 'active'
        """)
        active_subscriptions = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Converted'
        """)
        converted_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM tc)
                +
                (SELECT COUNT(*) FROM bonafide) AS total
        """)
        total_certificates = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT
                s.name AS school_name,
                s.email,
                COALESCE(st_count.total_students, 0) AS students,
                COALESCE(sub_latest.plan_name, '-') AS plan_name,
                COALESCE(rev.total_revenue, 0) AS revenue,
                (
                    COALESCE(tc_count.total_tc, 0)
                    +
                    COALESCE(bon_count.total_bonafide, 0)
                ) AS certificates,
                COALESCE(sub_latest.status, '-') AS status
            FROM schools s

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_students
                FROM students
                GROUP BY school_id
            ) st_count ON s.school_id = st_count.school_id

            LEFT JOIN (
                SELECT school_id, MAX(id) AS latest_id
                FROM subscriptions
                GROUP BY school_id
            ) latest_sub ON s.school_id = latest_sub.school_id

            LEFT JOIN subscriptions sub_latest
                ON latest_sub.latest_id = sub_latest.id

            LEFT JOIN (
                SELECT school_id, SUM(amount) AS total_revenue
                FROM payment_logs
                WHERE LOWER(payment_status) = 'success'
                GROUP BY school_id
            ) rev ON s.school_id = rev.school_id

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_tc
                FROM tc
                GROUP BY school_id
            ) tc_count ON s.school_id = tc_count.school_id

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_bonafide
                FROM bonafide
                GROUP BY school_id
            ) bon_count ON s.school_id = bon_count.school_id

            ORDER BY revenue DESC, certificates DESC
        """)
        schools = cursor.fetchall()

        output = io.BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"

        # styles
        title_font = Font(bold=True, size=18, color="0F172A")
        subtitle_font = Font(size=11, color="64748B")
        header_fill = PatternFill("solid", fgColor="0EA5A4")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        

        ws.merge_cells("A1:G1")
        ws["A1"] = "SchoolSphere ERP - Reports & Analytics"
        ws["A1"].font = title_font

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated On: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
        ws["A2"].font = subtitle_font

        kpis = [
            ["Total Revenue", total_revenue],
            ["Total Schools", total_schools],
            ["Active Subscriptions", active_subscriptions],
            ["Converted Leads", converted_leads],
            ["Certificates", total_certificates],
        ]

        row = 4
        for label, value in kpis:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        table_start = 11

        headers = [
            "School Name",
            "Email",
            "Students",
            "Plan",
            "Revenue",
            "Certificates",
            "Status"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for r, school in enumerate(schools, table_start + 1):
            values = [
                school["school_name"],
                school["email"],
                school["students"],
                school["plan_name"],
                school["revenue"],
                school["certificates"],
                school["status"]
            ]

            for c, value in enumerate(values, 1):
                cell = ws.cell(row=r, column=c)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        last_row = table_start + len(schools)

        ws.freeze_panes = "A12"
        ws.auto_filter.ref = f"A{table_start}:G{last_row}"

        widths = {
            "A": 35,
            "B": 35,
            "C": 12,
            "D": 16,
            "E": 14,
            "F": 14,
            "G": 14
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

            for cell in ws["E"]:
                if cell.row > table_start:
                    cell.number_format = '₹#,##0.00'

        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="admin_reports_analytics.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("❌ REPORT EXCEL EXPORT ERROR:", e)
        return f"Excel export failed ❌ {str(e)}"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

 
 # =========================================================
# 📄 SUPER ADMIN - EXPORT REPORT PDF USING PDFKIT
# =========================================================

@app.route("/superadmin/reports/export/pdf")
@admin_required
def export_admin_reports_pdf():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                COALESCE(SUM(amount), 0) AS total_revenue
            FROM payment_logs
            WHERE LOWER(payment_status) = 'success'
        """)
        total_revenue = cursor.fetchone()["total_revenue"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM schools
        """)
        total_schools = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM subscriptions
            WHERE LOWER(status) = 'active'
        """)
        active_subscriptions = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM lead_requests
            WHERE status = 'Converted'
        """)
        converted_leads = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT
                (
                    SELECT COUNT(*) FROM tc
                )
                +
                (
                    SELECT COUNT(*) FROM bonafide
                ) AS total
        """)
        total_certificates = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT
                s.name AS school_name,
                s.email,
                COALESCE(st_count.total_students, 0) AS students,
                COALESCE(sub_latest.plan_name, '-') AS plan_name,
                COALESCE(rev.total_revenue, 0) AS revenue,
                (
                    COALESCE(tc_count.total_tc, 0)
                    +
                    COALESCE(bon_count.total_bonafide, 0)
                ) AS certificates,
                COALESCE(sub_latest.status, '-') AS status
            FROM schools s

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_students
                FROM students
                GROUP BY school_id
            ) st_count
                ON s.school_id = st_count.school_id

            LEFT JOIN (
                SELECT school_id, MAX(id) AS latest_id
                FROM subscriptions
                GROUP BY school_id
            ) latest_sub
                ON s.school_id = latest_sub.school_id

            LEFT JOIN subscriptions sub_latest
                ON latest_sub.latest_id = sub_latest.id

            LEFT JOIN (
                SELECT school_id, SUM(amount) AS total_revenue
                FROM payment_logs
                WHERE LOWER(payment_status) = 'success'
                GROUP BY school_id
            ) rev
                ON s.school_id = rev.school_id

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_tc
                FROM tc
                GROUP BY school_id
            ) tc_count
                ON s.school_id = tc_count.school_id

            LEFT JOIN (
                SELECT school_id, COUNT(*) AS total_bonafide
                FROM bonafide
                GROUP BY school_id
            ) bon_count
                ON s.school_id = bon_count.school_id

            ORDER BY revenue DESC, certificates DESC
        """)

        schools = cursor.fetchall()

        html = render_template_string("""

        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">

            <style>
                body{
                    font-family: Arial, sans-serif;
                    color:#0f172a;
                    padding:25px;
                }

                h1{
                    color:#0ea5a4;
                    margin-bottom:5px;
                }

                .subtitle{
                    color:#64748b;
                    margin-bottom:25px;
                }

                .kpi-table{
                    width:100%;
                    border-collapse:separate;
                    border-spacing:8px;
                    margin-bottom:25px;
                }
                                      
                .kpi{
                    border:1px solid #dce3ec;
                    border-radius:8px;
                    padding:12px;
                    background:#f8fafc;
                }

                .kpi label{
                    display:block;
                    font-size:11px;
                    color:#64748b;
                    margin-bottom:6px;
                }

                .kpi strong{
                    font-size:18px;
                }

                table{
                    width:100%;
                    border-collapse:collapse;
                    margin-top:15px;
                    font-size:11px;
                }

                th{
                    background:#0ea5a4;
                    color:#fff;
                    padding:8px;
                    text-align:left;
                }

                td{
                    padding:8px;
                    border:1px solid #e2e8f0;
                }

                tr:nth-child(even){
                    background:#f8fafc;
                }

                .footer{
                    margin-top:30px;
                    font-size:11px;
                    color:#64748b;
                }
            </style>
        </head>

        <body>

            <h1>SchoolSphere ERP - Reports & Analytics</h1>

            <div class="subtitle">
                Generated report summary
            </div>

           <table class="kpi-table">
            <tr>

                <td class="kpi">
                    <label>Total Revenue</label>
                    <strong>Rs. {{ total_revenue }}</strong>
                </td>

                <td class="kpi">
                    <label>Total Schools</label>
                    <strong>{{ total_schools }}</strong>
                </td>

                <td class="kpi">
                    <label>Active Subscriptions</label>
                    <strong>{{ active_subscriptions }}</strong>
                </td>

                <td class="kpi">
                    <label>Converted Leads</label>
                    <strong>{{ converted_leads }}</strong>
                </td>

                <td class="kpi">
                    <label>Certificates</label>
                    <strong>{{ total_certificates }}</strong>
                </td>

            </tr>
            </table>

            <h3>Top Schools Report</h3>

            <table>
                <thead>
                    <tr>
                        <th>School</th>
                        <th>Email</th>
                        <th>Students</th>
                        <th>Plan</th>
                        <th>Revenue</th>
                        <th>Certificates</th>
                        <th>Status</th>
                    </tr>
                </thead>

                <tbody>
                    {% for s in schools %}
                    <tr>
                        <td>{{ s.school_name }}</td>
                        <td>{{ s.email or '-' }}</td>
                        <td>{{ s.students }}</td>
                        <td>{{ s.plan_name }}</td>
                        <td>Rs. {{ s.revenue }}</td>
                        <td>{{ s.certificates }}</td>
                        <td>{{ s.status }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>

            <div class="footer">
                Generated by SchoolSphere ERP Admin Panel
            </div>

        </body>
        </html>

        """,
            total_revenue=total_revenue,
            total_schools=total_schools,
            active_subscriptions=active_subscriptions,
            converted_leads=converted_leads,
            total_certificates=total_certificates,
            schools=schools
        )

        options = {
            "page-size": "A4",
            "encoding": "UTF-8",
            "margin-top": "10mm",
            "margin-right": "10mm",
            "margin-bottom": "10mm",
            "margin-left": "10mm",
            "enable-local-file-access": None
        }

        pdf = pdfkit.from_string(
            html,
            False,
            configuration=pdf_config,
            options=options
        )

        response = make_response(pdf)

        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = (
            "attachment; filename=admin_reports.pdf"
        )

        return response

    except Exception as e:

        print("❌ ADMIN REPORT PDF ERROR:", e)
        return f"PDF export failed ❌ {str(e)}"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# ⚙️ SUPER ADMIN - SETTINGS PAGE
# =========================================================
@app.route("/superadmin/settings")
@admin_required
def superadmin_settings():

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        # ================= SETTINGS =================
        cursor.execute("""
            SELECT *
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        row = cursor.fetchone()

        if not row:
            return "System settings record missing ❌"

        columns = [col[0] for col in cursor.description]
        settings = dict(zip(columns, row))

        # ================= ACTIVE SCHOOLS =================
        cursor.execute("""
            SELECT COUNT(*)
            FROM schools
            WHERE is_active = 1
        """)
        total_schools = cursor.fetchone()[0] or 0

        # ================= ALERTS =================
        cursor.execute("""
            SELECT COUNT(*)
            FROM lead_requests
            WHERE status = 'New'
        """)
        new_leads = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*)
            FROM subscriptions
            WHERE end_date <= DATE_ADD(CURDATE(), INTERVAL 7 DAY)
        """)
        renewal_alerts = cursor.fetchone()[0] or 0

        pending_alerts = new_leads + renewal_alerts

        # ================= SUBSCRIPTION PLANS =================
        cursor.execute("""
            SELECT
                id,
                plan_name,
                monthly_price
            FROM subscription_plans
            WHERE is_active = 1
            ORDER BY monthly_price ASC
        """)

        plans = [
            {
                "id": p[0],
                "plan_name": p[1],
                "monthly_price": p[2]
            }
            for p in cursor.fetchall()
        ]

        # ================= DEFAULT PLAN NAME =================
        default_plan_name = "No Plan Selected"

        if settings.get("default_plan_id"):
            for plan in plans:
                if int(plan["id"]) == int(settings["default_plan_id"]):
                    default_plan_name = plan["plan_name"]
                    break

        # ================= BACKUP LOGS =================
        cursor.execute("""
            SELECT
                id,
                backup_file,
                backup_status,
                backup_size,
                backup_date,
                backup_type
            FROM system_backup_logs
            ORDER BY id DESC
            LIMIT 10
        """)

        backup_rows = cursor.fetchall()
        backup_columns = [col[0] for col in cursor.description]

        backup_logs = [
            dict(zip(backup_columns, row))
            for row in backup_rows
        ]

        return render_template(
            "superadmin/superadmin_settings.html",
            settings=settings,
            plans=plans,
            default_plan_name=default_plan_name,
            backup_logs=backup_logs,
            total_schools=total_schools,
            pending_alerts=pending_alerts,
            role="admin",
            school_name="Admin Panel",
            active_page="settings"
        )

    except Exception as e:
        print("❌ SETTINGS ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            
# =========================================================
# 💾 SAVE GENERAL SETTINGS
# =========================================================
@app.route("/superadmin/settings/general", methods=["POST"])
@admin_required
def save_general_settings():

    conn = None
    cursor = None

    try:

        # ================= FORM DATA =================
        system_name = (
            request.form.get("system_name") or ""
        ).strip()

        support_email = (
            request.form.get("support_email") or ""
        ).strip().lower()

        support_phone = (
            request.form.get("support_phone") or ""
        ).strip()

        default_language = (
            request.form.get("default_language") or "English"
        ).strip()

        timezone_value = (
            request.form.get("timezone") or "Asia/Kolkata"
        ).strip()

        # ================= VALIDATION =================
        if not system_name:
            return "System name required ❌"

        if len(system_name) > 150:
            return "System name too long ❌"

        if support_email:

            if len(support_email) > 150:
                return "Support email too long ❌"

            if not is_valid_email(support_email):
                return "Invalid support email ❌"

        if support_phone:

            clean_phone = support_phone.replace("+", "")

            if not clean_phone.isdigit():
                return "Invalid support phone ❌"

            if len(support_phone) > 20:
                return "Support phone too long ❌"

        valid_languages = [
            "English",
            "Marathi"
        ]

        if default_language not in valid_languages:
            return "Invalid language ❌"

        valid_timezones = [
            "Asia/Kolkata"
        ]

        if timezone_value not in valid_timezones:
            return "Invalid timezone ❌"

        # ================= DB =================
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                id,
                system_logo
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings_row = cursor.fetchone()

        if not settings_row:
            return "System settings record missing ❌"

        settings_id = settings_row[0]
        old_logo = settings_row[1]

        # ================= LOGO UPLOAD =================
        logo_filename = None

        logo = request.files.get("system_logo")

        if logo and logo.filename:

            original_filename = secure_filename(
                logo.filename
            )

            if "." not in original_filename:
                return "Invalid logo file ❌"

            extension = (
                original_filename
                .rsplit(".", 1)[1]
                .lower()
            )

            allowed_extensions = [
                "png",
                "jpg",
                "jpeg",
                "webp"
            ]

            if extension not in allowed_extensions:
                return "Invalid logo file type ❌"

            # Optional size check: 2 MB
            logo.seek(0, os.SEEK_END)
            file_size = logo.tell()
            logo.seek(0)

            if file_size > 2 * 1024 * 1024:
                return "Logo size must be below 2 MB ❌"

            upload_folder = os.path.join(
                "static",
                "uploads",
                "system"
            )

            os.makedirs(
                upload_folder,
                exist_ok=True
            )

            new_filename = (
                f"system_logo_{settings_id}_"
                f"{int(datetime.now().timestamp())}."
                f"{extension}"
            )

            upload_path = os.path.join(
                upload_folder,
                new_filename
            )

            logo.save(upload_path)

            logo_filename = (
                f"uploads/system/{new_filename}"
            )

            # ================= DELETE OLD LOGO =================
            if old_logo:

                old_logo_path = os.path.join(
                    "static",
                    old_logo
                )

                if os.path.exists(old_logo_path):

                    try:
                        os.remove(old_logo_path)

                    except Exception as delete_error:
                        print(
                            "⚠️ OLD LOGO DELETE ERROR:",
                            delete_error
                        )

        # ================= UPDATE SETTINGS =================
        if logo_filename:

            cursor.execute("""
                UPDATE system_settings
                SET
                    system_name = %s,
                    support_email = %s,
                    support_phone = %s,
                    default_language = %s,
                    timezone = %s,
                    system_logo = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (
                system_name,
                support_email,
                support_phone,
                default_language,
                timezone_value,
                logo_filename,
                settings_id
            ))

        else:

            cursor.execute("""
                UPDATE system_settings
                SET
                    system_name = %s,
                    support_email = %s,
                    support_phone = %s,
                    default_language = %s,
                    timezone = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (
                system_name,
                support_email,
                support_phone,
                default_language,
                timezone_value,
                settings_id
            ))

        if cursor.rowcount < 1:
            conn.rollback()
            return "No changes detected ❌"

        conn.commit()

        print("✅ GENERAL SETTINGS UPDATED")

        return redirect(
            url_for("superadmin_settings") + "#general"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ GENERAL SETTINGS ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 💳 SAVE SUBSCRIPTION SETTINGS
# =========================================================
@app.route("/superadmin/settings/subscription", methods=["POST"])
@admin_required
def save_subscription_settings():

    conn = None
    cursor = None

    try:

        default_plan_id = (
            request.form.get("default_plan_id") or ""
        ).strip()

        trial_days = (
            request.form.get("trial_days") or ""
        ).strip()

        grace_period = (
            request.form.get("grace_period") or ""
        ).strip()

        # ================= VALIDATION =================
        if not default_plan_id:
            return "Default plan missing ❌"

        if not trial_days:
            return "Trial days missing ❌"

        if not grace_period:
            return "Grace period missing ❌"

        if not default_plan_id.isdigit():
            return "Invalid default plan ❌"

        if not trial_days.isdigit():
            return "Invalid trial days ❌"

        if not grace_period.isdigit():
            return "Invalid grace period ❌"

        default_plan_id = int(default_plan_id)
        trial_days = int(trial_days)
        grace_period = int(grace_period)

        if trial_days < 0 or trial_days > 365:
            return "Trial days must be between 0 and 365 ❌"

        if grace_period < 0 or grace_period > 90:
            return "Grace period must be between 0 and 90 ❌"

        conn = get_connection()
        cursor = conn.cursor()

        # ================= CHECK PLAN =================
        cursor.execute("""
            SELECT id
            FROM subscription_plans
            WHERE id = %s
            AND is_active = 1
            LIMIT 1
        """, (default_plan_id,))

        if not cursor.fetchone():
            return "Selected subscription plan not found ❌"

        # ================= CHECK SETTINGS =================
        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        settings_id = settings[0]

        # ================= UPDATE =================
        cursor.execute("""
            UPDATE system_settings
            SET
                default_plan_id = %s,
                trial_days = %s,
                grace_period = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            default_plan_id,
            trial_days,
            grace_period,
            settings_id
        ))

        conn.commit()

        print("✅ SUBSCRIPTION SETTINGS UPDATED")

        return redirect(
            url_for("superadmin_settings") + "#subscription"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ SUBSCRIPTION SETTINGS ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🔐 SAVE SECURITY SETTINGS
# =========================================================
@app.route("/superadmin/settings/security", methods=["POST"])
@admin_required
def save_security_settings():

    conn = None
    cursor = None

    try:
        password_length = (request.form.get("password_length") or "").strip()
        login_attempt_limit = (request.form.get("login_attempt_limit") or "").strip()
        session_timeout = (request.form.get("session_timeout") or "").strip()

        if not password_length or not login_attempt_limit or not session_timeout:
            return "Required fields missing ❌"

        if not password_length.isdigit():
            return "Invalid password length ❌"

        if not login_attempt_limit.isdigit():
            return "Invalid login attempt limit ❌"

        if not session_timeout.isdigit():
            return "Invalid session timeout ❌"

        password_length = int(password_length)
        login_attempt_limit = int(login_attempt_limit)
        session_timeout = int(session_timeout)

        if password_length < 6 or password_length > 50:
            return "Password length must be between 6 and 50 ❌"

        if login_attempt_limit < 1 or login_attempt_limit > 20:
            return "Login attempts must be between 1 and 20 ❌"

        if session_timeout < 5 or session_timeout > 1440:
            return "Session timeout must be between 5 and 1440 minutes ❌"

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        settings_id = settings[0]

        cursor.execute("""
            UPDATE system_settings
            SET
                password_length = %s,
                login_attempt_limit = %s,
                session_timeout = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            password_length,
            login_attempt_limit,
            session_timeout,
            settings_id
        ))

        conn.commit()

        return redirect(
            url_for("superadmin_settings") + "#security"
        )

    except Exception as e:
        if conn:
            conn.rollback()

        print("❌ SECURITY SETTINGS ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()

        if conn:
            conn.close()
 
# =========================================================
# 💳 RENEW SUBSCRIPTION PAGE
# =========================================================

@app.route("/clerk/subscription/renew")
@login_required
def renew_subscription():

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    school_id = session.get(
        "clerk_school_id"
    )

    if not school_id:
        return "School session missing ❌"

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor(
            dictionary=True
        )

        # =========================================
        # CURRENT SUBSCRIPTION
        # =========================================

        cursor.execute("""

            SELECT

                s.id,
                s.plan_id,
                s.amount,
                s.end_date,
                s.status,

                p.plan_name,
                p.student_limit,
                p.staff_limit,
                p.storage_limit,
                p.support_type,

                p.enable_tc_management,
                p.enable_bonafide_management,
                p.enable_import_export,
                p.enable_attendance,
                p.enable_fee_management,
                p.enable_teacher_management,
                p.enable_results,
                p.enable_timetable,
                p.enable_notice_board

            FROM subscriptions s

            LEFT JOIN subscription_plans p
                ON s.plan_id = p.id

            WHERE s.school_id = %s

            ORDER BY s.id DESC

            LIMIT 1

        """, (
            school_id,
        ))

        subscription = cursor.fetchone()

        if not subscription:
            return "Subscription not found ❌"

        # =========================================
        # AVAILABLE PLANS
        # =========================================

        cursor.execute("""

            SELECT

                id,
                plan_name,

                monthly_price,
                yearly_price,

                student_limit,
                staff_limit,

                storage_limit,
                support_type,

                enable_tc_management,
                enable_bonafide_management,
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board

            FROM subscription_plans

            WHERE is_active = 1

            AND plan_name IN (
                'Starter',
                'Essential',
                'Professional'
            )

            ORDER BY monthly_price ASC

        """)

        plans = cursor.fetchall()

        # =========================================
        # CURRENT PLAN FEATURES COUNT
        # =========================================

        current_feature_count = 0

        feature_columns = [

            "enable_tc_management",
            "enable_bonafide_management",
            "enable_import_export",
            "enable_attendance",
            "enable_fee_management",
            "enable_teacher_management",
            "enable_results",
            "enable_timetable",
            "enable_notice_board"

        ]

        if subscription["plan_id"]:

            cursor.execute("""

                SELECT *

                FROM subscription_plans

                WHERE id = %s

            """, (

                subscription["plan_id"],

            ))

            current_plan = cursor.fetchone()

            if current_plan:

                for col in feature_columns:

                    if current_plan.get(col) == "Enabled":
                        current_feature_count += 1

            # =========================================
            # DAYS LEFT
            # =========================================

            days_left = -1

            if subscription.get("end_date"):

                end_date = subscription["end_date"]

                days_left = (
                    end_date - date.today()
                ).days

                if days_left < 0:
                    days_left = 0

        # =========================================
        # RENDER
        # =========================================

        return render_template(

            "subscription/renew.html",

            subscription=subscription,

            plans=plans,

            current_feature_count=current_feature_count,

            days_left=days_left,

            role="clerk",

            active_page="subscription"

        )

    except Exception as e:

        print(
            "❌ RENEW PAGE ERROR:",
            e
        )

        return "Renew page failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📄 SAVE CERTIFICATE SETTINGS
# =========================================================
@app.route("/superadmin/settings/certificate", methods=["POST"])
@admin_required
def save_certificate_settings():

    conn = None
    cursor = None

    try:
        tc_prefix = (
            request.form.get("tc_prefix") or "TC"
        ).strip().upper()

        bonafide_prefix = (
            request.form.get("bonafide_prefix") or "BON"
        ).strip().upper()

        auto_numbering = (
            request.form.get("auto_numbering") or "Enabled"
        ).strip()

        enable_certificate_labels = (
            request.form.get("enable_certificate_labels") or "Enabled"
        ).strip()

        show_tc_logo = (
            request.form.get("show_tc_logo") or "Enabled"
        ).strip()

        show_tc_watermark = (
            request.form.get("show_tc_watermark") or "Enabled"
        ).strip()

        show_bonafide_logo = (
            request.form.get("show_bonafide_logo") or "Enabled"
        ).strip()

        show_bonafide_watermark = (
            request.form.get("show_bonafide_watermark") or "Enabled"
        ).strip()

        # ================= VALIDATION =================
        if len(tc_prefix) > 10:
            return "TC prefix too long ❌"

        if len(bonafide_prefix) > 10:
            return "Bonafide prefix too long ❌"

        if not re.fullmatch(r"[A-Z0-9_-]+", tc_prefix):
            return "Invalid TC prefix ❌"

        if not re.fullmatch(r"[A-Z0-9_-]+", bonafide_prefix):
            return "Invalid Bonafide prefix ❌"

        valid_options = ["Enabled", "Disabled"]

        for value in [
            auto_numbering,
            enable_certificate_labels,
            show_tc_logo,
            show_tc_watermark,
            show_bonafide_logo,
            show_bonafide_watermark
        ]:
            if value not in valid_options:
                return "Invalid certificate setting ❌"

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings_row = cursor.fetchone()

        if not settings_row:
            return "System settings not found ❌"

        settings_id = settings_row[0]

        cursor.execute("""
            UPDATE system_settings
            SET
                tc_prefix = %s,
                bonafide_prefix = %s,
                auto_numbering = %s,
                enable_certificate_labels = %s,
                show_tc_logo = %s,
                show_tc_watermark = %s,
                show_bonafide_logo = %s,
                show_bonafide_watermark = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            tc_prefix,
            bonafide_prefix,
            auto_numbering,
            enable_certificate_labels,
            show_tc_logo,
            show_tc_watermark,
            show_bonafide_logo,
            show_bonafide_watermark,
            settings_id
        ))

        conn.commit()

        return redirect(
            url_for("superadmin_settings") + "#certificate"
        )

    except Exception as e:
        if conn:
            conn.rollback()

        print("❌ CERTIFICATE SETTINGS ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()

        if conn:
            conn.close()

 
# =========================================================
# 🧩 SAVE FEATURE MODULE SETTINGS
# =========================================================
@app.route("/superadmin/settings/modules", methods=["POST"])
@admin_required
def save_feature_module_settings():

    conn = None
    cursor = None

    try:

        fields = [
            "enable_tc_management",
            "enable_bonafide_management",
            "enable_attendance",
            "enable_fee_management",
            "enable_teacher_management",
            "enable_results",
            "enable_import_export",
            "enable_timetable",
            "enable_notice_board"
        ]

        values = {}

        for field in fields:

            value = (
                request.form.get(field) or "Disabled"
            ).strip()

            if value not in ["Enabled", "Disabled"]:
                return "Invalid feature setting ❌"

            values[field] = value

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings_row = cursor.fetchone()

        if not settings_row:
            return "System settings not found ❌"

        settings_id = settings_row[0]

        cursor.execute("""
            UPDATE system_settings
            SET
                enable_tc_management = %s,
                enable_bonafide_management = %s,
                enable_attendance = %s,
                enable_fee_management = %s,
                enable_teacher_management = %s,
                enable_results = %s,
                enable_import_export = %s,
                enable_timetable = %s,
                enable_notice_board = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            values["enable_tc_management"],
            values["enable_bonafide_management"],
            values["enable_attendance"],
            values["enable_fee_management"],
            values["enable_teacher_management"],
            values["enable_results"],
            values["enable_import_export"],
            values["enable_timetable"],
            values["enable_notice_board"],
            settings_id
        ))

        conn.commit()

        print("✅ FEATURE MODULE SETTINGS UPDATED")

        return redirect(
            url_for("superadmin_settings") + "#modules"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ FEATURE MODULE SETTINGS ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()



# =========================================================
# 💳 SAVE PAYMENT SETTINGS
# =========================================================
@app.route("/superadmin/settings/payment", methods=["POST"])
@admin_required
def save_payment_settings():

    conn = None
    cursor = None

    try:
        payment_gateway = (
            request.form.get("payment_gateway") or "Razorpay"
        ).strip()

        razorpay_key_id = (
            request.form.get("razorpay_key_id") or ""
        ).strip()

        razorpay_mode = (
            request.form.get("razorpay_mode") or "Sandbox"
        ).strip()

        currency = (
            request.form.get("currency") or "INR"
        ).strip().upper()

        gst_percentage = (
            request.form.get("gst_percentage") or "18"
        ).strip()

        # ================= VALIDATION =================
        allowed_gateways = [
            "Razorpay",
            "Disabled"
        ]

        if payment_gateway not in allowed_gateways:
            return "Invalid payment gateway ❌"

        allowed_modes = [
            "Sandbox",
            "Live"
        ]

        if razorpay_mode not in allowed_modes:
            return "Invalid payment mode ❌"

        allowed_currencies = [
            "INR"
        ]

        if currency not in allowed_currencies:
            return "Invalid currency ❌"

        try:
            gst_percentage = float(gst_percentage)
        except ValueError:
            return "Invalid GST percentage ❌"

        if gst_percentage < 0 or gst_percentage > 100:
            return "GST must be between 0 and 100 ❌"

        if len(razorpay_key_id) > 200:
            return "Razorpay key too long ❌"

        if payment_gateway == "Razorpay" and not razorpay_key_id:
            return "Razorpay Key ID required ❌"

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings_row = cursor.fetchone()

        if not settings_row:
            return "System settings not found ❌"

        settings_id = settings_row[0]

        cursor.execute("""
            UPDATE system_settings
            SET
                payment_gateway = %s,
                razorpay_key_id = %s,
                razorpay_mode = %s,
                currency = %s,
                gst_percentage = %s,
                updated_at = NOW()
            WHERE id = %s
        """, (
            payment_gateway,
            razorpay_key_id,
            razorpay_mode,
            currency,
            gst_percentage,
            settings_id
        ))

        conn.commit()

        print("✅ PAYMENT SETTINGS UPDATED")

        return redirect(
            url_for("superadmin_settings") + "#payment"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ PAYMENT SETTINGS ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📧 SAVE SMTP SETTINGS
# =========================================================
@app.route(
    "/superadmin/settings/smtp",
    methods=["POST"]
)
@admin_required
def save_smtp_settings():

    conn = None
    cursor = None

    try:

        smtp_email = (
            request.form.get(
                "smtp_email"
            ) or ""
        ).strip().lower()

        smtp_password = (
            request.form.get(
                "smtp_password"
            ) or ""
        ).strip()

        smtp_server = (
            request.form.get(
                "smtp_server"
            ) or ""
        ).strip()

        smtp_port = (
            request.form.get(
                "smtp_port"
            ) or ""
        ).strip()

        smtp_tls = (
            request.form.get(
                "smtp_tls"
            ) or "Enabled"
        ).strip()

        # =====================================
        # REQUIRED
        # =====================================

        if not smtp_email:
            return "SMTP Email required ❌"

        if not smtp_server:
            return "SMTP Server required ❌"

        if not smtp_port:
            return "SMTP Port required ❌"

        # =====================================
        # EMAIL
        # =====================================

        if not is_valid_email(
            smtp_email
        ):
            return "Invalid SMTP Email ❌"

        if len(smtp_email) > 150:
            return "SMTP email too long ❌"

        if len(smtp_server) > 255:
            return "SMTP server too long ❌"
       
        # =====================================
        # PORT
        # =====================================

        if not smtp_port.isdigit():
            return "Invalid SMTP Port ❌"

        smtp_port = int(smtp_port)

        if smtp_port < 1 or smtp_port > 65535:
            return "SMTP Port out of range ❌"

        # =====================================
        # TLS
        # =====================================

        if smtp_tls not in [
            "Enabled",
            "Disabled"
        ]:
            return "Invalid TLS setting ❌"

        # =====================================
        # DB
        # =====================================

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                id,
                smtp_password
            FROM system_settings
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        settings_id = settings[0]

        existing_password = settings[1]

        # =====================================
        # KEEP OLD PASSWORD
        # =====================================

        if not smtp_password:
            smtp_password = existing_password

        if not smtp_password:
            return "SMTP password required ❌"

        # =====================================
        # UPDATE
        # =====================================

        cursor.execute("""
            UPDATE system_settings
            SET

                smtp_email = %s,
                smtp_password = %s,
                smtp_server = %s,
                smtp_port = %s,
                smtp_tls = %s,

                updated_at = NOW()

            WHERE id = %s
        """, (

            smtp_email,
            smtp_password,
            smtp_server,
            smtp_port,
            smtp_tls,
            settings_id

        ))

        conn.commit()

        print(
            "✅ SMTP SETTINGS UPDATED"
        )

        return redirect(
            url_for("superadmin_settings") + "#smtp"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ SMTP SETTINGS ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 💾 CREATE REAL DATABASE BACKUP
# =========================================================

# @app.route(
#     "/superadmin/settings/backup",
#     methods=["POST"]
# )
# @admin_required
# def create_backup():

#     conn = None
#     cursor = None

#     try:

        # =========================================
        # DB CONNECTION
        # =========================================

        # conn = get_connection()

        # BACKUP DATABASE requires autocommit

        # conn.autocommit = True

        # cursor = conn.cursor()

        # =========================================
        # ADMIN SESSION
        # =========================================

        # admin_id = session.get(
        #     "admin_user_id"
        # )

        # if not admin_id:
        #     return "Admin session missing ❌"

        # =========================================
        # BACKUP DIRECTORY
        # =========================================

        # backup_dir = os.path.join(
        #     os.getcwd(),
        #     "backups"
        # )

        # os.makedirs(
        #     backup_dir,
        #     exist_ok=True
        # )

        # =========================================
        # FILE NAME
        # =========================================

        # backup_date = datetime.now()

        # backup_file = (
        #     f"backup_"
        #     f"{backup_date.strftime('%Y%m%d_%H%M%S')}"
        #     f".bak"
        # )

        # backup_path = os.path.join(
        #     backup_dir,
        #     backup_file
        # )

        # =========================================
        # SQL SERVER BACKUP COMMAND
        # =========================================

        # sql = f"""
        # BACKUP DATABASE SchoolERP
        # TO DISK = '{backup_path}'
        # WITH INIT,
        # NAME = 'Full Backup of SchoolERP';
        # """
        # cursor.execute(sql)

        # =========================================
        # FILE SIZE
        # =========================================

        # size_bytes = os.path.getsize(
        #     backup_path
        # )

        # size_mb = round(
        #     size_bytes / (1024 * 1024),
        #     2
        # )

        # backup_size = f"{size_mb} MB"

        # =========================================
        # SAVE BACKUP LOG
        # =========================================

        # cursor.execute("""

        #     INSERT INTO system_backup_logs
        #     (
        #         backup_file,
        #         backup_status,
        #         backup_size,
        #         backup_date,
        #         backup_type,
        #         created_by
        #     )

        #     VALUES (%s, %s, %s, %s, %s, %s)

        # """, (

        #     backup_file,
        #     "Success",
        #     backup_size,
        #     backup_date,
        #     "manual",
        #     admin_id

        # ))

        # =========================================
        # UPDATE SETTINGS
        # =========================================

    #     cursor.execute("""

    #         UPDATE system_settings

    #         SET

    #             last_backup = %s,
    #             backup_status = %s,
    #             updated_at = NOW()

    #         WHERE id = 1

    #     """, (

    #         backup_date,
    #         "Enabled"

    #     ))

    #     conn.commit()

    #     print("✅ REAL BACKUP CREATED")

    #     return redirect(
    #         url_for("superadmin_settings")
    #     )

    # except Exception as e:

    #     if conn:
    #         conn.rollback()

    #     print(
    #         "❌ BACKUP ERROR:",
    #         e
    #     )

    #     return "Something went wrong ❌"

    # finally:

    #     if cursor:
    #         cursor.close()

    #     if conn:
    #         conn.close()

 # =========================================================
# 📥 DOWNLOAD BACKUP
# =========================================================

@app.route(
    "/download-backup/<filename>"
)
@admin_required
def download_backup(filename):

    filename = secure_filename(filename)

    backup_path = os.path.join(
        os.getcwd(),
        "backups",
        filename
    )

    if not os.path.exists(
        backup_path
    ):
        return "Backup file not found ❌"

    return send_file(

        backup_path,

        as_attachment=True

    )

# =========================================================
# 🤖 AUTO BACKUP FUNCTION
# =========================================================

# def auto_backup():

#     conn = None
#     cursor = None

#     try:

        # =========================================
        # DB CONNECTION
        # =========================================

        # conn = get_connection()

        # conn.autocommit = True

        # cursor = conn.cursor()

        # =========================================
        # BACKUP DIRECTORY
        # =========================================

        # backup_dir = os.path.join(
        #     os.getcwd(),
        #     "backups"
        # )

        # os.makedirs(
        #     backup_dir,
        #     exist_ok=True
        # )

        # =========================================
        # BACKUP FILE NAME
        # =========================================

        # backup_date = datetime.now()

        # backup_file = (
        #     f"auto_backup_"
        #     f"{backup_date.strftime('%Y%m%d_%H%M%S')}"
        #     f".bak"
        # )

        # backup_path = os.path.join(
        #     backup_dir,
        #     backup_file
        # )

        # =========================================
        # SQL SERVER BACKUP
        # =========================================

        # sql = f"""

        # BACKUP DATABASE SchoolERP

        # TO DISK = '{backup_path}'

        # WITH INIT,
        # NAME = 'Auto Backup';

        # """

        # cursor.execute(sql)

        # =========================================
        # FILE SIZE
        # =========================================

        # size_bytes = os.path.getsize(
        #     backup_path
        # )

        # size_mb = round(
        #     size_bytes / (1024 * 1024),
        #     2
        # )

        # backup_size = f"{size_mb} MB"

        # =========================================
        # SAVE BACKUP LOG
        # =========================================

        # cursor.execute("""

        #     INSERT INTO system_backup_logs
        #     (
        #         backup_file,
        #         backup_status,
        #         backup_size,
        #         backup_date,
        #         backup_type,
        #         created_by
        #     )

        #     VALUES (%s, %s, %s, %s, %s, %s)

        # """, (

        #     backup_file,
        #     "Success",
        #     backup_size,
        #     backup_date,
        #     "automatic",
        #     1

        # ))

        # =========================================
        # KEEP ONLY LATEST 10 BACKUPS
        # =========================================

        # cursor.execute("""

        #     SELECT

        #         id,
        #         backup_file

        #     FROM system_backup_logs
                       
        #     WHERE backup_type = 'automatic'

        #     ORDER BY backup_date DESC

        # """)

        # logs = cursor.fetchall()

        # =========================================
        # DELETE OLD BACKUPS
        # =========================================

        # if len(logs) > 10:

        #     old_logs = logs[10:]

        #     for log in old_logs:

        #         old_id = log[0]

        #         old_file = log[1]

        #         old_path = os.path.join(
        #             backup_dir,
        #             old_file
        #         )

                # =================================
                # DELETE FILE
                # =================================

                # if os.path.exists(old_path):

                #     os.remove(old_path)

                # =================================
                # DELETE DB LOG
                # =================================

                # cursor.execute("""

                #     DELETE FROM system_backup_logs

                #     WHERE id = %s

                # """, (

                #     old_id,

                # ))

        # =========================================
        # SAVE ALL CHANGES
        # =========================================

    #     conn.commit()

    #     print("✅ AUTO BACKUP CREATED")

    # except Exception as e:

    #     print(
    #         "❌ AUTO BACKUP ERROR:",
    #         e
    #     )

    # finally:

    #     # =========================================
    #     # CLOSE DB
    #     # =========================================

    #     if cursor:
    #         cursor.close()

    #     if conn:
    #         conn.close()


# =========================================================
# ⏰ AUTO BACKUP SCHEDULER
# =========================================================

# scheduler = BackgroundScheduler(
#     daemon=True
# )

# =========================================================
# TEST MODE
# EVERY 30 SECONDS
# =========================================================

# scheduler.add_job(

#     func=auto_backup,

#     trigger="cron",

#     hour=2,
#     minute=0

# )

# =========================================================
# START SCHEDULER
# =========================================================

# if not scheduler.running:
#     scheduler.start()

# =========================================================
# 🗑 DELETE BACKUP
# =========================================================

# @app.route(
#     "/delete-backup/<int:backup_id>"
# )
# @admin_required
# def delete_backup(backup_id):

#     conn = None
#     cursor = None

#     try:

#         conn = get_connection()
#         cursor = conn.cursor()

        # =========================================
        # GET BACKUP FILE
        # =========================================

        # cursor.execute("""

        #     SELECT backup_file
        #     FROM system_backup_logs
        #     WHERE id = %s

        # """, (backup_id,))

        # backup = cursor.fetchone()

        # if not backup:
        #     return "Backup not found ❌"

        # backup_file = backup[0]

        # =========================================
        # FILE PATH
        # =========================================

        # backup_path = os.path.join(
        #     os.getcwd(),
        #     "backups",
        #     backup_file
        # )

        # =========================================
        # DELETE FILE
        # =========================================

        # if os.path.exists(backup_path):

        #     os.remove(backup_path)

        # =========================================
        # DELETE DB LOG
        # =========================================

    #     cursor.execute("""

    #         DELETE FROM system_backup_logs
    #         WHERE id = %s

    #     """, (backup_id,))

    #     conn.commit()

    #     print("✅ BACKUP DELETED")

    #     return redirect(
    #         url_for("superadmin_settings")
    #     )

    # except Exception as e:

    #     if conn:
    #         conn.rollback()

    #     print("❌ DELETE BACKUP ERROR:", e)

    #     return "Something went wrong ❌"

    # finally:

    #     if cursor:
    #         cursor.close()

    #     if conn:
    #         conn.close()

 # =========================================================
# ♻ SAFE BACKUP VERIFY RESTORE
# =========================================================

# @app.route(
#     "/restore-backup/<filename>"
# )
# @admin_required
# def restore_backup(filename):

#     filename = secure_filename(filename)

#     conn = None
#     cursor = None

#     try:

        # =========================================
        # BACKUP PATH
        # =========================================

        # backup_path = os.path.join(
        #     os.getcwd(),
        #     "backups",
        #     filename
        # )

        # if not os.path.exists(backup_path):

        #     return "Backup file not found ❌"
        
        # if not filename.endswith(".bak"):
            
        #     return "Invalid backup file ❌"

        # =========================================
        # CONNECT
        # =========================================

        # conn = get_connection()

        # conn.autocommit = True

        # cursor = conn.cursor()

        # =========================================
        # TEMP DATABASE NAME
        # =========================================

        # temp_db = "SchoolERP_TestRestore"

        # =========================================
        # DELETE OLD TEST DB
        # =========================================

        # cursor.execute(f"""

        # IF DB_ID('{temp_db}') IS NOT NULL
        # BEGIN

        #     ALTER DATABASE [{temp_db}]
        #     SET SINGLE_USER
        #     WITH ROLLBACK IMMEDIATE;

        #     DROP DATABASE [{temp_db}];

        # END

        # """)

        # =========================================
        # RESTORE TO TEST DATABASE
        # =========================================

    #     restore_sql = f"""

    #     RESTORE DATABASE [{temp_db}]

    #     FROM DISK = '{backup_path}'

    #     WITH
    #     MOVE 'SchoolERP'
    #     TO 'C:\\Program Files\\Microsoft SQL Server\\MSSQL17.SQLEXPRESS\\MSSQL\\DATA\\{temp_db}.mdf',

    #     MOVE 'SchoolERP_log'
    #     TO 'C:\\Program Files\\Microsoft SQL Server\\MSSQL17.SQLEXPRESS\\MSSQL\\DATA\\{temp_db}_log.ldf',

    #     REPLACE

    #     """

    #     cursor.execute(restore_sql)

    #     print("✅ BACKUP VERIFIED SUCCESSFULLY")

    #     return redirect(
    #         url_for("superadmin_settings")
    #     )

    # except Exception as e:

    #     print("❌ RESTORE ERROR:", e)

    #     return "Something went wrong ❌"

    # finally:

    #     if cursor:
    #         cursor.close()

    #     if conn:
    #         conn.close()
            
# =========================================================
# 🛠 SAVE MAINTENANCE SETTINGS (PRODUCTION SAFE)
# =========================================================
@app.route(
    "/superadmin/settings/maintenance",
    methods=["POST"]
)
@admin_required
def save_maintenance_settings():

    conn = None
    cursor = None

    try:

        # =========================================
        # GET FORM DATA
        # =========================================

        maintenance_mode = (
            request.form.get("maintenance_mode") or "OFF"
        ).strip().upper()

        maintenance_message = (
            request.form.get("maintenance_message") or ""
        ).strip()

        # =========================================
        # VALIDATION
        # =========================================

        if maintenance_mode not in [
            "ON",
            "OFF"
        ]:
            return "Invalid maintenance mode ❌"

        if len(maintenance_message) > 1000:
            return "Maintenance message too long ❌"

        # =========================================
        # DEFAULT MESSAGE
        # =========================================

        if not maintenance_message:

            if maintenance_mode == "ON":

                maintenance_message = (
                    "ERP system is currently "
                    "under maintenance. "
                    "Please try again later."
                )

            else:

                maintenance_message = (
                    "System running normally"
                )

               

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()

        if not conn:
            return "Database connection failed ❌"

        cursor = conn.cursor()

        # =========================================
        # CHECK SETTINGS EXISTS
        # =========================================

        cursor.execute("""
            SELECT id
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        settings_id = settings[0]

        # =========================================
        # UPDATE SETTINGS
        # =========================================

        cursor.execute("""

            UPDATE system_settings

            SET

                maintenance_mode = %s,
                maintenance_message = %s,
                updated_at = NOW()

            WHERE id = %s

        """, (

            maintenance_mode,
            maintenance_message,
            settings_id

        ))

        if cursor.rowcount == 0:
            return "No settings updated ❌"

        # =========================================
        # SAVE
        # =========================================

        conn.commit()

        # =========================================
        # TERMINAL LOG
        # =========================================

        print(
            f"✅ MAINTENANCE MODE UPDATED: "
            f"{maintenance_mode}"
        )

        # =========================================
        # REDIRECT
        # =========================================

        return redirect(
            url_for("superadmin_settings") + "#maintenance"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ MAINTENANCE SETTINGS ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


 # =========================================================
# 🎨 SAVE BRANDING SETTINGS
# =========================================================
@app.route("/superadmin/save-branding-settings", methods=["POST"])
@admin_required
def save_branding_settings():

    conn = None
    cursor = None

    try:

        primary_color = (
            request.form.get("primary_color") or "#0EA5A4"
        ).strip()

        secondary_color = (
            request.form.get("secondary_color") or "#14B8A6"
        ).strip()

        button_color = (
            request.form.get("button_color") or "#38BDF8"
        ).strip()

        erp_version = (
            request.form.get("erp_version") or ""
        ).strip()

        footer_text = (
            request.form.get("footer_text") or ""
        ).strip()

        powered_by = (
            request.form.get("powered_by") or ""
        ).strip()

        website_url = (
            request.form.get("website_url") or ""
        ).strip()

        # ================= COLOR VALIDATION =================
        color_pattern = r"^#[0-9A-Fa-f]{6}$"

        for color in [
            primary_color,
            secondary_color,
            button_color
        ]:
            if not re.fullmatch(color_pattern, color):
                return "Invalid color code ❌"

        # ================= LENGTH VALIDATION =================
        if len(erp_version) > 50:
            return "ERP version too long ❌"

        if len(footer_text) > 255:
            return "Footer text too long ❌"

        if len(powered_by) > 150:
            return "Powered by text too long ❌"

        if len(website_url) > 255:
            return "Website URL too long ❌"

        # ================= URL VALIDATION =================
        if website_url:
            if not (
                website_url.startswith("http://")
                or website_url.startswith("https://")
            ):
                return "Invalid website URL ❌"

        conn = get_connection()

        if not conn:
            return "Database connection failed ❌"

        cursor = conn.cursor()

        # ================= GET SETTINGS =================
        cursor.execute("""
            SELECT
                id,
                favicon
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        settings = cursor.fetchone()

        if not settings:
            return "System settings not found ❌"

        settings_id = settings[0]
        old_favicon = settings[1]

        # ================= FAVICON UPLOAD =================
        favicon_filename = None
        favicon = request.files.get("favicon")

        if favicon and favicon.filename:

            favicon.seek(0, os.SEEK_END)
            size = favicon.tell()
            favicon.seek(0)

            if size > 2 * 1024 * 1024:
                return "Favicon size exceeds 2MB ❌"

            original_filename = secure_filename(
                favicon.filename
            )

            allowed_extensions = (
                ".png",
                ".jpg",
                ".jpeg",
                ".ico",
                ".webp"
            )

            if not original_filename.lower().endswith(
                allowed_extensions
            ):
                return "Invalid favicon type ❌"

            filename = (
                f"{int(datetime.now().timestamp())}_"
                + original_filename
            )

            upload_folder = os.path.join(
                "static",
                "uploads",
                "branding"
            )

            os.makedirs(
                upload_folder,
                exist_ok=True
            )

            favicon_path = os.path.join(
                upload_folder,
                filename
            )

            favicon.save(favicon_path)

            favicon_filename = (
                "uploads/branding/" + filename
            )

        # ================= UPDATE SETTINGS =================
        if favicon_filename:

            cursor.execute("""
                UPDATE system_settings
                SET
                    primary_color = %s,
                    secondary_color = %s,
                    button_color = %s,
                    erp_version = %s,
                    footer_text = %s,
                    powered_by = %s,
                    website_url = %s,
                    favicon = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (
                primary_color,
                secondary_color,
                button_color,
                erp_version,
                footer_text,
                powered_by,
                website_url,
                favicon_filename,
                settings_id
            ))

        else:

            cursor.execute("""
                UPDATE system_settings
                SET
                    primary_color = %s,
                    secondary_color = %s,
                    button_color = %s,
                    erp_version = %s,
                    footer_text = %s,
                    powered_by = %s,
                    website_url = %s,
                    updated_at = NOW()
                WHERE id = %s
            """, (
                primary_color,
                secondary_color,
                button_color,
                erp_version,
                footer_text,
                powered_by,
                website_url,
                settings_id
            ))

        conn.commit()

        # ================= DELETE OLD FAVICON AFTER SUCCESS =================
        if favicon_filename and old_favicon:

            old_path = os.path.join(
                "static",
                old_favicon
            )

            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception as delete_error:
                    print(
                        "⚠️ OLD FAVICON DELETE ERROR:",
                        delete_error
                    )

        print("✅ BRANDING SETTINGS UPDATED")

        return redirect(
            url_for("superadmin_settings") + "#branding"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ BRANDING SETTINGS ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


#_________________________________________________________________________________________         

# =========================================================
# 💰 PAYMENT PAGE (PRODUCTION READY)
# =========================================================

@app.route(
    "/clerk/subscription/payment",
    methods=["POST"]
)
@login_required
def subscription_payment():

    conn = None
    cursor = None

    try:

        # =========================================
        # CLERK VALIDATION
        # =========================================

        if session.get("clerk_role") != "clerk":
            return "Unauthorized ❌"

        school_id = session.get(
            "clerk_school_id"
        )

        if not school_id:
            return "School session missing ❌"

        # =========================================
        # FORM DATA
        # =========================================

        subscription_id = (
            request.form.get(
                "subscription_id",
                ""
            ).strip()
        )

        plan_id = (
            request.form.get(
                "plan_id",
                ""
            ).strip()
        )

        if not subscription_id:
            return "Subscription missing ❌"

        if not plan_id:
            return "Plan missing ❌"

        if not subscription_id.isdigit():
            return "Invalid subscription ❌"

        if not plan_id.isdigit():
            return "Invalid plan ❌"

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()

        if not conn:
            return "Database connection failed ❌"

        cursor = conn.cursor()

        # =========================================
        # VERIFY SUBSCRIPTION BELONGS TO SCHOOL
        # =========================================

        cursor.execute("""
            SELECT
                id
            FROM subscriptions
            WHERE id = %s
            AND school_id = %s
        """, (
            int(subscription_id),
            school_id
        ))

        subscription = cursor.fetchone()

        if not subscription:
            return "Invalid subscription ❌"

        # =========================================
        # GET ACTIVE PLAN
        # =========================================

        cursor.execute("""
            SELECT
                id,
                plan_name,
                monthly_price,
                yearly_price,
                duration_months
            FROM subscription_plans
            WHERE id = %s
            AND is_active = 1
        """, (
            int(plan_id),
        ))

        plan = cursor.fetchone()

        if not plan:
            return "Plan not found ❌"

        # =========================================
        # AMOUNT VALIDATION
        # =========================================

        if (
            plan[2] is None
            or float(plan[2]) <= 0
        ):
            return "Invalid plan amount ❌"

        amount = float(plan[2])

        # =========================================
        # CREATE RAZORPAY CLIENT
        # =========================================

        client = razorpay.Client(
            auth=(
                RAZORPAY_KEY_ID,
                RAZORPAY_KEY_SECRET
            )
        )

        amount_in_paise = int(
            amount * 100
        )

        # =========================================
        # CREATE ORDER
        # =========================================

        razorpay_order = client.order.create({

            "amount": amount_in_paise,

            "currency": "INR",

            "payment_capture": 1

        })

        if not razorpay_order:
            return "Unable to create payment order ❌"

        # =========================================
        # LOAD PAYMENT PAGE
        # =========================================

        return render_template(

            "subscription/payment.html",

            subscription_id=int(
                subscription_id
            ),

            plan_id=plan[0],

            plan_name=plan[1],

            amount=amount,

            duration_months=plan[4],

            razorpay_order_id=(
                razorpay_order["id"]
            ),

            razorpay_key=(
                RAZORPAY_KEY_ID
            )

        )

    except Exception as e:

        print(
            "❌ PAYMENT PAGE ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# ✅ PAYMENT SUCCESS (PRODUCTION READY)
# =========================================================
@app.route(
    "/clerk/subscription/payment-success",
    methods=["POST"]
)
@login_required
def payment_success():

    conn = None
    cursor = None

    try:

        # =========================================
        # SESSION VALIDATION
        # =========================================

        if session.get("clerk_role") != "clerk":
            return "Unauthorized ❌"

        school_id = session.get(
            "clerk_school_id"
        )

        if not school_id:
            return "School session missing ❌"

        # =========================================
        # FORM DATA
        # =========================================

        plan_id = (
            request.form.get(
                "plan_id",
                ""
            ).strip()
        )

        subscription_id = (
            request.form.get(
                "subscription_id",
                ""
            ).strip()
        )

        razorpay_payment_id = (
            request.form.get(
                "razorpay_payment_id",
                ""
            ).strip()
        )

        razorpay_order_id = (
            request.form.get(
                "razorpay_order_id",
                ""
            ).strip()
        )

        razorpay_signature = (
            request.form.get(
                "razorpay_signature",
                ""
            ).strip()
        )

        # =========================================
        # VALIDATION
        # =========================================

        if not plan_id:
            return "Plan missing ❌"

        if not subscription_id:
            return "Subscription missing ❌"

        if not razorpay_payment_id:
            return "Payment ID missing ❌"

        if not razorpay_order_id:
            return "Order ID missing ❌"

        if not razorpay_signature:
            return "Payment signature missing ❌"

        if not plan_id.isdigit():
            return "Invalid plan ❌"

        if not subscription_id.isdigit():
            return "Invalid subscription ❌"

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()

        if not conn:
            return "Database connection failed ❌"

        cursor = conn.cursor()

        # =========================================
        # VERIFY SUBSCRIPTION OWNERSHIP
        # =========================================

        cursor.execute("""
            SELECT id
            FROM subscriptions
            WHERE id = %s
            AND school_id = %s
        """, (
            int(subscription_id),
            school_id
        ))

        subscription = cursor.fetchone()

        if not subscription:
            return "Invalid subscription ❌"

        # =========================================
        # PREVENT DUPLICATE PAYMENT
        # =========================================

        cursor.execute("""
            SELECT id
            FROM payment_logs
            WHERE order_id = %s
        """, (
            razorpay_order_id,
        ))

        existing_order = cursor.fetchone()

        if existing_order:
            return "Payment already processed ✅"

        # =========================================
        # VERIFY RAZORPAY SIGNATURE
        # =========================================

        client = razorpay.Client(
            auth=(
                RAZORPAY_KEY_ID,
                RAZORPAY_KEY_SECRET
            )
        )

        verify_data = {

            "razorpay_order_id":
            razorpay_order_id,

            "razorpay_payment_id":
            razorpay_payment_id,

            "razorpay_signature":
            razorpay_signature

        }

        try:

            client.utility.verify_payment_signature(
                verify_data
            )

        except Exception as verify_error:

            cursor.execute("""
                INSERT INTO payment_logs
                (
                    school_id,
                    subscription_id,
                    plan_id,
                    amount,

                    payment_status,
                    payment_id,
                    order_id,

                    transaction_type,
                    payment_gateway,

                )
                VALUES
                (
                    %s,%s,%s,%s,
                    %s,%s,%s,
                    %s,%s
                )
            """, (

                school_id,
                subscription_id,
                plan_id,
                0,

                "failed",
                razorpay_payment_id,
                razorpay_order_id,

                "Subscription Renewal",
                "Razorpay",
                str(verify_error)

            ))

            conn.commit()

            return (
                f"Payment verification failed ❌ "
                f"{verify_error}"
            )

        # =========================================
        # GET ACTIVE PLAN
        # =========================================

        cursor.execute("""
            SELECT

                plan_name,
                monthly_price,
                duration_months

            FROM subscription_plans

            WHERE id = %s
            AND is_active = 1
        """, (
            int(plan_id),
        ))

        plan = cursor.fetchone()

        if not plan:
            return "Plan not found ❌"

        if plan[2] is None:
            return "Plan duration missing ❌"

        amount_paid = float(plan[1])

        if amount_paid <= 0:
            return "Invalid plan amount ❌"

        # =========================================
        # GET SCHOOL DETAILS
        # =========================================

        cursor.execute("""
            SELECT
                name,
                email
            FROM schools
            WHERE school_id = %s
        """, (
            school_id,
        ))

        school = cursor.fetchone()

        if not school:
            return "School not found ❌"

        school_name = school[0]
        school_email = school[1]

        # =========================================
        # UPDATE SUBSCRIPTION
        # =========================================

        cursor.execute("""
            UPDATE subscriptions
            SET

                plan_id = %s,
                plan_name = %s,
                amount = %s,

                start_date = NOW(),

                end_date = DATE_ADD(
                    NOW(),
                    INTERVAL %s MONTH
                ),

                status = 'active'

            WHERE school_id = %s
            AND id = %s
        """, (

            int(plan_id),
            plan[0],
            amount_paid,

            int(plan[2]),

            school_id,
            int(subscription_id)

        ))

        if cursor.rowcount == 0:
            return "Subscription update failed ❌"


        # =========================================
        # GENERATE INVOICE NUMBER
        # =========================================

        invoice_number = (
            f"INV-{datetime.now().year}-{razorpay_payment_id[-8:]}"
        )    

        # =========================================
        # PAYMENT LOG
        # =========================================

        cursor.execute("""
            INSERT INTO payment_logs
            (

                school_id,
                subscription_id,
                plan_id,
                amount,

                payment_status,
                payment_id,
                order_id,
                       
                invoice_number,

                transaction_type,
                payment_gateway

            )

            VALUES
            (
                %s,%s,%s,%s,
                %s,%s,%s,
                %s,
                %s,%s
            )
        """, (

            school_id,
            subscription_id,
            plan_id,
            amount_paid,

            "success",
            razorpay_payment_id,
            razorpay_order_id,

            invoice_number,

            "Subscription Renewal",
            "Razorpay"

        ))

        # =========================================
        # SAVE
        # =========================================

        conn.commit()

        # =========================================
        # APPLY PLAN FEATURES
        # =========================================

        apply_plan_features(
            school_id,
            int(plan_id)
            )

        # =========================================
        # EMAIL RECEIPT
        # =========================================

        try:

            subject = (
                "Payment Successful - "
                "SPL ShalaSarthi ERP"
            )

            body = f"""
            <div style="font-family:Arial;padding:20px;">

                <h2 style="color:#10b981;">
                    Payment Successful
                </h2>

                <p>Dear {school_name},</p>

                <p>
                    Your subscription payment
                    was completed successfully.
                </p>

                <hr>

                <p><b>Plan:</b> {plan[0]}</p>

                <p><b>Amount Paid:</b>
                Rs. {amount_paid}</p>

                <p><b>Payment ID:</b>
                {razorpay_payment_id}</p>

                <p><b>Order ID:</b>
                {razorpay_order_id}</p>

                <p><b>Duration:</b>
                {plan[2]} Month(s)</p>

                <p><b>Status:</b> Success</p>

                <hr>

                <p>
                    Thank you for renewing
                    your SPL ShalaSarthi ERP subscription.
                </p>

            </div>
            """

            send_email(
                school_email,
                subject,
                body
            )

            print(
                "✅ PAYMENT RECEIPT EMAIL SENT"
            )

        except Exception as email_error:

            print(
                "❌ PAYMENT EMAIL ERROR:",
                email_error
            )

        # =========================================
        # REDIRECT
        # =========================================

        return redirect(
            url_for(
                "subscription_success",
                plan=plan[0],
                amount=amount_paid
            )
        )
    
    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "❌ PAYMENT VERIFY ERROR:",
            e
        )

        return (
            f"Payment verification failed ❌ {e}"
        )

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 # =========================================================
# 📜 SUBSCRIPTION HISTORY
# =========================================================

@app.route("/clerk/subscription/history")
@login_required
def subscription_history():

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    school_id = session.get(
        "clerk_school_id"
    )

    if not school_id:
        return "School session missing ❌"

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor(
            dictionary=True
        )

        # =========================================
        # PAYMENT HISTORY
        # =========================================

        cursor.execute("""

            SELECT

                pl.id,

                pl.invoice_number,

                pl.amount,

                pl.payment_status,

                pl.payment_id,

                pl.order_id,

                pl.payment_gateway,

                pl.transaction_type,

                pl.created_at,

                sp.plan_name

            FROM payment_logs pl

            LEFT JOIN subscription_plans sp
                ON pl.plan_id = sp.id

            WHERE pl.school_id = %s

            ORDER BY pl.id DESC

        """, (

            school_id,

        ))

        history = cursor.fetchall()

        # =========================================
        # TOTAL PAYMENTS
        # =========================================

        cursor.execute("""

            SELECT
                COUNT(*) AS total_transactions,

                COALESCE(
                    SUM(amount),
                    0
                ) AS total_amount

            FROM payment_logs

            WHERE school_id = %s

            AND payment_status = 'success'

        """, (

            school_id,

        ))

        summary = cursor.fetchone()


        
        # ================= SCHOOL DETAILS =================
        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        # =========================================
        # RENDER
        # =========================================

        return render_template(

            "subscription/history.html",

            history=history,

            summary=summary,

            role="clerk",

            school_name=school["school_name"],
            

            active_page="subscription_history"

        )

    except Exception as e:

        print(
            "❌ SUBSCRIPTION HISTORY ERROR:",
            e
        )

        return "History page failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📄 VIEW INVOICE
# =========================================================

@app.route(
    "/clerk/subscription/invoice/<int:payment_log_id>"
)
@login_required
def view_invoice(payment_log_id):

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    school_id = session.get(
        "clerk_school_id"
    )

    if not school_id:
        return "School session missing ❌"

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor(
            dictionary=True
        )

        # =========================================
        # GET INVOICE DATA
        # =========================================

        cursor.execute("""

            SELECT

                pl.id,
                pl.invoice_number,
                pl.amount,
                pl.payment_status,
                pl.payment_id,
                pl.order_id,
                pl.payment_gateway,
                pl.created_at,

                sp.plan_name,

                s.name AS school_name,
                s.email AS school_email,
                s.address

            FROM payment_logs pl

            LEFT JOIN subscription_plans sp
                ON pl.plan_id = sp.id

            LEFT JOIN schools s
                ON pl.school_id = s.school_id

            WHERE pl.id = %s
            AND pl.school_id = %s

            LIMIT 1

        """, (

            payment_log_id,
            school_id

        ))

        invoice = cursor.fetchone()

        if not invoice:
            return "Invoice not found ❌"

        return render_template(

            "subscription/invoice.html",

            invoice=invoice

        )

    except Exception as e:

        print(
            "❌ INVOICE ERROR:",
            e
        )

        return "Invoice loading failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📄 DOWNLOAD INVOICE PDF
# =========================================================

@app.route(
    "/clerk/subscription/invoice/pdf/<int:payment_log_id>"
)
@login_required
def download_invoice_pdf(payment_log_id):

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    school_id = session.get(
        "clerk_school_id"
    )

    conn = None
    cursor = None

    try:

        conn = get_connection()

        cursor = conn.cursor(
            dictionary=True
        )

        cursor.execute("""

            SELECT

                pl.id,
                pl.invoice_number,
                pl.amount,
                pl.payment_status,
                pl.payment_id,
                pl.order_id,
                pl.payment_gateway,
                pl.created_at,

                sp.plan_name,

                s.name AS school_name,
                s.email AS school_email,
                s.address

            FROM payment_logs pl

            LEFT JOIN subscription_plans sp
                ON pl.plan_id = sp.id

            LEFT JOIN schools s
                ON pl.school_id = s.school_id

            WHERE pl.id = %s
            AND pl.school_id = %s

            LIMIT 1

        """, (

            payment_log_id,
            school_id

        ))

        invoice = cursor.fetchone()

        if not invoice:
            return "Invoice not found ❌"

        # =====================================
        # RENDER HTML
        # =====================================

        rendered = render_template(

            "subscription/invoice.html",

            invoice=invoice,

            is_pdf=True

        )

        # =====================================
        # PDF OPTIONS
        # =====================================
        options = {

            "page-size": "A4",

            "encoding": "UTF-8",

            "enable-local-file-access": "",

            "print-media-type": "",

            "dpi": 300,

            "image-quality": 100,

            "margin-top": "0mm",
            "margin-right": "0mm",
            "margin-bottom": "0mm",
            "margin-left": "0mm",

            "disable-smart-shrinking": ""
        }

        pdf = pdfkit.from_string(

            rendered,

            False,

            configuration=pdf_config,

            options=options

        )

        response = make_response(pdf)

        response.headers[
            "Content-Type"
        ] = "application/pdf"

        response.headers[
            "Content-Disposition"
        ] = f'attachment; filename="{invoice["invoice_number"]}.pdf"'

        return response

    except Exception as e:

        print(
            "❌ INVOICE PDF ERROR:",
            e
        )

        return "PDF generation failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

 # ==================================================
# SUBSCRIPTION SUCCESS PAGE
# ==================================================

@app.route("/clerk/subscription/success")
@login_required
def subscription_success():

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    plan = request.args.get(
        "plan",
        "Subscription"
    )

    amount = request.args.get(
        "amount",
        "0"
    )

    return render_template(
        "subscription/success.html",
        plan=plan,
        amount=amount
    )

 # ==================================================
# PAYMENT FAILED PAGE
# ==================================================

@app.route("/clerk/subscription/payment-failed")
@login_required
def payment_failed():

    if session.get("clerk_role") != "clerk":
        return "Unauthorized ❌"

    return render_template(
        "subscription/payment_failed.html"
    )

# =============================================================================================



# =========================================================
# 🏫 CLERK DASHBOARD
# Real-time clerk dashboard with KPIs, charts, profile,
# subscription alert, reminder email and recent activities.
# =========================================================
@app.route("/clerk/dashboard")
@login_required
def clerk_dashboard():

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        school_id = session.get("clerk_school_id")
        clerk_user_id = session.get("clerk_user_id")

        # =====================================================
        # SESSION SAFETY
        # =====================================================
        if not school_id or not clerk_user_id:
            return redirect(url_for("login"))

        # =====================================================
        # DEFAULT SUBSCRIPTION VALUES
        # Prevents variable errors if subscription is missing
        # =====================================================
        subscription_id = None
        end_date = None
        current_status = "none"
        subscription_alert = None

        # =====================================================
        # LATEST SUBSCRIPTION
        # =====================================================
        cursor.execute("""
            SELECT
                id,
                end_date,
                status
            FROM subscriptions
            WHERE school_id = %s
            ORDER BY id DESC
            LIMIT 1
        """, (school_id,))

        sub = cursor.fetchone()

        if sub:

            subscription_id = sub[0]
            end_date = sub[1]
            current_status = sub[2] or "none"

            if end_date and isinstance(end_date, datetime):
                end_date = end_date.date()

            # =================================================
            # GET GLOBAL GRACE PERIOD
            # =================================================
            cursor.execute("""
                SELECT COALESCE(grace_period, 0)
                FROM system_settings
                ORDER BY id ASC
                LIMIT 1
            """)

            settings_row = cursor.fetchone()
            grace_days = int(settings_row[0] or 0) if settings_row else 0

            grace_end_date = None

            if end_date:
                grace_end_date = end_date + timedelta(days=grace_days)

            today_date = datetime.now().date()

            # =================================================
            # AUTO EXPIRE AFTER GRACE PERIOD
            # =================================================
            if (
                grace_end_date
                and today_date > grace_end_date
                and current_status == "active"
            ):

                cursor.execute("""
                    UPDATE subscriptions
                    SET status = 'expired'
                    WHERE id = %s
                """, (subscription_id,))

                conn.commit()

                current_status = "expired"

            # =================================================
            # SUBSCRIPTION REMINDER EMAIL
            # Sends only once for 7, 3, 1 and expiry day
            # =================================================
            if end_date:

                remaining_days_for_email = (end_date - today_date).days
                email_type = None

                if remaining_days_for_email == 7:
                    email_type = "7_days_warning"

                elif remaining_days_for_email == 3:
                    email_type = "3_days_warning"

                elif remaining_days_for_email == 1:
                    email_type = "1_day_warning"

                elif remaining_days_for_email == 0:
                    email_type = "expiry_today"

                if email_type and subscription_id:

                    cursor.execute("""
                        SELECT id
                        FROM subscription_email_logs
                        WHERE school_id = %s
                        AND subscription_id = %s
                        AND email_type = %s
                        LIMIT 1
                    """, (
                        school_id,
                        subscription_id,
                        email_type
                    ))

                    already_sent = cursor.fetchone()

                    if not already_sent:

                        cursor.execute("""
                            SELECT
                                name,
                                email
                            FROM schools
                            WHERE school_id = %s
                            LIMIT 1
                        """, (school_id,))

                        school_data = cursor.fetchone()

                        if school_data and school_data[1]:

                            school_name_email = school_data[0]
                            school_email = school_data[1]

                            subject = "Subscription Reminder - SPL ShalaSarthi ERP"

                            body = f"""
                            <div style="font-family:Arial;padding:20px;">
                                <h2 style="color:#f59e0b;">
                                    Subscription Expiry Reminder
                                </h2>

                                <p>Dear {school_name_email},</p>

                                <p>
                                    Your subscription will expire in
                                    <b>{remaining_days_for_email} day(s)</b>.
                                </p>

                                <p>
                                    Please renew your subscription to continue using ERP services.
                                </p>

                                <hr>

                                <p>
                                    Thank you,<br>
                                    <b>SPL ShalaSarthi ERP Team</b>
                                </p>
                            </div>
                            """

                            email_sent = send_email(
                                school_email,
                                subject,
                                body
                            )

                            if email_sent:

                                cursor.execute("""
                                    INSERT INTO subscription_email_logs
                                    (
                                        school_id,
                                        subscription_id,
                                        email_type,
                                        sent_at
                                    )
                                    VALUES (%s, %s, %s, NOW())
                                """, (
                                    school_id,
                                    subscription_id,
                                    email_type
                                ))

                                conn.commit()

        # =====================================================
        # SCHOOL DATA
        # =====================================================
        cursor.execute("""
            SELECT name
            FROM schools
            WHERE school_id = %s
            LIMIT 1
        """, (school_id,))

        school = cursor.fetchone()

        if not school:
            return "School not found ❌"

        school_name = school[0]

        # =====================================================
        # KPI COUNTS
        # =====================================================
        cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE school_id = %s
        """, (school_id,))

        total_students = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*)
            FROM tc
            WHERE school_id = %s
        """, (school_id,))

        total_tc = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT COUNT(*)
            FROM bonafide
            WHERE school_id = %s
        """, (school_id,))

        total_bonafide = cursor.fetchone()[0] or 0


        # =====================================================
        # TODAY'S SUMMARY
        # Shows today's clerk workload
        # =====================================================

        today = date.today()
        current_date = datetime.now().strftime("%d %b %Y")

        # Today new students
        cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE school_id = %s
            AND DATE(created_at) = %s
        """, (
            school_id,
            today
        ))

        today_students = cursor.fetchone()[0] or 0


        # Today TC generated
        cursor.execute("""
            SELECT COUNT(*)
            FROM tc
            WHERE school_id = %s
            AND DATE(created_at) = %s
        """, (
            school_id,
            today
        ))

        today_tc = cursor.fetchone()[0] or 0


        # Today Bonafide issued
        cursor.execute("""
            SELECT COUNT(*)
            FROM bonafide
            WHERE school_id = %s
            AND DATE(created_at) = %s
        """, (
            school_id,
            today
        ))

        today_bonafide = cursor.fetchone()[0] or 0


        # =====================================================
        # PENDING TASKS
        # Counts students with incomplete important profile data
        # Based on add_student form fields
        # =====================================================

        cursor.execute("""
            SELECT COUNT(*)
            FROM students
            WHERE school_id = %s
            AND (
                school_register_no IS NULL OR school_register_no = ''
                OR name IS NULL OR name = ''
                OR father_name IS NULL OR father_name = ''
                OR mother_name IS NULL OR mother_name = ''
                OR aadhaar IS NULL OR aadhaar = ''
                OR dob IS NULL
                OR birth_place IS NULL OR birth_place = ''
                OR nationality IS NULL OR nationality = ''
                OR mother_tongue IS NULL OR mother_tongue = ''
                OR religion IS NULL OR religion = ''
                OR caste IS NULL OR caste = ''
                OR city IS NULL OR city = ''
                OR taluka IS NULL OR taluka = ''
                OR district IS NULL OR district = ''
                OR state IS NULL OR state = ''
                OR admission_no IS NULL OR admission_no = ''
                OR admission_date IS NULL
                OR class IS NULL OR class = ''
                OR section IS NULL OR section = ''
                OR primary_mobile IS NULL OR primary_mobile = ''
            )
        """, (school_id,))

        pending_tasks = cursor.fetchone()[0] or 0

        # =========================================
        # TOTAL MISSING FIELDS COUNT
        # =========================================

        cursor.execute("""
        SELECT
        (
            SUM(CASE WHEN school_register_no IS NULL OR school_register_no='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN father_name IS NULL OR father_name='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN mother_name IS NULL OR mother_name='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN aadhaar IS NULL OR aadhaar='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN dob IS NULL THEN 1 ELSE 0 END) +
            SUM(CASE WHEN birth_place IS NULL OR birth_place='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN nationality IS NULL OR nationality='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN mother_tongue IS NULL OR mother_tongue='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN religion IS NULL OR religion='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN caste IS NULL OR caste='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN city IS NULL OR city='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN taluka IS NULL OR taluka='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN district IS NULL OR district='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN state IS NULL OR state='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN admission_date IS NULL THEN 1 ELSE 0 END) +
            SUM(CASE WHEN class IS NULL OR class='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN section IS NULL OR section='' THEN 1 ELSE 0 END) +
            SUM(CASE WHEN primary_mobile IS NULL OR primary_mobile='' THEN 1 ELSE 0 END)
        )
        FROM students
        WHERE school_id = %s
        """, (school_id,))

        pending_fields = cursor.fetchone()[0] or 0


        # =====================================================
        # PENDING STUDENT DETAILS LIST
        # Shows all missing important fields per student
        # =====================================================

        cursor.execute("""
            SELECT
                id,
                admission_no,
                name,
                class,
                section,

                CONCAT_WS(', ',

                    CASE WHEN school_register_no IS NULL OR school_register_no = '' THEN 'Register No' END,
                    CASE WHEN name IS NULL OR name = '' THEN 'Name' END,
                    CASE WHEN father_name IS NULL OR father_name = '' THEN 'Father Name' END,
                    CASE WHEN mother_name IS NULL OR mother_name = '' THEN 'Mother Name' END,
                    CASE WHEN aadhaar IS NULL OR aadhaar = '' THEN 'Aadhaar' END,
                    CASE WHEN dob IS NULL THEN 'DOB' END,
                    CASE WHEN birth_place IS NULL OR birth_place = '' THEN 'Birth Place' END,
                    CASE WHEN nationality IS NULL OR nationality = '' THEN 'Nationality' END,
                    CASE WHEN mother_tongue IS NULL OR mother_tongue = '' THEN 'Mother Tongue' END,
                    CASE WHEN religion IS NULL OR religion = '' THEN 'Religion' END,
                    CASE WHEN caste IS NULL OR caste = '' THEN 'Caste' END,
                    CASE WHEN city IS NULL OR city = '' THEN 'City' END,
                    CASE WHEN taluka IS NULL OR taluka = '' THEN 'Taluka' END,
                    CASE WHEN district IS NULL OR district = '' THEN 'District' END,
                    CASE WHEN state IS NULL OR state = '' THEN 'State' END,
                    CASE WHEN admission_no IS NULL OR admission_no = '' THEN 'Admission No' END,
                    CASE WHEN admission_date IS NULL THEN 'Admission Date' END,
                    CASE WHEN class IS NULL OR class = '' THEN 'Class' END,
                    CASE WHEN section IS NULL OR section = '' THEN 'Section' END,
                    CASE WHEN primary_mobile IS NULL OR primary_mobile = '' THEN 'Mobile' END

                ) AS pending_reason

            FROM students
            WHERE school_id = %s
            AND (
                school_register_no IS NULL OR school_register_no = ''
                OR name IS NULL OR name = ''
                OR father_name IS NULL OR father_name = ''
                OR mother_name IS NULL OR mother_name = ''
                OR aadhaar IS NULL OR aadhaar = ''
                OR dob IS NULL
                OR birth_place IS NULL OR birth_place = ''
                OR nationality IS NULL OR nationality = ''
                OR mother_tongue IS NULL OR mother_tongue = ''
                OR religion IS NULL OR religion = ''
                OR caste IS NULL OR caste = ''
                OR city IS NULL OR city = ''
                OR taluka IS NULL OR taluka = ''
                OR district IS NULL OR district = ''
                OR state IS NULL OR state = ''
                OR admission_no IS NULL OR admission_no = ''
                OR admission_date IS NULL
                OR class IS NULL OR class = ''
                OR section IS NULL OR section = ''
                OR primary_mobile IS NULL OR primary_mobile = ''
            )
            ORDER BY id DESC
            LIMIT 6
        """, (school_id,))

        pending_students = cursor.fetchall()


      # =====================================================
      # STUDENT GROWTH CHART DATA - LAST 5 MONTHS
      # =====================================================

        cursor.execute("""
            SELECT
                DATE_FORMAT(created_at, '%b') AS month_name,
                COUNT(*) AS total
            FROM students
            WHERE school_id = %s
            GROUP BY
                YEAR(created_at),
                MONTH(created_at),
                DATE_FORMAT(created_at, '%b')
            ORDER BY
                YEAR(created_at) DESC,
                MONTH(created_at) DESC
            LIMIT 5
        """, (school_id,))

        rows = cursor.fetchall()[::-1]

        growth_labels = [row[0] for row in rows]
        monthly_counts = [row[1] for row in rows]

        while len(growth_labels) < 5:
            growth_labels.insert(0, "-")
            monthly_counts.insert(0, 0)

        growth_data = []
        running_total = 0

        for count in monthly_counts:
            running_total += count
            growth_data.append(running_total)

        # =====================================================
        # TC CHART DATA - LAST 5 ACTIVE DAYS
        # =====================================================

        cursor.execute("""
            SELECT
                DATE_FORMAT(tc_date, '%d %b') AS tc_day,
                COUNT(*) AS total
            FROM tc
            WHERE school_id = %s
            GROUP BY DATE(tc_date), DATE_FORMAT(tc_date, '%d %b')
            ORDER BY DATE(tc_date) DESC
            LIMIT 5
        """, (school_id,))

        rows = cursor.fetchall()[::-1]

        tc_labels = [row[0] for row in rows]
        tc_data = [row[1] for row in rows]

        while len(tc_labels) < 5:
            tc_labels.insert(0, "-")
            tc_data.insert(0, 0)

        # =====================================================
        # BONAFIDE DOUGHNUT DATA
        # =====================================================

        bon_count = total_bonafide or 0

        remaining_students = max(
            0,
            (total_students or 0) - bon_count
        )

        bonafide_data = [
            bon_count,
            remaining_students
        ]

        if bonafide_data == [0, 0]:
            bonafide_data = [0, 1]

        # =====================================================
        # GROWTH PERCENT
        # =====================================================

        if len(growth_data) >= 2 and growth_data[-2] > 0:

            growth_percent = round(
                ((growth_data[-1] - growth_data[-2]) / growth_data[-2]) * 100
            )

        elif total_students > 0:

            growth_percent = 100

        else:

            growth_percent = 0

        growth_percent = max(
            0,
            min(growth_percent, 100)
        )

        # =====================================================
        # USER PROFILE
        # =====================================================
        cursor.execute("""
            SELECT
                u.id,
                u.name,
                u.email,
                u.phone,
                u.address,
                u.designation,
                s.name AS school_name,
                sub.plan_name,
                CASE
                    WHEN sub.end_date IS NULL THEN 0
                    WHEN DATEDIFF(sub.end_date, NOW()) < 0 THEN 0
                    ELSE DATEDIFF(sub.end_date, NOW())
                END AS remaining_days
            FROM users u
            JOIN schools s
                ON u.school_id = s.school_id
            LEFT JOIN (
                SELECT
                    s1.school_id,
                    s1.plan_name,
                    s1.end_date
                FROM subscriptions s1
                INNER JOIN (
                    SELECT
                        school_id,
                        MAX(id) AS latest_id
                    FROM subscriptions
                    GROUP BY school_id
                ) s2
                    ON s1.id = s2.latest_id
            ) sub
                ON sub.school_id = s.school_id
            WHERE u.id = %s
            LIMIT 1
        """, (clerk_user_id,))

        row = cursor.fetchone()

        if row:

            columns = [
                col[0]
                for col in cursor.description
            ]

            user_profile = dict(
                zip(columns, row)
            )

        else:

            user_profile = {}

        # =====================================================
        # SUBSCRIPTION ALERT SYSTEM
        # =====================================================
        remaining_days = int(
            user_profile.get("remaining_days") or 0
        )

        if current_status == "expired":

            subscription_alert = {
                "type": "expired",
                "title": "Subscription Expired",
                "message": (
                    "Your subscription has expired. "
                    "Renew now to continue ERP services."
                )
            }

        elif current_status == "none":

            subscription_alert = {
                "type": "danger",
                "title": "No Subscription Found",
                "message": (
                    "No active subscription is linked with this school. "
                    "Please contact administrator."
                )
            }

        elif remaining_days <= 3:

            subscription_alert = {
                "type": "danger",
                "title": "Urgent Renewal Required",
                "message": (
                    f"Your subscription expires in {remaining_days} day(s). "
                    f"Renew immediately."
                )
            }

        elif remaining_days <= 7:

            subscription_alert = {
                "type": "warning",
                "title": "Subscription Expiring Soon",
                "message": (
                    f"Your subscription expires in {remaining_days} day(s). "
                    f"Please renew before expiry."
                )
            }

        # =====================================================
        # RECENT ACTIVITIES
        # Store real activity date for proper sorting
        # =====================================================
        activities = []

        # ================= STUDENTS =================
        cursor.execute("""
            SELECT
                name,
                created_at
            FROM students
            WHERE school_id = %s
            ORDER BY created_at DESC
            LIMIT 3
        """, (school_id,))

        for s in cursor.fetchall():

            activities.append({
                "type": "teal",
                "title": f"New student admission: {s[0]}",
                "title_mr": "नवीन विद्यार्थी प्रवेश",
                "time": "Recently",
                "sort_date": normalize_datetime(s[1])
            })

        # ================= TC =================
        cursor.execute("""
            SELECT
                tc.tc_number,
                st.name,
                tc.tc_date
            FROM tc
            JOIN students st
                ON tc.student_id = st.id
            WHERE tc.school_id = %s
            ORDER BY tc.tc_date DESC
            LIMIT 3
        """, (school_id,))

        for tc in cursor.fetchall():

            activities.append({
                "type": "orange",
                "title": f"TC issued for {tc[1]} (TC No: {tc[0]})",
                "title_mr": "टीसी जारी केले",
                "time": "Recently",
                "sort_date": normalize_datetime(tc[2])
            })

        # ================= BONAFIDE =================
        cursor.execute("""
            SELECT
                st.name,
                b.date
            FROM bonafide b
            JOIN students st
                ON b.student_id = st.id
            WHERE b.school_id = %s
            ORDER BY b.date DESC
            LIMIT 3
        """, (school_id,))

        for b in cursor.fetchall():

            activities.append({
                "type": "blue",
                "title": f"Bonafide certificate issued for {b[0]}",
                "title_mr": "बोनाफाईड प्रमाणपत्र जारी केले",
                "time": "Recently",
                "sort_date": normalize_datetime(b[1])
            })

        activities = sorted(
            activities,
            key=lambda x: x["sort_date"],
            reverse=True
        )[:6]

        # Remove internal sorting key before sending to template
        for activity in activities:
            activity.pop("sort_date", None)

        
        # =====================================================
        # GLOBAL FEATURE SETTINGS
        # Used to lock/hide dashboard action cards
        # =====================================================
        cursor.execute("""
            SELECT
                enable_tc_management,
                enable_bonafide_management,
                enable_import_export,
                enable_attendance,
                enable_fee_management,
                enable_teacher_management,
                enable_results,
                enable_timetable,
                enable_notice_board
            FROM system_settings
            ORDER BY id ASC
            LIMIT 1
        """)

        feature_row = cursor.fetchone()

        features = {
            "tc": feature_row[0] if feature_row else "Disabled",
            "bonafide": feature_row[1] if feature_row else "Disabled",
            "import_export": feature_row[2] if feature_row else "Disabled",
            "attendance": feature_row[3] if feature_row else "Disabled",
            "fee": feature_row[4] if feature_row else "Disabled",
            "teacher": feature_row[5] if feature_row else "Disabled",
            "results": feature_row[6] if feature_row else "Disabled",
            "timetable": feature_row[7] if feature_row else "Disabled",
            "notice": feature_row[8] if feature_row else "Disabled"
        }

        # =====================================================
        # FINAL RENDER
        # =====================================================
        return render_template(
            "dashboard/clerk.html",
            role="clerk",
            school_name=school_name,
            active_page="dashboard",
            user_profile=user_profile,
            current_date=current_date,
            total_students=total_students,
            total_tc=total_tc,
            total_bonafide=total_bonafide,
            today_students=today_students,
            today_tc=today_tc,
            today_bonafide=today_bonafide,
            pending_tasks=pending_tasks,
            pending_fields=pending_fields,
            pending_students=pending_students,
            growth_labels=growth_labels,
            growth_data=growth_data,
            tc_labels=tc_labels,
            tc_data=tc_data,
            bonafide_data=bonafide_data,
            growth_percent=growth_percent,
            activities=activities,
            subscription_alert=subscription_alert,
            features=features
        )

    except Exception as e:

        print("❌ DASHBOARD ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
            
# =========================================================
# UPDATE USER PROFILE
# =========================================================

@app.route("/clerk/profile/update", methods=["POST"])
@login_required
def clerk_profile_update():

    conn = None
    cursor = None

    try:
        user_id = session.get("clerk_user_id")

        if not user_id:
            return {
                "status":"error",
                "message":"Session expired"
            }

        name = request.form.get("name","").strip()
        phone = request.form.get("phone","").strip()
        address = request.form.get("address","").strip()
        designation = request.form.get("designation","").strip()

        if not name:
            return {
                "status":"error",
                "message":"Name required"
            }
        
        if phone and (not phone.isdigit() or len(phone) != 10):
            return {
                "status":"error",
                "message":"Invalid phone number"
            }

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE users
            SET
                name=%s,
                phone=%s,
                       
                address=%s,
                designation=%s,
                       
                updated_at=NOW()
                       
            WHERE id=%s
            AND role='clerk'      
                           
        """,(
            name,
            phone,
            address,
            designation,
            user_id
        ))

        conn.commit()

        return {
            "status":"success",
            "message":"Profile updated successfully"
        }

    except Exception as e:
        
        print("PROFILE UPDATE ERROR:", e)
        return jsonify({
            "status": "error",
            "message": "Something went wrong"
        })

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# =========================================================
# 🔐 SEND PASSWORD RESET OTP
# PURPOSE:
# Send OTP to registered email
# Store OTP in DB
# OTP valid for 5 minutes
# Includes resend cooldown and failed-email cleanup
# =========================================================

@app.route(
    "/clerk/profile/send-otp",
    methods=["POST"]
)
@login_required
def clerk_send_password_otp():

    conn = None
    cursor = None

    try:

        # =========================================
        # GET REQUEST DATA
        # =========================================

        data = request.get_json() or {}

        email = (
            data.get("email") or ""
        ).strip()

        if not email:

            return jsonify({
                "status": "error",
                "message": "Email required"
            })

        # =========================================
        # SESSION USER
        # =========================================

        user_id = session.get(
            "clerk_user_id"
        )

        if not user_id:

            return jsonify({
                "status": "error",
                "message": "User session missing"
            })

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()

        if not conn:

            return jsonify({
                "status": "error",
                "message": "Database connection failed"
            })

        cursor = conn.cursor()

        # =========================================
        # CHECK USER EMAIL
        # =========================================

        cursor.execute("""
            SELECT
                email,
                name
            FROM users
            WHERE id = %s
            AND role = 'clerk'
            LIMIT 1
        """, (
            user_id,
        ))

        user = cursor.fetchone()

        if not user:

            return jsonify({
                "status": "error",
                "message": "User not found"
            })

        db_email = (
            user[0] or ""
        ).strip()

        user_name = (
            user[1] or "User"
        )

        # =========================================
        # SECURITY EMAIL MATCH
        # Email must match logged-in clerk email
        # =========================================

        if email.lower() != db_email.lower():

            return jsonify({
                "status": "error",
                "message": "Email does not match registered email"
            })

        # =========================================
        # OTP RESEND COOLDOWN
        # Prevent repeated OTP spam
        # =========================================

        cursor.execute("""
            SELECT
                created_at
            FROM password_reset_otp
            WHERE user_id = %s
            AND is_used = 0
            ORDER BY id DESC
            LIMIT 1
        """, (
            user_id,
        ))

        last_otp = cursor.fetchone()

        if last_otp and last_otp[0]:

            last_created = last_otp[0]

            if isinstance(last_created, date) and not isinstance(last_created, datetime):
                last_created = datetime.combine(
                    last_created,
                    datetime.min.time()
                )

            seconds_passed = (
                datetime.now() - last_created
            ).total_seconds()

            if seconds_passed < 60:

                return jsonify({
                    "status": "error",
                    "message": "Please wait 60 seconds before requesting another OTP"
                })

        # =========================================
        # REMOVE OLD UNUSED OTP
        # Keep only latest active OTP
        # =========================================

        cursor.execute("""
            DELETE FROM password_reset_otp
            WHERE user_id = %s
            AND is_used = 0
        """, (
            user_id,
        ))

        # =========================================
        # GENERATE OTP
        # =========================================

        otp = str(
            random.randint(
                100000,
                999999
            )
        )

        expiry_time = (
            datetime.now()
            + timedelta(minutes=5)
        )

        # =========================================
        # SAVE OTP IN DATABASE
        # =========================================

        cursor.execute("""
            INSERT INTO password_reset_otp (
                user_id,
                email,
                otp,
                expires_at,
                is_used,
                attempts,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, NOW())
        """, (
            user_id,
            email,
            otp,
            expiry_time,
            0,
            0
        ))

        conn.commit()

        # =========================================
        # TEMP SESSION SUPPORT
        # Used by password update step
        # =========================================

        session["otp_verified"] = False
        session.modified = True

        # =========================================
        # SEND OTP EMAIL
        # Uses global SMTP sender
        # =========================================

        email_body = f"""
        <p>Hello {user_name},</p>

        <p>Your OTP for password reset is:</p>

        <h2>{otp}</h2>

        <p>This OTP is valid for 5 minutes.</p>

        <p>Do not share this OTP with anyone.</p>

        <p>Regards,<br>SPL ShalaSarthi ERP</p>
        """

        email_sent = send_email(
            email,
            "Password Reset OTP",
            email_body
        )

        # =========================================
        # CLEAN OTP IF EMAIL FAILED
        # Prevent unused OTP remaining in DB
        # =========================================

        if not email_sent:

            cursor.execute("""
                DELETE FROM password_reset_otp
                WHERE user_id = %s
                AND otp = %s
                AND is_used = 0
            """, (
                user_id,
                otp
            ))

            conn.commit()

            return jsonify({
                "status": "error",
                "message": "Failed to send OTP email"
            })

        # =========================================
        # SUCCESS RESPONSE
        # =========================================

        return jsonify({
            "status": "success",
            "message": "OTP sent successfully"
        })

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "SEND OTP ERROR:",
            e
        )

        return jsonify({
            "status": "error",
            "message": "Something went wrong"
        })

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
 
# =========================================================
# ✅ VERIFY OTP
# PURPOSE:
# Verify OTP from database
# =========================================================

@app.route(
    "/clerk/profile/check-otp",
    methods=["POST"]
)
@login_required
def clerk_check_password_otp():

    conn = None
    cursor = None

    try:

        # =========================================
        # GET REQUEST DATA
        # =========================================

        data = request.get_json() or {}

        entered_otp = (
            data.get("otp") or ""
        ).strip()

        # =========================================
        # VALIDATION
        # =========================================

        if not entered_otp:

            return jsonify({

                "status": "error",
                "message": "OTP required"

            })

        if (
            not entered_otp.isdigit()
            or len(entered_otp) != 6
        ):

            return jsonify({

                "status": "error",
                "message": "Invalid OTP format"

            })

        # =========================================
        # SESSION USER
        # =========================================

        user_id = session.get(
            "clerk_user_id"
        )

        if not user_id:

            return jsonify({

                "status": "error",
                "message": "User session missing"

            })

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # GET LATEST UNUSED OTP
        # =========================================

        cursor.execute("""

            SELECT

                id,
                otp,
                expires_at,
                attempts,
                is_used

            FROM password_reset_otp

            WHERE user_id = %s
            AND is_used = 0

            ORDER BY id DESC
            LIMIT 1

        """, (user_id,))

        otp_row = cursor.fetchone()

        # =========================================
        # OTP NOT FOUND
        # =========================================

        if not otp_row:

            return jsonify({

                "status": "error",
                "message": "OTP not found"

            })

        otp_id = otp_row[0]
        saved_otp = otp_row[1]
        expiry_time = otp_row[2]
        attempts = otp_row[3] or 0
        is_used = otp_row[4]

        # =========================================
        # ALREADY USED
        # =========================================

        if is_used == 1:

            return jsonify({

                "status": "error",
                "message": "OTP already used"

            })
        # =========================================
        # BLOCK AFTER 5 ATTEMPTS
        # OTP becomes unusable
        # =========================================

        if attempts >= 5:

            cursor.execute("""

                UPDATE password_reset_otp

                SET is_used = 1

                WHERE id = %s

            """, (otp_id,))

            conn.commit()

            return jsonify({

                "status": "error",

                "message":
                "Too many invalid attempts. Please request a new OTP."

            })
        # =========================================
        # CHECK EXPIRY
        # =========================================

        if datetime.now() > expiry_time:

            return jsonify({

                "status": "error",
                "message": "OTP expired"

            })

        # =========================================
        # INVALID OTP
        # =========================================

        if entered_otp != saved_otp:

            cursor.execute("""

                UPDATE password_reset_otp

                SET attempts = attempts + 1

                WHERE id = %s

            """, (otp_id,))

            conn.commit()

            attempts += 1

            if attempts >= 5:

                cursor.execute("""

                    UPDATE password_reset_otp

                    SET is_used = 1

                    WHERE id = %s

                """, (otp_id,))

                conn.commit()

                return jsonify({

                    "status": "error",

                    "message":
                    "OTP blocked after 5 invalid attempts"

                })

            return jsonify({

                "status": "error",

                "message":
                f"Invalid OTP. {5 - attempts} attempts remaining"

            })

        # =========================================
        # OTP VERIFIED
        # =========================================

        cursor.execute("""

            UPDATE password_reset_otp

            SET

                is_used = 1,
                verified_at = NOW()

            WHERE id = %s

        """, (otp_id,))

        conn.commit()

        # =========================================
        # SESSION VERIFIED
        # =========================================

        session["otp_verified"] = True

        session.modified = True

        # =========================================
        # SUCCESS
        # =========================================

        return jsonify({

            "status": "success",
            "message": "OTP Verified"

        })

    except Exception as e:

        print(
            "VERIFY OTP ERROR:",
            e
        )

        return jsonify({

            "status": "error",
            "message": "Something went wrong"

        })

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()
    
 
# =========================================================
# 🔒 UPDATE PASSWORD
# PURPOSE:
# Final password update after OTP verification
# =========================================================

@app.route(
    "/clerk/profile/update-password",
    methods=["POST"]
)
@login_required
def clerk_update_password():

    conn = None
    cursor = None

    try:

        # =========================================
        # SESSION USER
        # =========================================

        user_id = session.get(
            "clerk_user_id"
        )

        if not user_id:

            return jsonify({

                "status": "error",
                "message": "User session missing"

            })

            # =========================================
            # OTP SESSION CHECK
            # User must verify OTP in current session
            # =========================================

            if session.get("otp_verified") is not True:

                return jsonify({
                    "status": "error",
                    "message": "OTP verification required"
                })

        # =========================================
        # GET REQUEST DATA
        # =========================================

        data = request.get_json() or {}

        new_password = (
            data.get("password") or ""
        ).strip()

        # =========================================
        # PASSWORD REQUIRED
        # =========================================

        if not new_password:

            return jsonify({

                "status": "error",
                "message": "Password required"

            })

        # =========================================
        # STRONG PASSWORD VALIDATION
        # =========================================

        if (

            len(new_password) < 8
            or not any(
                c.isupper()
                for c in new_password
            )
            or not any(
                c.islower()
                for c in new_password
            )
            or not any(
                c.isdigit()
                for c in new_password
            )
            or not any(
                not c.isalnum()
                for c in new_password
            )

        ):

            return jsonify({

                "status": "error",

                "message":
                "Password must contain uppercase, lowercase, number and special character"

            })

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # CHECK VERIFIED OTP
        # =========================================

        cursor.execute("""

            SELECT

                id,
                verified_at

            FROM password_reset_otp

            WHERE user_id = %s
            AND is_used = 1

            ORDER BY id DESC

            LIMIT 1

        """, (user_id,))

        otp_row = cursor.fetchone()

        if not otp_row:

            return jsonify({

                "status": "error",

                "message":
                "OTP verification required"

            })

        verified_time = otp_row[1]

        # =========================================
        # OTP TIME LIMIT AFTER VERIFY
        # OPTIONAL SECURITY
        # =========================================

        if verified_time:

            minutes_passed = (

                datetime.now()
                - verified_time

            ).total_seconds() / 60

            if minutes_passed > 10:

                return jsonify({

                    "status": "error",

                    "message":
                    "OTP verification expired"

                })

        # =========================================
        # HASH PASSWORD
        # =========================================

        hashed_password = (

            bcrypt
            .generate_password_hash(
                new_password
            )
            .decode("utf-8")

        )

        # =========================================
        # UPDATE USER PASSWORD
        # =========================================

        cursor.execute("""

            UPDATE users

            SET

                password = %s,
                last_password_change = NOW(),
                updated_at = NOW()

            WHERE id = %s

        """, (

            hashed_password,
            user_id

        ))

        # =========================================
        # CLEAN OLD OTP RECORDS
        # =========================================

        cursor.execute("""

            DELETE FROM password_reset_otp

            WHERE user_id = %s

        """, (user_id,))

        conn.commit()

        # =========================================
        # CLEAR SESSION
        # =========================================

        session.pop(
            "otp_verified",
            None
        )

        # =========================================
        # SUCCESS
        # =========================================

        return jsonify({

            "status": "success",

            "message":
            "Password updated successfully"

        })

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "UPDATE PASSWORD ERROR:",
            e
        )

        return jsonify({

            "status": "error",
            "message": "Something went wrong"

        })

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


    
# =========================================================
# 🎓 AUTO GENERATE ADMISSION NUMBER
# (MYSQL SAFE + ATOMIC + PRODUCTION READY)
# =========================================================

def generate_admission_no(school_id):

    conn = None
    cursor = None

    try:

        # =========================================
        # GET SCHOOL CODE
        # =========================================

        school_code = get_school_code(
            school_id
        )

        conn = get_connection()

        cursor = conn.cursor()

        # =========================================
        # GET CURRENT NUMBER
        # =========================================

        cursor.execute("""

            SELECT admission_last_number

            FROM school_sequences

            WHERE school_id = %s

        """, (school_id,))

        row = cursor.fetchone()

        # =========================================
        # VALIDATION
        # =========================================

        if not row:

            raise Exception(
                "School sequence not found ❌"
            )

        current_number = row[0] or 0

        # =========================================
        # NEXT NUMBER
        # =========================================

        next_number = current_number + 1

        # =========================================
        # UPDATE NEW NUMBER
        # =========================================

        cursor.execute("""

            UPDATE school_sequences

            SET
                admission_last_number = %s

            WHERE school_id = %s

        """, (
            next_number,
            school_id
        ))

        conn.commit()

        # =========================================
        # FINAL ADMISSION NUMBER
        # =========================================

        admission_no = (
            f"{school_code}-ADM-"
            f"{str(next_number).zfill(4)}"
        )

        print(
            "✅ Generated Admission Number:",
            admission_no
        )

        return admission_no

    except Exception as e:

        if conn:
            conn.rollback()

        print(
            "ADMISSION NUMBER ERROR:",
            e
        )

        raise

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

            
# =========================================================
# ➕ ADD STUDENT
# Safe student registration with validation, duplicate check,
# optional unique fields and auto admission number.
# =========================================================
@app.route("/clerk/add-student", methods=["GET", "POST"])
@login_required
@subscription_required
def add_student():

    school_id = session.get("clerk_school_id")
    next_admission = "Auto Generated On Save"

    if not school_id:
        return redirect(url_for("login"))

    if request.method == "POST":

        conn = None
        cursor = None

        try:
            conn = get_connection()
            cursor = conn.cursor()

            # ================= SAFE FORM VALUE =================
            def form_value(field):
                value = request.form.get(field)
                return value.strip() if value else ""

            def optional_value(field):
                value = form_value(field)
                return value if value and value != "None" else None

            # ================= GET DATA =================
            school_register_no = form_value("school_register_no")
            name = form_value("name")
            father_name = form_value("father_name")
            mother_name = form_value("mother_name")

            student_uid = optional_value("student_uid")
            apaar_id = optional_value("apaar_id")

            aadhaar = form_value("aadhaar")

            dob = parse_date(request.form.get("dob"))
            birth_place = form_value("birth_place")
            nationality = form_value("nationality")
            mother_tongue = form_value("mother_tongue")
            religion = form_value("religion")
            caste = form_value("caste")

            city = form_value("city")
            taluka = form_value("taluka")
            district = form_value("district")
            state = form_value("state")

            admission_date = parse_date(request.form.get("admission_date"))
            student_class = form_value("class")
            section = form_value("section")
            previous_school = optional_value("previous_school")

            last_exam = optional_value("last_exam")
            result_status = optional_value("result_status")
            progress = optional_value("progress")
            conduct = optional_value("conduct")

            primary_mobile = form_value("primary_mobile")
            alternate_mobile = optional_value("alternate_mobile")
            email = optional_value("email")
            occupation = optional_value("occupation")
            income = optional_value("income")
            guardian_name = optional_value("guardian_name")
            guardian_mobile = optional_value("guardian_mobile")

            if email:
                email = email.lower()

            # ================= REQUIRED VALIDATION =================
            required_fields = {
                "School Register No": school_register_no,
                "Student Name": name,
                "Father Name": father_name,
                "Mother Name": mother_name,
                "Aadhaar": aadhaar,
                "Date of Birth": dob,
                "Mother Tongue": mother_tongue,
                "Religion": religion,
                "Caste": caste,
                "City": city,
                "Taluka": taluka,
                "District": district,
                "State": state,
                "Admission Date": admission_date,
                "Class": student_class,
                "Section": section,
                "Primary Mobile": primary_mobile
            }

            for label, value in required_fields.items():
                if not value:
                    return f"{label} is required ❌"

            # ================= FORMAT VALIDATION =================
            if not is_valid_aadhaar(aadhaar):
                return "Invalid Aadhaar number ❌"

            if not is_valid_phone(primary_mobile):
                return "Invalid primary mobile number ❌"

            if alternate_mobile and not is_valid_phone(alternate_mobile):
                return "Invalid alternate mobile number ❌"

            if guardian_mobile and not is_valid_phone(guardian_mobile):
                return "Invalid guardian mobile number ❌"

            if email and not is_valid_email(email):
                return "Invalid email ❌"

            # ================= LENGTH VALIDATION =================
            if len(school_register_no) > 50:
                return "School register number too long ❌"

            if len(name) > 200:
                return "Student name too long ❌"

            if len(father_name) > 200:
                return "Father name too long ❌"

            if len(mother_name) > 200:
                return "Mother name too long ❌"

            if student_uid and len(student_uid) > 50:
                return "Student UID too long ❌"

            if apaar_id and len(apaar_id) > 50:
                return "APAAR ID too long ❌"

            if email and len(email) > 255:
                return "Email too long ❌"

            # ================= DUPLICATE CHECK =================
            cursor.execute("""
                SELECT id
                FROM students
                WHERE school_id = %s
                AND (
                    school_register_no = %s
                    OR aadhaar = %s
                    OR (
                        student_uid = %s
                        AND %s IS NOT NULL
                        AND %s != ''
                    )
                    OR (
                        apaar_id = %s
                        AND %s IS NOT NULL
                        AND %s != ''
                    )
                )
                LIMIT 1
            """, (
                school_id,
                school_register_no,
                aadhaar,
                student_uid,
                student_uid,
                student_uid,
                apaar_id,
                apaar_id,
                apaar_id
            ))

            if cursor.fetchone():
                return "Student with same Register No, Aadhaar, Student UID or APAAR ID already exists ❌"

            # ================= GENERATE ADMISSION NO =================
            admission_no = generate_admission_no(school_id)

            # ================= INSERT STUDENT =================
            cursor.execute("""
                INSERT INTO students (
                    school_id,
                    school_register_no,
                    name,
                    father_name,
                    mother_name,
                    student_uid,
                    aadhaar,
                    apaar_id,
                    dob,
                    birth_place,
                    nationality,
                    mother_tongue,
                    religion,
                    caste,
                    city,
                    taluka,
                    district,
                    state,
                    admission_no,
                    admission_date,
                    class,
                    section,
                    previous_school,
                    last_exam,
                    result_status,
                    progress,
                    conduct,
                    primary_mobile,
                    alternate_mobile,
                    email,
                    occupation,
                    income,
                    guardian_name,
                    guardian_mobile
                )
                VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """, (
                school_id,
                school_register_no,
                name,
                father_name,
                mother_name,
                student_uid,
                aadhaar,
                apaar_id,
                dob,
                birth_place,
                nationality,
                mother_tongue,
                religion,
                caste,
                city,
                taluka,
                district,
                state,
                admission_no,
                admission_date,
                student_class,
                section,
                previous_school,
                last_exam,
                result_status,
                progress,
                conduct,
                primary_mobile,
                alternate_mobile,
                email,
                occupation,
                income,
                guardian_name,
                guardian_mobile
            ))

            conn.commit()

            print("✅ Student Saved:", admission_no)

            return redirect(url_for("clerk_students"))

        except Exception as e:

            if conn:
                conn.rollback()

            print("❌ ADD STUDENT ERROR:", e)

            return "Something went wrong ❌"

        finally:

            if cursor:
                cursor.close()

            if conn:
                conn.close()

    school = get_school_details(school_id)

    return render_template(
        "clerk/add_student.html",
        next_admission=next_admission,
        role="clerk",
        school_name=school["school_name"],
        school_udise=school["school_udise"],
        active_page="add_student"
    )


# =========================================================
# ✏️ EDIT STUDENT (CLERK + ADMIN SAFE)
# =========================================================
@app.route("/clerk/edit-student/<int:id>", methods=["GET", "POST"])
@login_required
@subscription_required
def edit_student(id):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # CHECK WHO IS ACCESSING
        # Clerk is preferred if both sessions exist
        # =========================================

        is_clerk_request = (
            session.get("clerk_logged_in") is True
            and session.get("clerk_role") == "clerk"
            and session.get("clerk_school_id")
        )

        is_admin_request = (
            session.get("admin_logged_in") is True
            and session.get("admin_role") == "admin"
        )

        if not is_clerk_request and not is_admin_request:
            return redirect(url_for("login"))

        # =========================================
        # FETCH STUDENT SAFELY
        # Clerk can access only own school student
        # Admin can access any student
        # =========================================

        if is_clerk_request:

            school_id = session.get("clerk_school_id")

            cursor.execute("""
                SELECT *
                FROM students
                WHERE id = %s
                AND school_id = %s
                LIMIT 1
            """, (
                id,
                school_id
            ))

        else:

            cursor.execute("""
                SELECT *
                FROM students
                WHERE id = %s
                LIMIT 1
            """, (
                id,
            ))

        row = cursor.fetchone()

        if not row:
            return "Student Not Found ❌"

        columns = [
            column[0]
            for column in cursor.description
        ]

        student = dict(
            zip(columns, row)
        )

        student_school_id = student["school_id"]

        # =========================================
        # UPDATE STUDENT
        # =========================================

        if request.method == "POST":

            # ================= SAFE FORM HELPERS =================

            def form_value(field):

                value = request.form.get(field)

                return value.strip() if value else ""


            def optional_value(field):

                value = form_value(field)

                if value in ["", "None", "none", "NULL", "null"]:
                    return None

                return value


            def required_date(field):

                value = request.form.get(field)

                return parse_date(value) if value else None

            # ================= GET DATA =================

            school_register_no = form_value("school_register_no")
            name = form_value("name")
            father_name = form_value("father_name")
            mother_name = form_value("mother_name")

            student_uid = optional_value("student_uid")
            apaar_id = optional_value("apaar_id")

            aadhaar = form_value("aadhaar")

            dob = required_date("dob")
            birth_place = form_value("birth_place")
            nationality = form_value("nationality")
            mother_tongue = form_value("mother_tongue")
            religion = form_value("religion")
            caste = form_value("caste")

            city = form_value("city")
            taluka = form_value("taluka")
            district = form_value("district")
            state = form_value("state")

            admission_date = required_date("admission_date")
            student_class = form_value("class")
            section = form_value("section")
            previous_school = optional_value("previous_school")

            last_exam = optional_value("last_exam")
            result_status = optional_value("result_status")
            progress = optional_value("progress")
            conduct = optional_value("conduct")

            primary_mobile = form_value("primary_mobile")
            alternate_mobile = optional_value("alternate_mobile")
            email = optional_value("email")
            occupation = optional_value("occupation")
            income = optional_value("income")
            guardian_name = optional_value("guardian_name")
            guardian_mobile = optional_value("guardian_mobile")

            if email:
                email = email.lower()

            # =========================================
            # REQUIRED FIELD VALIDATION
            # =========================================

            required_fields = {
                "School Register No": school_register_no,
                "Student Name": name,
                "Father Name": father_name,
                "Mother Name": mother_name,
                "Aadhaar": aadhaar,
                "Date of Birth": dob,
                "Birth Place": birth_place,
                "Nationality": nationality,
                "Mother Tongue": mother_tongue,
                "Religion": religion,
                "Caste": caste,
                "City": city,
                "Taluka": taluka,
                "District": district,
                "State": state,
                "Admission Date": admission_date,
                "Class": student_class,
                "Section": section,
                "Primary Mobile": primary_mobile
            }

            for field_name, field_value in required_fields.items():

                if not field_value:
                    return f"{field_name} is required ❌"

            # =========================================
            # FORMAT VALIDATION
            # =========================================

            if not is_valid_phone(primary_mobile):
                return "Invalid primary mobile number ❌"

            if alternate_mobile and not is_valid_phone(alternate_mobile):
                return "Invalid alternate mobile number ❌"

            if guardian_mobile and not is_valid_phone(guardian_mobile):
                return "Invalid guardian mobile number ❌"

            if email and not is_valid_email(email):
                return "Invalid email ❌"

            if not is_valid_aadhaar(aadhaar):
                return "Invalid Aadhaar number ❌"

            # =========================================
            # LENGTH VALIDATION
            # =========================================

            if len(school_register_no) > 50:
                return "School register number too long ❌"

            if len(name) > 200:
                return "Student name too long ❌"

            if len(father_name) > 200:
                return "Father name too long ❌"

            if len(mother_name) > 200:
                return "Mother name too long ❌"

            if student_uid and len(student_uid) > 50:
                return "Student UID too long ❌"

            if apaar_id and len(apaar_id) > 50:
                return "APAAR ID too long ❌"

            if len(birth_place) > 100:
                return "Birth place too long ❌"

            if len(nationality) > 50:
                return "Nationality too long ❌"

            if len(mother_tongue) > 50:
                return "Mother tongue too long ❌"

            if len(religion) > 50:
                return "Religion too long ❌"

            if len(caste) > 50:
                return "Caste too long ❌"

            if len(city) > 100:
                return "City name too long ❌"

            if len(taluka) > 100:
                return "Taluka name too long ❌"

            if len(district) > 100:
                return "District name too long ❌"

            if len(state) > 100:
                return "State name too long ❌"

            if len(student_class) > 50:
                return "Class too long ❌"

            if len(section) > 10:
                return "Section too long ❌"

            if previous_school and len(previous_school) > 200:
                return "Previous school name too long ❌"

            if last_exam and len(last_exam) > 100:
                return "Last exam too long ❌"

            if result_status and len(result_status) > 50:
                return "Result status too long ❌"

            if progress and len(progress) > 100:
                return "Progress too long ❌"

            if conduct and len(conduct) > 100:
                return "Conduct too long ❌"

            if email and len(email) > 255:
                return "Email too long ❌"

            if occupation and len(occupation) > 100:
                return "Occupation too long ❌"

            if income and len(income) > 50:
                return "Income too long ❌"

            if guardian_name and len(guardian_name) > 100:
                return "Guardian name too long ❌"

            # =========================================
            # DUPLICATE CHECK
            # Aadhaar = Mandatory Unique
            # Register No = School-wise Unique
            # Student UID = Optional Unique
            # APAAR ID = Optional Unique
            # =========================================

            cursor.execute("""
                SELECT id
                FROM students
                WHERE school_id = %s
                AND id != %s
                AND (
                    school_register_no = %s
                    OR aadhaar = %s
                    OR (
                        student_uid = %s
                        AND %s IS NOT NULL
                        AND %s != ''
                    )
                    OR (
                        apaar_id = %s
                        AND %s IS NOT NULL
                        AND %s != ''
                    )
                )
                LIMIT 1
            """, (
                student_school_id,
                id,

                school_register_no,
                aadhaar,

                student_uid,
                student_uid,
                student_uid,

                apaar_id,
                apaar_id,
                apaar_id
            ))

            existing_student = cursor.fetchone()

            if existing_student:

                return (
                    "Student with same Register No, Aadhaar, "
                    "Student UID or APAAR ID already exists ❌"
                )

            # =========================================
            # UPDATE QUERY
            # Admission number is not updated here
            # =========================================

            if is_clerk_request:

                cursor.execute("""
                    UPDATE students
                    SET
                        school_register_no = %s,
                        name = %s,
                        father_name = %s,
                        mother_name = %s,
                        student_uid = %s,
                        aadhaar = %s,
                        apaar_id = %s,
                        dob = %s,
                        birth_place = %s,
                        nationality = %s,
                        mother_tongue = %s,
                        religion = %s,
                        caste = %s,
                        city = %s,
                        taluka = %s,
                        district = %s,
                        state = %s,
                        admission_date = %s,
                        class = %s,
                        section = %s,
                        previous_school = %s,
                        last_exam = %s,
                        result_status = %s,
                        progress = %s,
                        conduct = %s,
                        primary_mobile = %s,
                        alternate_mobile = %s,
                        email = %s,
                        occupation = %s,
                        income = %s,
                        guardian_name = %s,
                        guardian_mobile = %s
                    WHERE id = %s
                    AND school_id = %s
                """, (
                    school_register_no,
                    name,
                    father_name,
                    mother_name,
                    student_uid,
                    aadhaar,
                    apaar_id,
                    dob,
                    birth_place,
                    nationality,
                    mother_tongue,
                    religion,
                    caste,
                    city,
                    taluka,
                    district,
                    state,
                    admission_date,
                    student_class,
                    section,
                    previous_school,
                    last_exam,
                    result_status,
                    progress,
                    conduct,
                    primary_mobile,
                    alternate_mobile,
                    email,
                    occupation,
                    income,
                    guardian_name,
                    guardian_mobile,
                    id,
                    school_id
                ))

            else:

                cursor.execute("""
                    UPDATE students
                    SET
                        school_register_no = %s,
                        name = %s,
                        father_name = %s,
                        mother_name = %s,
                        student_uid = %s,
                        aadhaar = %s,
                        apaar_id = %s,
                        dob = %s,
                        birth_place = %s,
                        nationality = %s,
                        mother_tongue = %s,
                        religion = %s,
                        caste = %s,
                        city = %s,
                        taluka = %s,
                        district = %s,
                        state = %s,
                        admission_date = %s,
                        class = %s,
                        section = %s,
                        previous_school = %s,
                        last_exam = %s,
                        result_status = %s,
                        progress = %s,
                        conduct = %s,
                        primary_mobile = %s,
                        alternate_mobile = %s,
                        email = %s,
                        occupation = %s,
                        income = %s,
                        guardian_name = %s,
                        guardian_mobile = %s
                    WHERE id = %s
                """, (
                    school_register_no,
                    name,
                    father_name,
                    mother_name,
                    student_uid,
                    aadhaar,
                    apaar_id,
                    dob,
                    birth_place,
                    nationality,
                    mother_tongue,
                    religion,
                    caste,
                    city,
                    taluka,
                    district,
                    state,
                    admission_date,
                    student_class,
                    section,
                    previous_school,
                    last_exam,
                    result_status,
                    progress,
                    conduct,
                    primary_mobile,
                    alternate_mobile,
                    email,
                    occupation,
                    income,
                    guardian_name,
                    guardian_mobile,
                    id
                ))

            conn.commit()

            print("✅ Student Updated in DB")

            if is_clerk_request:
                return redirect(url_for("clerk_dashboard"))

            return redirect(url_for("superadmin_all_students"))

        # =========================================
        # GET SCHOOL DETAILS FOR PAGE HEADER
        # =========================================

        school = get_school_details(student_school_id)

        if not school:
            return "School not found ❌"

        return render_template(
            "clerk/edit_student.html",
            student=student,
            role="clerk" if is_clerk_request else "admin",
            school_name=school["school_name"],
            school_udise=school["school_udise"],
            active_page="students"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ EDIT STUDENT ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📋 VIEW STUDENTS LIST PAGE
# =========================================================
@app.route("/clerk/students")
@login_required
@subscription_required
def clerk_students():

    conn = None
    cursor = None

    try:
        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        search = (request.args.get("search") or "").strip()
        class_filter = (request.args.get("class") or "").strip()

        page = request.args.get("page", 1, type=int)
        per_page = 10
        offset = (page - 1) * per_page

        conn = get_connection()

        if not conn:
            return "Database connection failed ❌"

        cursor = conn.cursor(dictionary=True)

        where_query = """
            FROM students
            WHERE school_id = %s
        """

        params = [school_id]

        if search:
            where_query += """
                AND (
                    name LIKE %s
                    OR admission_no LIKE %s
                    OR school_register_no LIKE %s
                    OR student_uid LIKE %s
                    OR primary_mobile LIKE %s
                )
            """

            like_search = f"%{search}%"

            params.extend([
                like_search,
                like_search,
                like_search,
                like_search,
                like_search
            ])

        if class_filter:
            where_query += " AND class = %s"
            params.append(class_filter)

        # TOTAL COUNT
        cursor.execute(
            "SELECT COUNT(*) AS total " + where_query,
            tuple(params)
        )

        total_records = cursor.fetchone()["total"] or 0

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        # STUDENT DATA
        query = """
            SELECT
                id,
                school_register_no,
                name,
                father_name,
                mother_name,
                class,
                section,
                admission_no,
                student_uid,
                apaar_id,
                aadhaar,
                dob,
                primary_mobile,
                progress,
                conduct,
                created_at
        """ + where_query + """
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        """

        data_params = params + [per_page, offset]

        cursor.execute(query, tuple(data_params))

        students = cursor.fetchall()

        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        return render_template(
            "clerk/students.html",
            students=students,
            search=search,
            class_filter=class_filter,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total_records=total_records,
            role="clerk",
            school_name=school["school_name"],
            active_page="students"
        )

    except Exception as e:
        print("❌ STUDENTS FETCH ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 📄 AUTO GENERATE TC NUMBER - SCHOOL SETTINGS BASED
# PURPOSE:
# Generate TC number using same DB transaction
# Prevents duplicate TC numbers during multiple users
# =========================================================

def generate_tc_number(cursor, school_id):

    # =====================================
    # GET SCHOOL CERTIFICATE SETTINGS
    # =====================================

    cursor.execute("""
        SELECT
            school_code,
            tc_prefix,
            auto_numbering
        FROM schools
        WHERE school_id = %s
        LIMIT 1
    """, (
        school_id,
    ))

    school = cursor.fetchone()

    if not school:
        raise Exception("School not found ❌")

    school_code = school[0]
    tc_prefix = school[1] or "TC"
    auto_numbering = school[2] or "Enabled"

    # =====================================
    # IF AUTO NUMBERING DISABLED
    # =====================================

    if auto_numbering != "Enabled":
        raise Exception("Auto numbering disabled for this school ❌")

    # =====================================
    # LOCK SCHOOL SEQUENCE
    # IMPORTANT: Requires same transaction
    # =====================================

    cursor.execute("""
        SELECT tc_last_number
        FROM school_sequences
        WHERE school_id = %s
        FOR UPDATE
    """, (
        school_id,
    ))

    row = cursor.fetchone()

    if not row:
        raise Exception("School sequence not found ❌")

    next_number = (row[0] or 0) + 1

    # =====================================
    # UPDATE SEQUENCE
    # =====================================

    cursor.execute("""
        UPDATE school_sequences
        SET tc_last_number = %s
        WHERE school_id = %s
    """, (
        next_number,
        school_id
    ))

    # =====================================
    # FINAL TC NUMBER
    # =====================================

    tc_number = (
        f"{school_code}-"
        f"{tc_prefix}-"
        f"{str(next_number).zfill(4)}"
    )

    return tc_number


# =========================================================
# 📄 TC FORM
# PURPOSE:
# Clerk can generate TC only for own school student
# TC number + TC insert happen in one transaction
# =========================================================

@app.route("/clerk/tc-form/<int:id>", methods=["GET", "POST"])
@login_required
@subscription_required
@feature_required("enable_tc_management")
def tc_form(id):

    conn = None
    cursor = None

    try:

        # =========================================
        # CLERK SESSION CHECK
        # =========================================

        school_id = session.get("clerk_school_id")

        if not school_id:
            return redirect(url_for("login"))

        if session.get("clerk_role") != "clerk":
            return "Unauthorized ❌"

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # FETCH STUDENT ONLY FROM CLERK SCHOOL
        # Prevents clerk accessing other school student
        # =========================================

        cursor.execute("""
            SELECT *
            FROM students
            WHERE id = %s
            AND school_id = %s
            LIMIT 1
        """, (
            id,
            school_id
        ))

        row = cursor.fetchone()

        if not row:
            return "Student Not Found ❌"

        columns = [column[0] for column in cursor.description]
        student = dict(zip(columns, row))

        # =========================================
        # POST: CREATE TC
        # =========================================

        if request.method == "POST":

            # =========================================
            # GET FORM DATA
            # =========================================

            tc_date_raw = (
                request.form.get("tc_date") or ""
            ).strip()

            leaving_date_raw = (
                request.form.get("leaving_date") or ""
            ).strip()

            leaving_reason = (
                request.form.get("leaving_reason") or ""
            ).strip()

            remark = (
                request.form.get("remark") or ""
            ).strip()

            # =========================================
            # PARSE DATES
            # =========================================

            tc_date = parse_date(tc_date_raw)
            leaving_date = parse_date(leaving_date_raw)

            # =========================================
            # VALIDATION
            # =========================================

            if not tc_date or not leaving_date:
                return "TC Date / Leaving Date invalid ❌"

            if not leaving_reason:
                return "Leaving reason required ❌"

            if len(leaving_reason) > 255:
                return "Leaving reason too long ❌"

            if remark and len(remark) > 500:
                return "Remark too long ❌"

            # =========================================
            # CONVERT DATETIME TO DATE FOR DB
            # =========================================

            tc_date_value = (
                tc_date.date()
                if hasattr(tc_date, "date")
                else tc_date
            )

            leaving_date_value = (
                leaving_date.date()
                if hasattr(leaving_date, "date")
                else leaving_date
            )

            # =========================================
            # LEAVING DATE CHECK
            # Leaving date cannot be before admission date
            # =========================================

            admission_date = student.get("admission_date")

            if admission_date:

                admission_date_value = (
                    admission_date.date()
                    if hasattr(admission_date, "date")
                    else admission_date
                )

                if leaving_date_value < admission_date_value:
                    return "Leaving date cannot be before admission date ❌"

            # =========================================
            # EXISTING TC CHECK
            # Prevents duplicate TC for same student
            # =========================================

            cursor.execute("""
                SELECT id
                FROM tc
                WHERE student_id = %s
                AND school_id = %s
                LIMIT 1
            """, (
                id,
                school_id
            ))

            existing_tc = cursor.fetchone()

            if existing_tc:
                return redirect(
                    url_for("view_tc", tc_id=existing_tc[0])
                )

            # =========================================
            # GENERATE TC NUMBER
            # Same cursor + same transaction
            # =========================================

            tc_number = generate_tc_number(
                cursor,
                school_id
            )

            # =========================================
            # INSERT TC
            # =========================================

            cursor.execute("""
                INSERT INTO tc (
                    school_id,
                    student_id,
                    tc_number,
                    tc_date,
                    leaving_date,
                    leaving_reason,
                    remark
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                school_id,
                id,
                tc_number,
                tc_date_value,
                leaving_date_value,
                leaving_reason,
                remark if remark else None
            ))

            new_tc_id = cursor.lastrowid

            # =========================================
            # COMMIT BOTH:
            # sequence update + TC insert
            # =========================================

            conn.commit()

            return redirect(
                url_for("view_tc", tc_id=new_tc_id)
            )

        # =========================================
        # GET: SHOW TC FORM
        # =========================================

        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        return render_template(
            "clerk/tc_form.html",
            student=student,
            tc_number="Auto Generate On Save",
            role="clerk",
            school_name=school["school_name"],
            school_udise=school["school_udise"],
            active_page="tc"
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ TC FORM ERROR:", e)

        return "TC form failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 👁️ VIEW TC (MAIN DISPLAY PAGE)
# =========================================================

@app.route("/clerk/tc/view/<int:tc_id>")
@login_required
@subscription_required
@feature_required("enable_tc_management")
def view_tc(tc_id):

    mode = request.args.get("mode", "").strip()

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        admin_mode = (
            session.get("admin_logged_in") is True
            and session.get("admin_role") == "admin"
        )

        if admin_mode:

            cursor.execute("""

                SELECT 
                    t.*,

                    s.name,
                    s.school_register_no,
                    s.student_uid,
                    s.apaar_id,
                    s.class AS class_name,
                    s.admission_no,
                    s.father_name,
                    s.mother_name,
                    s.nationality,
                    s.mother_tongue,
                    s.religion,
                    s.caste,
                    s.birth_place,
                    s.city,
                    s.taluka,
                    s.district,
                    s.state,
                    s.dob,
                    s.previous_school,
                    s.admission_date,
                    s.progress,
                    s.conduct,
                    s.aadhaar,
                    s.primary_mobile,
                    s.email,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,
                    sc.email AS school_email,
                    sc.udise_no,
                    sc.recognition_no,
                    sc.medium,
                    sc.school_index_no,
                    sc.board_name,
                    sc.logo_path,
                    sc.watermark_path,
                    sc.website,

                    sc.enable_certificate_labels,
                    sc.show_tc_logo,
                    sc.show_tc_watermark

                FROM tc t

                JOIN students s
                    ON t.student_id = s.id

                JOIN schools sc
                    ON t.school_id = sc.school_id

                WHERE t.id = %s

                LIMIT 1

            """, (tc_id,))

            role = "admin"

        else:

            school_id = session.get("clerk_school_id")

            if not school_id:
                return "School session missing ❌"

            cursor.execute("""

                SELECT 
                    t.*,

                    s.name,
                    s.school_register_no,
                    s.student_uid,
                    s.apaar_id,
                    s.class AS class_name,
                    s.admission_no,
                    s.father_name,
                    s.mother_name,
                    s.nationality,
                    s.mother_tongue,
                    s.religion,
                    s.caste,
                    s.birth_place,
                    s.city,
                    s.taluka,
                    s.district,
                    s.state,
                    s.dob,
                    s.previous_school,
                    s.admission_date,
                    s.progress,
                    s.conduct,
                    s.aadhaar,
                    s.primary_mobile,
                    s.email,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,
                    sc.email AS school_email,
                    sc.udise_no,
                    sc.recognition_no,
                    sc.medium,
                    sc.school_index_no,
                    sc.board_name,
                    sc.logo_path,
                    sc.watermark_path,
                    sc.website,

                    sc.enable_certificate_labels,
                    sc.show_tc_logo,
                    sc.show_tc_watermark

                FROM tc t

                JOIN students s
                    ON t.student_id = s.id

                JOIN schools sc
                    ON t.school_id = sc.school_id

                WHERE t.id = %s
                AND t.school_id = %s

                LIMIT 1

            """, (
                tc_id,
                school_id
            ))

            role = "clerk"

        row = cursor.fetchone()

        if not row:
            return "TC Not Found ❌"

        columns = [col[0] for col in cursor.description]
        row = dict(zip(columns, row))

        tc = {
            "id": row["id"],
            "tc_number": row["tc_number"],
            "tc_date": format_date(row["tc_date"]),
            "leaving_date": format_date(row["leaving_date"]),
            "leaving_reason": row["leaving_reason"],
            "remark": row["remark"]
        }

        student = {
            "id": row["student_id"],
            "school_register_no": row["school_register_no"] or "",
            "student_uid": row["student_uid"] or "",
            "apaar_id": row["apaar_id"] or "",
            "name": row["name"],
            "class_name": row["class_name"],
            "admission_no": row["admission_no"],
            "father_name": row["father_name"] or "",
            "mother_name": row["mother_name"] or "",
            "nationality": row["nationality"] or "",
            "mother_tongue": row["mother_tongue"] or "",
            "religion": row["religion"] or "",
            "caste": row["caste"] or "",
            "birth_place": row["birth_place"] or "",
            "city": row["city"] or "",
            "taluka": row["taluka"] or "",
            "district": row["district"] or "",
            "state": row["state"] or "",
            "dob": format_date(row["dob"]),
            "previous_school": row["previous_school"] or "",
            "admission_date": format_date(row["admission_date"]),
            "progress": row["progress"] or "",
            "conduct": row["conduct"] or "",
            "aadhaar": row["aadhaar"] or "",
            "primary_mobile": row["primary_mobile"] or "",
            "email": row["email"] or ""
        }

        school = {
            "name": row["school_name"],
            "address": row["address"] or "",
            "phone": row["phone"] or "",
            "email": row["school_email"] or "",
            "udise_no": row["udise_no"] or "",
            "recognition_no": row["recognition_no"] or "",
            "medium": row["medium"] or "",
            "school_index_no": row["school_index_no"] or "",
            "board_name": row["board_name"] or "",
            "logo_path": row["logo_path"] or "",
            "watermark_path": row["watermark_path"] or "",
            "website": row["website"] or "",

            "enable_certificate_labels": row["enable_certificate_labels"] or "Enabled",
            "show_tc_logo": row["show_tc_logo"] or "Disabled",
            "show_tc_watermark": row["show_tc_watermark"] or "Disabled"
        }

        return render_template(
            "clerk/tc_generate.html",
            student=student,
            tc=tc,
            school=school,
            role=role,
            school_name=school["name"],
            school_address=school["address"],
            school_phone=school["phone"],
            school_email=school["email"],
            school_udise=school["udise_no"],
            school_recognition_no=school["recognition_no"],
            school_medium=school["medium"],
            school_index_no=school["school_index_no"],
            school_board_name=school["board_name"],
            school_logo=school["logo_path"],
            school_watermark=school["watermark_path"],
            school_website=school["website"]
        )

    except Exception as e:

        print("❌ VIEW TC ERROR:", e)

        return "TC view failed ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

            
# =========================================================
# 📄 DOWNLOAD TC PDF (HTML → PDFKIT)
# =========================================================

import pdfkit
import os
from flask import render_template, send_file, request, session


@app.route("/clerk/tc/pdf/<int:tc_id>")
@login_required
@subscription_required
@feature_required("enable_tc_management")
def download_tc_pdf(tc_id):

    conn = None
    cursor = None

    try:
        mode = request.args.get("mode")
        school_id = session.get("clerk_school_id")

        if not school_id:
                return "School session missing ❌"

        conn = get_connection()
        cursor = conn.cursor()

        # =====================================================
        # ADMIN MODE
        # =====================================================
        if (
            mode == "admin"
            and session.get("admin_logged_in")
            and session.get("admin_role") == "admin"
        ):

            cursor.execute("""
            SELECT 
                t.*,

                s.school_register_no,
                s.student_uid,
                s.apaar_id,
                s.name,
                s.father_name,
                s.mother_name,
                s.class AS class_name,
                s.admission_no,
                s.dob,
                s.aadhaar,
                s.primary_mobile,
                s.email,

                s.birth_place,
                s.nationality,
                s.mother_tongue,
                s.religion,
                s.caste,

                s.city,
                s.taluka,
                s.district,
                s.state,

                s.admission_date,
                s.section,
                s.previous_school,
                s.last_exam,
                s.result_status,

                s.progress,
                s.conduct,

                sc.name AS school_name,
                sc.address,
                sc.phone,
                sc.email,
                sc.udise_no,
                sc.recognition_no,
                sc.medium,
                sc.school_index_no,
                sc.board_name,
                sc.logo_path,
                sc.watermark_path,
                sc.website

            FROM tc t
            JOIN students s ON t.student_id = s.id
            JOIN schools sc ON t.school_id = sc.school_id
            WHERE t.id = %s
            """, (tc_id,))

        # =====================================================
        # CLERK MODE
        # =====================================================
        else:

            cursor.execute("""
            SELECT 
                t.*,

                s.school_register_no,
                s.student_uid,
                s.apaar_id,
                s.name,
                s.father_name,
                s.mother_name,
                s.class AS class_name,
                s.admission_no,
                s.dob,
                s.aadhaar,
                s.primary_mobile,
                s.email,
                s.birth_place,
                s.nationality,
                s.mother_tongue,
                s.religion,
                s.caste,

                s.city,
                s.taluka,
                s.district,
                s.state,

                s.admission_date,
                s.section,
                s.previous_school,
                s.last_exam,
                s.result_status,

                s.progress,
                s.conduct,

                sc.name AS school_name,
                sc.address,
                sc.phone,
                sc.email,
                sc.udise_no,
                sc.recognition_no,
                sc.medium,
                sc.school_index_no,
                sc.board_name,
                sc.logo_path,
                sc.watermark_path,
                sc.website

            FROM tc t
            JOIN students s ON t.student_id = s.id
            JOIN schools sc ON t.school_id = sc.school_id
            WHERE t.id = %s AND t.school_id = %s
            """, (tc_id, school_id))

        row = cursor.fetchone()

        if not row:
            return "TC Not Found ❌"
        
        columns = [col[0] for col in cursor.description]
        row = dict(zip(columns, row))

        # =====================================================
        # TC DATA
        # =====================================================
        tc = {
            "id": row["id"],
            "tc_number": row["tc_number"],
            "tc_date": format_date(row["tc_date"]),
            "leaving_date": format_date(row["leaving_date"]),
            "leaving_reason": row["leaving_reason"],
            "remark": row["remark"]
        }

        # =====================================================
        # STUDENT DATA
        # =====================================================
        student = {
            "school_register_no": row["school_register_no"],
            "student_uid": row["student_uid"],
            "apaar_id": row["apaar_id"],
            "name": row["name"],
            "father_name": row["father_name"],
            "mother_name": row["mother_name"],
            "class_name": row["class_name"],
            "admission_no": row["admission_no"],
            "dob": format_date(row["dob"]),
            "aadhaar": row["aadhaar"],
            "primary_mobile": row["primary_mobile"] or "",
            "email": row["email"] or "",
            "birth_place": row["birth_place"],
            "nationality": row["nationality"],
            "mother_tongue": row["mother_tongue"],
            "religion": row["religion"],
            "caste": row["caste"],
            "city": row["city"],
            "taluka": row["taluka"],
            "district": row["district"],
            "state": row["state"],
            "admission_date": format_date(row["admission_date"]),
            "section": row["section"],
            "previous_school": row["previous_school"],
            "last_exam": row["last_exam"],
            "result_status": row["result_status"],
            "progress": row["progress"],
            "conduct": row["conduct"]
        }

        # =====================================================
        # SCHOOL DATA
        # =====================================================

        base_dir = os.path.abspath(os.path.dirname(__file__))

        logo_absolute = ""
        watermark_absolute = ""

        if row["logo_path"]:

            logo_absolute = "file:///" + os.path.join(

                base_dir,
                "static",
                row["logo_path"].replace("static/", "")

            ).replace("\\", "/")

        if row["watermark_path"]:

            watermark_absolute = "file:///" + os.path.join(

                base_dir,
                "static",
                row["watermark_path"].replace("static/", "")

            ).replace("\\", "/")

        school = {
            "name": row["school_name"],
            "address": row["address"] or "",
            "phone": row["phone"] or "",
            "email": row["email"] or "",
            "udise_no": row["udise_no"] or "",
            "recognition_no": row["recognition_no"] or "",
            "medium": row["medium"] or "",
            "school_index_no": row["school_index_no"] or "",
            "board_name": row["board_name"] or "",
            "logo_path": logo_absolute,
            "watermark_path": watermark_absolute,
            "website": row["website"] or ""
        }

        # =====================================================
        # RENDER HTML
        # =====================================================
        html = render_template(
            "clerk/tc_generate.html",
            tc=tc,
            student=student,
            school=school,
            is_pdf=True
        )

        # =====================================================
        # PDF SAVE FOLDER
        # =====================================================
        pdf_folder = os.path.join("static", "generated_tc")

        if not os.path.exists(pdf_folder):
            os.makedirs(pdf_folder)

        pdf_path = os.path.join(
            pdf_folder,
            f"tc_{tc['id']}.pdf"
        )

        # =====================================================
        # PDF OPTIONS
        # =====================================================
        options = {
            "page-size": "A4",
            "margin-top": "5mm",
            "margin-right": "5mm",
            "margin-bottom": "5mm",
            "margin-left": "5mm",
            "encoding": "UTF-8",
            "enable-local-file-access": ""
        }

        # =====================================================
        # GENERATE PDF
        # =====================================================

        pdfkit.from_string(

            html,
            pdf_path,
            configuration=pdf_config,
             options=options
        )

        # =====================================================
        # 📧 AUTO SEND TC PDF EMAIL
        # =====================================================

        try:

            student_email = student.get("email")

            if student_email:

                # =========================================
                # CHECK ALREADY SENT
                # =========================================

                cursor.execute("""

                    SELECT id
                    FROM tc_email_logs
                    WHERE school_id = %s
                    AND tc_id = %s

                """, (

                    school_id,
                    tc["id"]

                ))

                already_sent = cursor.fetchone()

                # =========================================
                # SEND ONLY ONCE
                # =========================================

                if not already_sent:

                    subject = (
                        f"Transfer Certificate - {student['name']}"
                    )

                    body = f"""

                    <div style="font-family:Arial;padding:20px;">

                        <h2 style="color:#14b8a6;">
                            Transfer Certificate Generated
                        </h2>

                        <p>
                            Dear Parent/Student,
                        </p>

                        <p>
                            Your Transfer Certificate has been generated successfully.
                        </p>

                        <hr>

                        <p>
                            <b>Student Name:</b>
                            {student['name']}
                        </p>

                        <p>
                            <b>TC Number:</b>
                            {tc['tc_number']}
                        </p>

                        <p>
                            <b>School:</b>
                            {school['name']}
                        </p>

                        <hr>

                        <p>
                            Please find the TC PDF attached.
                        </p>

                        <p>
                            Thank you,
                            <br>
                            <b>{school['name']}</b>
                        </p>

                    </div>

                    """

                    # =========================================
                    # SEND MAIL WITH ATTACHMENT
                    # =========================================

                    email_sent = send_email(

                        student_email,
                        subject,
                        body,
                        pdf_path,
                        f"{tc['tc_number']}.pdf"

                    )
                    # FOR SMS
                    # sms_sent = send_sms(
                    #     mobile_number,
                    #     f"Your OTP is {otp}"
                    # )

                    # =========================================
                    # EMAIL SUCCESS
                    # =========================================

                    if email_sent == True:

                        print("✅ TC PDF EMAIL SENT")

                        # =====================================
                        # SAVE EMAIL LOG
                        # =====================================

                        cursor.execute("""

                            INSERT INTO tc_email_logs
                            (
                                school_id,
                                tc_id,
                                student_email
                            )
                            VALUES (%s, %s, %s)

                        """, (

                            school_id,
                            tc["id"],
                            student_email

                        ))

                        conn.commit()

                    # =========================================
                    # EMAIL FAILED
                    # =========================================

                    else:

                        print(
                            "❌ TC EMAIL FAILED:",
                            email_sent
                        )

        except Exception as email_error:

            print(
                "❌ TC EMAIL ERROR:",
                email_error
            )

        # =====================================================
        # RETURN FILE
        # =====================================================
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=f"{tc['tc_number']}.pdf"
        )

    except Exception as e:
        print("❌ TC PDF ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 🔓 PUBLIC TC VIEW (SAFE PUBLIC VERIFY) FOR PARENT WP
# =========================================================
@app.route("/public/tc/<int:tc_id>")
def public_tc(tc_id):

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        # ================= GET FULL TC DATA =================
        cursor.execute("""
            SELECT
                t.*,

                s.name,
                s.father_name,
                s.mother_name,
                s.class AS class_name,
                s.admission_no,
                s.dob,
                s.aadhaar,
                s.birth_place,
                s.nationality,
                s.mother_tongue,
                s.religion,
                s.caste,
                s.city,
                s.taluka,
                s.district,
                s.state,
                s.admission_date,
                s.section,
                s.previous_school,
                s.last_exam,
                s.result_status,
                s.progress,
                s.conduct,

                sc.name AS school_name,
                sc.address,
                sc.phone,
                sc.email,
                sc.udise_no,
                sc.recognition_no,
                sc.medium,
                sc.school_index_no,
                sc.board_name,
                sc.logo_path,
                sc.watermark_path,
                sc.website

            FROM tc t
            JOIN students s
                ON t.student_id = s.id
            JOIN schools sc
                ON t.school_id = sc.school_id
            WHERE t.id = %s
        """, (tc_id,))

        row = cursor.fetchone()

        if not row:
            return "TC Not Found ❌"
        
        columns = [col[0] for col in cursor.description]
        row = dict(zip(columns, row))

        # ================= TC =================
        tc = {
            "id": row["id"],
            "tc_number": row["tc_number"],
            "tc_date": format_date(row["tc_date"]),
            "leaving_date": format_date(row["leaving_date"]),
            "leaving_reason": row["leaving_reason"],
            "remark": row["remark"]
        }

        # ================= STUDENT =================
        student = {
            "name": row["name"],
            "father_name": row["father_name"],
            "mother_name": row["mother_name"],
            "class_name": row["class_name"],
            "admission_no": row["admission_no"],
            "dob": format_date(row["dob"]),
            "aadhaar": "XXXX-XXXX-" + str(row["aadhaar"])[-4:] if row["aadhaar"] else "",
            "birth_place": row["birth_place"],
            "nationality": row["nationality"],
            "mother_tongue": row["mother_tongue"],
            "religion": row["religion"],
            "caste": row["caste"],
            "city": row["city"],
            "taluka": row["taluka"],
            "district": row["district"],
            "state": row["state"],
            "admission_date": format_date(row["admission_date"]),
            "section": row["section"],
            "previous_school": row["previous_school"],
            "last_exam": row["last_exam"],
            "result_status": row["result_status"],
            "progress": row["progress"],
            "conduct": row["conduct"]
        }

        # ================= SCHOOL =================
        school = {
            "name": row["school_name"],
            "address": row["address"] or "",
            "phone": row["phone"] or "",
            "email": row["email"] or "",
            "udise_no": row["udise_no"] or "",
            "recognition_no": row["recognition_no"] or "",
            "medium": row["medium"] or "",
            "school_index_no": row["school_index_no"] or "",
            "board_name": row["board_name"] or "",
            "logo_path": row["logo_path"] or "",
            "watermark_path": row["watermark_path"] or "",
            "website": row["website"] or ""
        }

        return render_template(
            "clerk/tc_generate.html",
            tc=tc,
            student=student,
            school=school,
            school_name=school["name"],
            school_address=school["address"],
            school_phone=school["phone"],
            school_email=school["email"],
            school_udise=school["udise_no"],
            school_recognition_no=school["recognition_no"],
            school_medium=school["medium"],
            school_index_no=school["school_index_no"],
            school_board_name=school["board_name"],
            school_logo_path=school["logo_path"],
            school_watermark_path=school["watermark_path"],
            school_website=school["website"],
            is_public=True
        )

    except Exception as e:
        print("❌ PUBLIC TC ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# =========================================================
# 📊 TC HISTORY PAGE (SAFE + BACKEND FILTER + PAGINATION)
# =========================================================
@app.route("/clerk/tc")
@login_required
@subscription_required
@feature_required("enable_tc_management")
def clerk_tc_page():

    conn = None
    cursor = None

    try:

        # =========================================
        # CLERK SCHOOL SESSION
        # =========================================

        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        # =========================================
        # FILTER VALUES
        # =========================================

        search = (request.args.get("search") or "").strip()
        class_filter = (request.args.get("class") or "").strip()
        year_filter = (request.args.get("year") or "").strip()

        # =========================================
        # PAGINATION VALUES
        # =========================================

        page = request.args.get("page", 1, type=int)

        if page < 1:
            page = 1

        per_page = 10
        offset = (page - 1) * per_page

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # COMMON WHERE QUERY
        # Ensures only logged-in school TC records
        # =========================================

        where_query = """
            FROM tc t
            JOIN students s
                ON t.student_id = s.id
                AND s.school_id = t.school_id
            WHERE t.school_id = %s
        """

        params = [school_id]

        # =========================================
        # SEARCH FILTER
        # =========================================

        if search:

            where_query += """
                AND (
                    s.name LIKE %s
                    OR s.admission_no LIKE %s
                    OR t.tc_number LIKE %s
                )
            """

            like_search = f"%{search}%"

            params.extend([
                like_search,
                like_search,
                like_search
            ])

        # =========================================
        # CLASS FILTER
        # =========================================

        if class_filter:

            where_query += """
                AND s.class = %s
            """

            params.append(class_filter)

        # =========================================
        # YEAR FILTER
        # =========================================

        if year_filter:

            where_query += """
                AND YEAR(t.tc_date) = %s
            """

            params.append(year_filter)

        # =========================================
        # TOTAL FILTERED RECORDS
        # =========================================

        cursor.execute(
            "SELECT COUNT(*) AS total " + where_query,
            tuple(params)
        )

        total_records = cursor.fetchone()["total"] or 0

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (page - 1) * per_page

        # =========================================
        # MAIN TC LIST WITH LIMIT
        # =========================================

        query = """
            SELECT
                t.id,
                t.tc_number,
                t.tc_date,
                s.name,
                s.class AS class_name,
                s.admission_no
        """ + where_query + """
            ORDER BY t.id DESC
            LIMIT %s OFFSET %s
        """

        cursor.execute(
            query,
            tuple(params + [per_page, offset])
        )

        rows = cursor.fetchall()

        tc_list = []

        for r in rows:

            tc_list.append({
                "id": r["id"],
                "tc_number": r["tc_number"],
                "tc_date": format_date(r["tc_date"]),
                "name": r["name"],
                "class": r["class_name"],
                "admission_no": r["admission_no"]
            })

        # =========================================
        # TOTAL TC COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM tc
            WHERE school_id = %s
        """, (school_id,))

        total_tc = cursor.fetchone()["total"] or 0

        # =========================================
        # TODAY TC COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM tc
            WHERE school_id = %s
            AND DATE(created_at) = CURDATE()
        """, (school_id,))

        today_tc = cursor.fetchone()["total"] or 0

        # =========================================
        # MONTH TC COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM tc
            WHERE school_id = %s
            AND MONTH(created_at) = MONTH(NOW())
            AND YEAR(created_at) = YEAR(NOW())
        """, (school_id,))

        month_tc = cursor.fetchone()["total"] or 0

        # =========================================
        # YEAR DROPDOWN DATA
        # =========================================

        cursor.execute("""
            SELECT DISTINCT YEAR(tc_date) AS year_no
            FROM tc
            WHERE school_id = %s
            AND tc_date IS NOT NULL
            ORDER BY year_no DESC
        """, (school_id,))

        year_rows = cursor.fetchall()

        years = [
            row["year_no"]
            for row in year_rows
            if row["year_no"]
        ]

        # =========================================
        # SCHOOL DETAILS
        # =========================================

        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        return render_template(
            "clerk/tc_search.html",
            role="clerk",
            school_name=school["school_name"],
            school_udise=school["school_udise"],
            active_page="tc",

            tc_list=tc_list,

            total_tc=total_tc,
            today_tc=today_tc,
            month_tc=month_tc,

            search=search,
            class_filter=class_filter,
            year_filter=year_filter,
            years=years,

            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total_records=total_records
        )

    except Exception as e:

        print("❌ TC PAGE ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()


# =========================================================
# 🧾 AUTO GENERATE BONAFIDE NUMBER - SCHOOL SETTINGS BASED
# PURPOSE:
# Generate bonafide number inside same DB transaction
# Prevents duplicate / lost bonafide numbers
# =========================================================

def generate_bonafide_number(cursor, school_id):

    # =====================================
    # GET SCHOOL CERTIFICATE SETTINGS
    # =====================================

    cursor.execute("""
        SELECT
            school_code,
            bonafide_prefix,
            auto_numbering
        FROM schools
        WHERE school_id = %s
        LIMIT 1
    """, (
        school_id,
    ))

    school = cursor.fetchone()

    if not school:
        raise Exception("School not found ❌")

    school_code = school[0]
    bonafide_prefix = school[1] or "BON"
    auto_numbering = school[2] or "Enabled"

    # =====================================
    # IF AUTO NUMBERING DISABLED
    # =====================================

    if auto_numbering != "Enabled":
        raise Exception("Auto numbering disabled for this school ❌")

    # =====================================
    # LOCK ONLY THIS SCHOOL SEQUENCE
    # Safe for multiple schools / clerks
    # =====================================

    cursor.execute("""
        SELECT bonafide_last_number
        FROM school_sequences
        WHERE school_id = %s
        FOR UPDATE
    """, (
        school_id,
    ))

    row = cursor.fetchone()

    if not row:
        raise Exception("School sequence not found ❌")

    next_number = (row[0] or 0) + 1

    # =====================================
    # UPDATE SEQUENCE
    # =====================================

    cursor.execute("""
        UPDATE school_sequences
        SET bonafide_last_number = %s
        WHERE school_id = %s
    """, (
        next_number,
        school_id
    ))

    # =====================================
    # FINAL BONAFIDE NUMBER
    # =====================================

    bonafide_number = (
        f"{school_code}-"
        f"{bonafide_prefix}-"
        f"{str(next_number).zfill(4)}"
    )

    return bonafide_number

 # =========================================================
# 👁️ VIEW BONAFIDE (PRINT PAGE)
# PURPOSE:
# Clerk can view only own school bonafide
# Admin can view all only from valid admin session
# =========================================================

@app.route("/clerk/bonafide/view/<int:bid>")
@login_required
@subscription_required
@feature_required("enable_bonafide_management")
def view_bonafide(bid):

    conn = None
    cursor = None

    try:

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # STRICT ADMIN CHECK
        # =========================================

        admin_mode = (
            session.get("admin_logged_in") is True
            and session.get("admin_role") == "admin"
        )

        # =========================================
        # ADMIN VIEW
        # =========================================

        if admin_mode:

            cursor.execute("""
                SELECT 
                    b.*,

                    s.name,
                    s.class AS class_name,
                    s.admission_date,
                    s.dob,
                    s.caste,
                    s.primary_mobile,
                    s.email AS student_email,
                    s.school_register_no,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,
                    sc.email AS school_email,
                    sc.logo_path,
                    sc.watermark_path,

                    sc.enable_certificate_labels,
                    sc.show_bonafide_logo,
                    sc.show_bonafide_watermark

                FROM bonafide b

                JOIN students s
                    ON b.student_id = s.id
                    AND b.school_id = s.school_id

                JOIN schools sc
                    ON b.school_id = sc.school_id

                WHERE b.id = %s

                LIMIT 1
            """, (bid,))

            role = "admin"

        # =========================================
        # CLERK VIEW
        # Clerk can view only own school bonafide
        # =========================================

        else:

            school_id = session.get("clerk_school_id")

            if not school_id:
                return "School session missing ❌"

            cursor.execute("""
                SELECT 
                    b.*,

                    s.name,
                    s.class AS class_name,
                    s.admission_date,
                    s.dob,
                    s.caste,
                    s.primary_mobile,
                    s.email AS student_email,
                    s.school_register_no,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,
                    sc.email AS school_email,
                    sc.logo_path,
                    sc.watermark_path,

                    sc.enable_certificate_labels,
                    sc.show_bonafide_logo,
                    sc.show_bonafide_watermark

                FROM bonafide b

                JOIN students s
                    ON b.student_id = s.id
                    AND b.school_id = s.school_id

                JOIN schools sc
                    ON b.school_id = sc.school_id

                WHERE b.id = %s
                AND b.school_id = %s

                LIMIT 1
            """, (
                bid,
                school_id
            ))

            role = "clerk"

        row = cursor.fetchone()

        if not row:
            return "Bonafide Not Found ❌"

        columns = [col[0] for col in cursor.description]
        data = dict(zip(columns, row))

        # =========================================
        # STUDENT DATA
        # =========================================

        student = {
            "name": data.get("name") or "",
            "class": data.get("class_name") or "",
            "admission_date": format_date(data.get("admission_date")),
            "dob": format_date(data.get("dob")),
            "caste": data.get("caste") or "",
            "primary_mobile": data.get("primary_mobile") or "",
            "email": data.get("student_email") or "",
            "school_register_no": data.get("school_register_no") or ""
        }

        # =========================================
        # BONAFIDE DATA
        # =========================================

        bonafide = {
            "id": data.get("id"),
            "bonafide_number": data.get("bonafide_number") or "",
            "purpose": data.get("purpose") or "",
            "date": format_date(data.get("date"))
        }

        # =========================================
        # SCHOOL DATA
        # =========================================

        school = {
            "name": data.get("school_name") or "",
            "address": data.get("address") or "",
            "phone": data.get("phone") or "",
            "email": data.get("school_email") or "",
            "logo_path": data.get("logo_path") or "",
            "watermark_path": data.get("watermark_path") or "",

            "enable_certificate_labels": data.get("enable_certificate_labels") or "Enabled",
            "show_bonafide_logo": data.get("show_bonafide_logo") or "Disabled",
            "show_bonafide_watermark": data.get("show_bonafide_watermark") or "Disabled"
        }

        return render_template(
            "clerk/bonafide_generate.html",
            student=student,
            bonafide=bonafide,
            school=school,
            role=role
        )

    except Exception as e:

        print("❌ VIEW BONAFIDE ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📜 BONAFIDE PAGE
# PURPOSE:
# Show bonafide records for logged-in clerk school only
# Backend search + class/year filter + pagination
# Production-safe for large data
# =========================================================

@app.route("/clerk/bonafide/")
@login_required
@subscription_required
@feature_required("enable_bonafide_management")
def clerk_bonafide_page():

    conn = None
    cursor = None

    try:

        # =========================================
        # CLERK SCHOOL SESSION
        # =========================================

        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        # =========================================
        # FILTER VALUES
        # =========================================

        search = (
            request.args.get("search") or ""
        ).strip()

        class_filter = (
            request.args.get("class") or ""
        ).strip()

        year_filter = (
            request.args.get("year") or ""
        ).strip()

        # =========================================
        # PAGINATION VALUES
        # =========================================

        page = request.args.get(
            "page",
            1,
            type=int
        )

        if page < 1:
            page = 1

        per_page = 10

        offset = (
            page - 1
        ) * per_page

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        # =========================================
        # SCHOOL INFO
        # =========================================

        cursor.execute("""
            SELECT
                school_id,
                name,
                school_code
            FROM schools
            WHERE school_id = %s
            LIMIT 1
        """, (
            school_id,
        ))

        school = cursor.fetchone()

        if not school:
            return "School not found ❌"

        school_name = school["name"]
        school_code = school["school_code"]

        # =========================================
        # COMMON WHERE QUERY
        # Only logged-in school bonafide records
        # =========================================

        where_query = """
            FROM bonafide b

            JOIN students s
                ON b.student_id = s.id
                AND s.school_id = b.school_id

            WHERE b.school_id = %s
        """

        params = [
            school_id
        ]

        # =========================================
        # SEARCH FILTER
        # =========================================

        if search:

            where_query += """
                AND (
                    s.name LIKE %s
                    OR s.admission_no LIKE %s
                    OR s.school_register_no LIKE %s
                    OR s.student_uid LIKE %s
                    OR s.apaar_id LIKE %s
                    OR b.bonafide_number LIKE %s
                )
            """

            like_search = f"%{search}%"

            params.extend([
                like_search,
                like_search,
                like_search,
                like_search,
                like_search,
                like_search
            ])

        # =========================================
        # CLASS FILTER
        # =========================================

        if class_filter:

            where_query += """
                AND s.class = %s
            """

            params.append(
                class_filter
            )

        # =========================================
        # YEAR FILTER
        # =========================================

        if year_filter:

            where_query += """
                AND YEAR(b.date) = %s
            """

            params.append(
                year_filter
            )

        # =========================================
        # TOTAL FILTERED RECORDS
        # =========================================

        cursor.execute(
            "SELECT COUNT(*) AS total " + where_query,
            tuple(params)
        )

        total_records = (
            cursor.fetchone()["total"] or 0
        )

        total_pages = max(
            1,
            (total_records + per_page - 1) // per_page
        )

        if page > total_pages:
            page = total_pages
            offset = (
                page - 1
            ) * per_page

        # =========================================
        # MAIN BONAFIDE LIST
        # =========================================

        query = """
            SELECT
                b.id,
                b.student_id,
                b.bonafide_number,
                b.date,
                b.purpose,

                s.school_register_no,
                s.name,
                s.admission_no,
                s.student_uid,
                s.apaar_id,
                s.primary_mobile,
                s.email,

                s.class AS class_name

        """ + where_query + """
            ORDER BY b.id DESC
            LIMIT %s OFFSET %s
        """

        cursor.execute(
            query,
            tuple(
                params + [
                    per_page,
                    offset
                ]
            )
        )

        rows = cursor.fetchall()

        bonafide_list = []

        for r in rows:

            bonafide_list.append({

                "id": r["id"],
                "student_id": r["student_id"],
                "bonafide_number": r["bonafide_number"],
                "date": format_date(r["date"]),
                "purpose": r["purpose"] or "",

                "school_register_no": r["school_register_no"] or "",
                "name": r["name"] or "",
                "admission_no": r["admission_no"] or "",
                "student_uid": r["student_uid"] or "",
                "apaar_id": r["apaar_id"] or "",

                "primary_mobile": r["primary_mobile"] or "",
                "email": r["email"] or "",

                "class": r["class_name"] or ""

            })

        # =========================================
        # TOTAL BONAFIDE COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM bonafide
            WHERE school_id = %s
        """, (
            school_id,
        ))

        total_bonafide = (
            cursor.fetchone()["total"] or 0
        )

        # =========================================
        # TODAY BONAFIDE COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM bonafide
            WHERE school_id = %s
            AND DATE(created_at) = CURDATE()
        """, (
            school_id,
        ))

        today_bonafide = (
            cursor.fetchone()["total"] or 0
        )

        # =========================================
        # MONTH BONAFIDE COUNT
        # =========================================

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM bonafide
            WHERE school_id = %s
            AND MONTH(created_at) = MONTH(NOW())
            AND YEAR(created_at) = YEAR(NOW())
        """, (
            school_id,
        ))

        month_bonafide = (
            cursor.fetchone()["total"] or 0
        )

        # =========================================
        # YEAR DROPDOWN DATA
        # =========================================

        cursor.execute("""
            SELECT DISTINCT YEAR(date) AS year_no
            FROM bonafide
            WHERE school_id = %s
            AND date IS NOT NULL
            ORDER BY year_no DESC
        """, (
            school_id,
        ))

        year_rows = cursor.fetchall()

        years = [
            row["year_no"]
            for row in year_rows
            if row["year_no"]
        ]


        # =========================================
        # HISTORY MAP FOR CURRENT PAGE STUDENTS
        # =========================================

        student_ids = list({
            b["student_id"]
            for b in bonafide_list
        })

        bonafide_history_map = {}

        if student_ids:

            placeholders = ",".join(["%s"] * len(student_ids))

            cursor.execute(f"""
                SELECT
                    id,
                    student_id,
                    bonafide_number,
                    date,
                    purpose
                FROM bonafide
                WHERE school_id = %s
                AND student_id IN ({placeholders})
                ORDER BY id DESC
            """, tuple([school_id] + student_ids))

            history_rows = cursor.fetchall()

            for h in history_rows:

                sid = h["student_id"]

                if sid not in bonafide_history_map:
                    bonafide_history_map[sid] = []

                bonafide_history_map[sid].append({
                    "id": h["id"],
                    "number": h["bonafide_number"],
                    "date": format_date(h["date"]),
                    "purpose": h["purpose"] or "-"
                })

        # =========================================
        # RENDER
        # =========================================

        return render_template(

            "clerk/bonafide.html",

            role="clerk",

            school_name=school_name,
            school_code=school_code,

            active_page="bonafide",

            bonafide_list=bonafide_list,
            

            total_bonafide=total_bonafide,
            today_bonafide=today_bonafide,
            month_bonafide=month_bonafide,

            search=search,
            class_filter=class_filter,
            year_filter=year_filter,
            years=years,

            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total_records=total_records,

            bonafide_history_map=bonafide_history_map,

            next_bonafide_number="Auto Generate On Save"

        )

    except Exception as e:

        print("❌ BONAFIDE PAGE ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

# =========================================================
# 📄 SAVE BONAFIDE
# PURPOSE:
# Clerk can generate bonafide only for own school student
# Bonafide number + insert happen in one transaction
# =========================================================

@app.route("/clerk/bonafide/save", methods=["POST"])
@login_required
@subscription_required
@feature_required("enable_bonafide_management")
def save_bonafide():

    conn = None
    cursor = None

    try:

        # =========================================
        # CLERK SESSION CHECK
        # =========================================

        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        if session.get("clerk_role") != "clerk":
            return "Unauthorized ❌"

        # =========================================
        # GET FORM DATA
        # =========================================

        student_id = (
            request.form.get("student_id") or ""
        ).strip()

        purpose = (
            request.form.get("purpose") or ""
        ).strip()

        date_raw = (
            request.form.get("date") or ""
        ).strip()

        certificate_date = parse_date(date_raw)

        # =========================================
        # VALIDATION
        # =========================================

        if not student_id:
            return "Student ID missing ❌"

        try:
            student_id = int(student_id)
        except ValueError:
            return "Invalid Student ID ❌"

        if not purpose:
            return "Purpose required ❌"

        if len(purpose) > 255:
            return "Purpose too long ❌"

        if not certificate_date:
            return "Invalid date ❌"

        certificate_date_value = (
            certificate_date.date()
            if hasattr(certificate_date, "date")
            else certificate_date
        )

        # =========================================
        # DB CONNECTION
        # =========================================

        conn = get_connection()
        cursor = conn.cursor()

        # =========================================
        # CHECK STUDENT BELONGS TO SAME SCHOOL
        # Prevents cross-school certificate creation
        # =========================================

        cursor.execute("""
            SELECT id
            FROM students
            WHERE id = %s
            AND school_id = %s
            LIMIT 1
        """, (
            student_id,
            school_id
        ))

        student = cursor.fetchone()

        if not student:
            return "Student Not Found ❌"

        # =========================================
        # CHECK EXISTING BONAFIDE
        # Same purpose certificate will not duplicate
        # =========================================

        cursor.execute("""
            SELECT id
            FROM bonafide
            WHERE student_id = %s
            AND school_id = %s
            AND purpose = %s
            LIMIT 1
        """, (
            student_id,
            school_id,
            purpose
        ))

        existing_bonafide = cursor.fetchone()

        if existing_bonafide:
            return redirect(
                url_for(
                    "view_bonafide",
                    bid=existing_bonafide[0]
                )
            )

        # =========================================
        # GENERATE BONAFIDE NUMBER
        # Same cursor + same transaction
        # =========================================

        bonafide_number = generate_bonafide_number(
            cursor,
            school_id
        )

        # =========================================
        # INSERT BONAFIDE
        # =========================================

        cursor.execute("""
            INSERT INTO bonafide (
                school_id,
                student_id,
                bonafide_number,
                purpose,
                date
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            school_id,
            student_id,
            bonafide_number,
            purpose,
            certificate_date_value
        ))

        bonafide_id = cursor.lastrowid

        # =========================================
        # COMMIT BOTH:
        # sequence update + bonafide insert
        # =========================================

        conn.commit()

        print("✅ BONAFIDE SAVED")

        return redirect(
            url_for(
                "view_bonafide",
                bid=bonafide_id
            )
        )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ BONAFIDE SAVE ERROR:", e)

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()

            
# =========================================================
# 📄 DOWNLOAD BONAFIDE PDF
# =========================================================
@app.route("/clerk/bonafide/pdf/<int:bid>")
@login_required
@subscription_required
@feature_required("enable_bonafide_management")
def download_bonafide_pdf(bid):

    conn = None
    cursor = None

    try:

        mode = request.args.get("mode")

        school_id = session.get("clerk_school_id")

        conn = get_connection()
        cursor = conn.cursor()

        # =====================================================
        # ADMIN MODE
        # =====================================================
        if (
            mode == "admin"
            and session.get("admin_logged_in")
            and session.get("admin_role") == "admin"
        ):

            cursor.execute("""

                SELECT

                    b.*,

                    s.name,
                    s.school_register_no,
                    s.`class` AS class_name,
                    s.admission_date,
                    s.dob,
                    s.caste,
                    s.primary_mobile,

                    s.email AS student_email,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,

                    sc.email AS school_email,

                    sc.logo_path,
                    sc.watermark_path,
                    sc.website

                FROM bonafide b

                JOIN students s
                    ON b.student_id = s.id

                JOIN schools sc
                    ON b.school_id = sc.school_id

                WHERE b.id = %s

            """, (bid,))

        # =====================================================
        # CLERK MODE
        # =====================================================
        else:

            if not school_id:
                return "School session missing ❌"

            cursor.execute("""

                SELECT

                    b.*,

                    s.name,
                    s.school_register_no,
                    s.`class` AS class_name,
                    s.admission_date,
                    s.dob,
                    s.caste,
                    s.primary_mobile,

                    s.email AS student_email,

                    sc.name AS school_name,
                    sc.address,
                    sc.phone,

                    sc.email AS school_email,

                    sc.logo_path,
                    sc.watermark_path,
                    sc.website

                FROM bonafide b

                JOIN students s
                    ON b.student_id = s.id

                JOIN schools sc
                    ON b.school_id = sc.school_id

                WHERE b.id = %s
                AND b.school_id = %s

            """, (

                bid,
                school_id

            ))

        row = cursor.fetchone()

        if not row:
            return "Bonafide Not Found ❌"

        columns = [col[0] for col in cursor.description]
        row = dict(zip(columns, row))

        # =====================================================
        # BONAFIDE DATA
        # =====================================================

        bonafide = {

            "id": row["id"],

            "bonafide_number": row["bonafide_number"],

            "purpose": row["purpose"] or "",

            "date": format_date(row["date"])

        }

        # =====================================================
        # STUDENT DATA
        # =====================================================

        student = {

            "name": row["name"] or "",

            "school_register_no": row["school_register_no"] or "",

            "class": row["class_name"] or "",

            "admission_date": format_date(
                row["admission_date"]
            ),

            "dob": format_date(row["dob"]),

            "caste": row["caste"] or "",

            "primary_mobile": row["primary_mobile"] or "",

            "email": row["student_email"] or ""

        }

        # =====================================================
        # SCHOOL DATA
        # =====================================================

        base_dir = os.path.abspath(
            os.path.dirname(__file__)
        )

        logo_absolute = ""
        watermark_absolute = ""

        if row["logo_path"]:

            logo_absolute = "file:///" + os.path.join(

                base_dir,
                "static",
                row["logo_path"].replace("static/", "")

            ).replace("\\", "/")

        if row["watermark_path"]:

            watermark_absolute = "file:///" + os.path.join(

                base_dir,
                "static",
                row["watermark_path"].replace("static/", "")

            ).replace("\\", "/")

        school = {

            "name": row["school_name"] or "",

            "address": row["address"] or "",

            "phone": row["phone"] or "",

            "email": row["school_email"] or "",

            "logo_path": logo_absolute,

            "watermark_path": watermark_absolute,

            "website": row["website"] or ""

        }

        # =====================================================
        # HTML TEMPLATE
        # =====================================================

        html = render_template(

            "clerk/bonafide_generate.html",

            student=student,

            bonafide=bonafide,

            school=school,

            is_pdf=True

        )

        # =====================================================
        # PDF FOLDER
        # =====================================================

        pdf_folder = os.path.join(

            "static",
            "generated_bonafide"

        )

        if not os.path.exists(pdf_folder):

            os.makedirs(pdf_folder)

        pdf_path = os.path.join(

            pdf_folder,

            f'bonafide_{bonafide["id"]}.pdf'

        )

        # =====================================================
        # PDF OPTIONS
        # =====================================================

        options = {

            "page-size": "A4",

            "margin-top": "5mm",

            "margin-right": "5mm",

            "margin-bottom": "5mm",

            "margin-left": "5mm",

            "encoding": "UTF-8",

            "enable-local-file-access": ""

        }

        # =====================================================
        # GENERATE PDF
        # =====================================================

        pdfkit.from_string(

            html,

            pdf_path,

            configuration=pdf_config,

            options=options

        )

        # =====================================================
        # 📧 AUTO SEND BONAFIDE PDF EMAIL
        # =====================================================

        try:

            student_email = student.get("email")

            if student_email:

                # =========================================
                # CHECK ALREADY SENT
                # =========================================

                cursor.execute("""

                    SELECT id
                    FROM bonafide_email_logs
                    WHERE school_id = %s
                    AND bonafide_id = %s

                """, (

                    row["school_id"],
                    bonafide["id"]

                ))

                already_sent = cursor.fetchone()

                # =========================================
                # SEND ONLY ONCE
                # =========================================

                if student_email and not already_sent:

                    subject = (
                        f'Bonafide Certificate - {student["name"]}'
                    )

                    body = f"""

                    <div style="font-family:Arial;padding:20px;">

                        <h2 style="color:#14b8a6;">
                            Bonafide Certificate Generated
                        </h2>

                        <p>
                            Dear Parent/Student,
                        </p>

                        <p>
                            Your Bonafide Certificate
                            has been generated successfully.
                        </p>

                        <hr>

                        <p>
                            <b>Student Name:</b>
                            {student["name"]}
                        </p>

                        <p>
                            <b>Bonafide Number:</b>
                            {bonafide["bonafide_number"]}
                        </p>

                        <p>
                            <b>School:</b>
                            {school["name"]}
                        </p>

                        <hr>

                        <p>
                            Please find the Bonafide PDF attached.
                        </p>

                        <p>
                            Thank you,
                            <br>
                            <b>{school["name"]}</b>
                        </p>

                    </div>

                    """

            # =========================================
            # SEND MAIL
            # =========================================

            email_sent = send_email(

                student_email,
                subject,
                body,
                pdf_path,
                f"{bonafide['bonafide_number']}.pdf"

            )

            # =========================================
            # EMAIL SUCCESS
            # =========================================

            if email_sent == True:

                print("✅ BONAFIDE PDF EMAIL SENT")

                # =====================================
                # SAVE EMAIL LOG
                # =====================================

                cursor.execute("""

                    INSERT INTO bonafide_email_logs
                    (
                        school_id,
                        bonafide_id,
                        student_email
                    )
                    VALUES (%s, %s, %s)

                """, (

                    row["school_id"],
                    bonafide["id"],
                    student_email

                ))

                conn.commit()

            # =========================================
            # EMAIL FAILED
            # =========================================

            else:

                print(
                    "❌ BONAFIDE EMAIL FAILED:",
                    email_sent
                )

        except Exception as email_error:

            print(
                "❌ BONAFIDE EMAIL ERROR:",
                email_error
            )

        # =====================================================
        # RETURN PDF
        # =====================================================

        return send_file(

            pdf_path,

            as_attachment=True,

            download_name=f'{bonafide["bonafide_number"]}.pdf'

        )

    except Exception as e:

        print(
            "❌ BONAFIDE PDF ERROR:",
            e
        )

        return "Something went wrong ❌"

    finally:

        if cursor:
            cursor.close()

        if conn:
            conn.close()            
            
 # =========================================================
# 🔓 PUBLIC BONAFIDE PDF DOWNLOAD - NO LOGIN
# =========================================================
@app.route("/public/bonafide/pdf/<int:bid>")
def public_bonafide_pdf(bid):

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                b.*,
                s.name,
                s.school_register_no,
                s.class AS class_name,
                s.admission_date,
                s.dob,
                s.caste,
                s.primary_mobile,
                s.email AS student_email,

                sc.name AS school_name,
                sc.address,
                sc.phone,
                sc.email AS school_email,
                sc.logo_path,
                sc.watermark_path,
                sc.website
            FROM bonafide b
            JOIN students s
                ON b.student_id = s.id
                AND b.school_id = s.school_id
            JOIN schools sc
                ON b.school_id = sc.school_id
            WHERE b.id = %s
            LIMIT 1
        """, (bid,))

        row = cursor.fetchone()

        if not row:
            return "Bonafide Not Found ❌"

        bonafide = {
            "id": row["id"],
            "bonafide_number": row["bonafide_number"],
            "purpose": row["purpose"] or "",
            "date": format_date(row["date"])
        }

        student = {
            "name": row["name"] or "",
            "school_register_no": row["school_register_no"] or "",
            "class": row["class_name"] or "",
            "admission_date": format_date(row["admission_date"]),
            "dob": format_date(row["dob"]),
            "caste": row["caste"] or "",
            "primary_mobile": row["primary_mobile"] or "",
            "email": row["student_email"] or ""
        }

        base_dir = os.path.abspath(os.path.dirname(__file__))

        logo_absolute = ""
        watermark_absolute = ""

        if row["logo_path"]:
            logo_absolute = "file:///" + os.path.join(
                base_dir,
                "static",
                row["logo_path"].replace("static/", "")
            ).replace("\\", "/")

        if row["watermark_path"]:
            watermark_absolute = "file:///" + os.path.join(
                base_dir,
                "static",
                row["watermark_path"].replace("static/", "")
            ).replace("\\", "/")

        school = {
            "name": row["school_name"] or "",
            "address": row["address"] or "",
            "phone": row["phone"] or "",
            "email": row["school_email"] or "",
            "logo_path": logo_absolute,
            "watermark_path": watermark_absolute,
            "website": row["website"] or ""
        }

        html = render_template(
            "clerk/bonafide_generate.html",
            student=student,
            bonafide=bonafide,
            school=school,
            is_pdf=True,
            is_public=True
        )

        pdf_folder = os.path.join("static", "generated_bonafide")
        os.makedirs(pdf_folder, exist_ok=True)

        pdf_path = os.path.join(
            pdf_folder,
            f"bonafide_{bonafide['id']}.pdf"
        )

        options = {
            "page-size": "A4",
            "margin-top": "5mm",
            "margin-right": "5mm",
            "margin-bottom": "5mm",
            "margin-left": "5mm",
            "encoding": "UTF-8",
            "enable-local-file-access": ""
        }

        pdfkit.from_string(
            html,
            pdf_path,
            configuration=pdf_config,
            options=options
        )

        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=f"{bonafide['bonafide_number']}.pdf"
        )

    except Exception as e:
        print("❌ PUBLIC BONAFIDE PDF ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# =========================================================
# IMPORT - EXPORT PAGE ROUTE
# =========================================================
@app.route("/clerk/import-export")
@login_required
@subscription_required
@feature_required("enable_import_export")
def import_export_page():

    conn = None
    cursor = None

    try:
        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM students
            WHERE school_id = %s
        """, (school_id,))

        total_students = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM students
            WHERE school_id = %s
            AND DATE(created_at) = CURDATE()
        """, (school_id,))

        today_students = cursor.fetchone()["total"] or 0

        cursor.execute("""
            SELECT DISTINCT YEAR(admission_date) AS year_no
            FROM students
            WHERE school_id = %s
            AND admission_date IS NOT NULL
            ORDER BY year_no DESC
        """, (school_id,))

        years = [
            row["year_no"]
            for row in cursor.fetchall()
            if row["year_no"]
        ]

        return render_template(
            "clerk/import_export.html",
            role="clerk",
            school_name=school["school_name"],
            school_udise=school["school_udise"],
            active_page="import_export",
            total_students=total_students,
            today_students=today_students,
            years=years
        )

    except Exception as e:
        print("❌ IMPORT EXPORT PAGE ERROR:", e)
        return "Something went wrong ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# =========================================================
# 📥 IMPORT STUDENTS FROM EXCEL (SAFE + ATOMIC)
# =========================================================
@app.route("/clerk/import-students", methods=["GET", "POST"])
@login_required
@subscription_required
@feature_required("enable_import_export")
def import_students():

    if request.method == "GET":
        return redirect("/clerk/import-export")

    print("🔥 IMPORT STARTED")

    conn = None
    cursor = None

    try:
        file = request.files.get("file")

        # ================= FILE CHECK =================
        if not file or not file.filename:

            return "No file uploaded ❌"

        # ================= SAFE FILENAME =================
        filename = secure_filename(file.filename)

        # ================= EXTENSION CHECK =================
        if not filename.lower().endswith(".xlsx"):

            return "Only .xlsx files allowed ❌"
        

        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        # ================= READ EXCEL =================
        df = pd.read_excel(file, dtype=str)

        if df.empty:
            return "Excel file is empty ❌"

        # ================= LIMIT ROWS =================
        if len(df) > 5000:
            return "Maximum 5000 rows allowed ❌"

        # ================= NORMALIZE COLUMNS =================
        df.columns = df.columns.str.strip().str.lower()

        print("📊 Columns:", df.columns)

       # ================= REQUIRED COLUMN CHECK =================
        required_columns = ["name", "class"]

        for col in required_columns:
            if col not in df.columns:
                return f"Missing required column: {col} ❌"

        conn = get_connection()
        cursor = conn.cursor()

        inserted_count = 0

        for _, row in df.iterrows():

            # ================= CLEAN NULLS =================
            row = row.where(pd.notnull(row), None)

            # ================= BASIC REQUIRED =================
            name = str(row.get("name")).strip()
            class_name = str(row.get("class")).strip()

            # ================= SKIP INVALID =================
            if not name or not class_name:
                print("⚠️ Row skipped (missing required data)")
                continue

            # ================= NEW FIELDS =================
            school_register_no = row.get("school_register_no")

            student_uid = row.get("student_uid")
            if student_uid:
                student_uid = str(student_uid).split(".")[0].strip()

            apaar_id = row.get("apaar_id")
            if apaar_id:
                apaar_id = str(apaar_id).split(".")[0].strip()

           # ================= DATE FIX =================
            dob = row.get("dob")

            if dob:
                try:
                    if hasattr(dob, "strftime"):
                        dob = dob.strftime("%Y-%m-%d")
                    else:
                        dob = pd.to_datetime(
                            str(dob),
                            dayfirst=True
                        ).strftime("%Y-%m-%d")
                except:
                    dob = None


            admission_date = row.get("admission_date")

            if admission_date:
                try:
                    if hasattr(admission_date, "strftime"):
                        admission_date = admission_date.strftime("%Y-%m-%d")
                    else:
                        admission_date = pd.to_datetime(
                            str(admission_date),
                            dayfirst=True
                        ).strftime("%Y-%m-%d")
                except:
                    admission_date = None

            # ================= OTHER FIELDS =================
            caste = row.get("caste")
            father_name = row.get("father_name")
            mother_name = row.get("mother_name")
            aadhaar = row.get("aadhaar")

            if aadhaar:
                aadhaar = str(aadhaar).split(".")[0].strip()
            else:
                aadhaar = None

            birth_place = row.get("birth_place")
            nationality = row.get("nationality")
            mother_tongue = row.get("mother_tongue")
            religion = row.get("religion")

            city = row.get("city")
            taluka = row.get("taluka")
            district = row.get("district")
            state = row.get("state")

            section = row.get("section")
            previous_school = row.get("previous_school")

            last_exam = row.get("last_exam")
            result_status = row.get("result_status")

            progress = row.get("progress")
            conduct = row.get("conduct")

            primary_mobile = row.get("primary_mobile")
            alternate_mobile = row.get("alternate_mobile")

            email = row.get("email")

            if email:
                email = str(email).strip().lower()

                if not is_valid_email(email):
                    print("⚠️ Invalid email skipped")
                    continue

            occupation = row.get("occupation")

            income = row.get("income")

            if income:
                income = str(income).replace(",", "").strip()
                try:
                    income = int(float(income))
                except:
                    income = None
            else:
                income = None

            guardian_name = row.get("guardian_name")
            guardian_mobile = row.get("guardian_mobile")

            # ================= AADHAAR VALIDATION =================
            if aadhaar:
                aadhaar = str(aadhaar).strip()

                if not aadhaar.isdigit() or len(aadhaar) != 12:
                    print("⚠️ Invalid Aadhaar skipped")
                    continue

            # ================= MOBILE CLEANUP =================
            if primary_mobile:
                primary_mobile = str(primary_mobile).split(".")[0].strip()

            if alternate_mobile:
                alternate_mobile = str(alternate_mobile).split(".")[0].strip()

            if guardian_mobile:
                guardian_mobile = str(guardian_mobile).split(".")[0].strip()

            print("Processing:", name)


            # ================= DUPLICATE CHECK (FIXED) =================
            existing_student = None

            # check aadhaar only if valid and exists
            if aadhaar:
                cursor.execute("""
                    SELECT id
                    FROM students
                    WHERE school_id = %s
                    AND aadhaar = %s
                               LIMIT 1
                """, (
                    school_id,
                    aadhaar
                ))
                existing_student = cursor.fetchone()

            # check student uid only if aadhaar not matched
            if not existing_student and student_uid:
                cursor.execute("""
                    SELECT id
                    FROM students
                    WHERE school_id = %s
                    AND student_uid = %s
                     LIMIT 1  
                """, (
                    school_id,
                    student_uid
                ))
                existing_student = cursor.fetchone()

            # skip only real duplicates
            if existing_student:
                print(
                    "⚠️ Duplicate student skipped:",
                    name,
                    aadhaar,
                    student_uid
                )
                continue
            # ================= SAFE ATOMIC ADMISSION =================
            admission_no = generate_admission_no(school_id)


               # ================= INSERT =================
            cursor.execute("""
                INSERT INTO students (
                    school_id,
                    school_register_no,
                    name,
                    father_name,
                    mother_name,
                    student_uid,
                    aadhaar,
                    apaar_id,
                    dob,
                    birth_place,
                    nationality,
                    mother_tongue,
                    religion,
                    caste,
                    city,
                    taluka,
                    district,
                    state,
                    admission_no,
                    admission_date,
                    `class`,
                    section,
                    previous_school,
                    last_exam,
                    result_status,
                    progress,
                    conduct,
                    primary_mobile,
                    alternate_mobile,
                    occupation,
                    income,
                    guardian_name,
                    guardian_mobile,
                    email
                )
                VALUES (%s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                school_id,
                school_register_no,
                name,
                father_name,
                mother_name,
                student_uid,
                aadhaar,
                apaar_id,
                dob,
                birth_place,
                nationality,
                mother_tongue,
                religion,
                caste,
                city,
                taluka,
                district,
                state,
                admission_no,
                admission_date,
                class_name,
                section,
                previous_school,
                last_exam,
                result_status,
                progress,
                conduct,
                primary_mobile,
                alternate_mobile,
                occupation,
                income,
                guardian_name,
                guardian_mobile,
                email
            ))

            inserted_count += 1

        conn.commit()

        print("✅ IMPORT SUCCESS:", inserted_count)

        return redirect(
        f"/clerk/import-export?success={inserted_count}"
    )

    except Exception as e:

        if conn:
            conn.rollback()

        print("❌ IMPORT ERROR:", e)
        return "Import failed ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# =========================================================
# EXPORT STUDENTS TO EXCEL
# =========================================================
@app.route("/clerk/export-students")
@login_required
@subscription_required
@feature_required("enable_import_export")
def export_students():

    conn = None
    cursor = None

    try:
        import io
        import pandas as pd
        from datetime import datetime

        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        cls = (request.args.get("class") or "").strip()
        month = (request.args.get("month") or "").strip()
        year = (request.args.get("year") or "").strip()

        if cls and not cls.isdigit():
            return "Invalid class filter ❌"

        if month and (not month.isdigit() or int(month) < 1 or int(month) > 12):
            return "Invalid month filter ❌"

        if year and not year.isdigit():
            return "Invalid year filter ❌"

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
            SELECT
                school_register_no,
                name,
                father_name,
                mother_name,
                student_uid,
                aadhaar,
                apaar_id,
                dob,
                birth_place,
                nationality,
                mother_tongue,
                religion,
                caste,
                city,
                taluka,
                district,
                state,
                admission_no,
                admission_date,
                `class`,
                section,
                previous_school,
                last_exam,
                result_status,
                progress,
                conduct,
                primary_mobile,
                alternate_mobile,
                email,
                occupation,
                income,
                guardian_name,
                guardian_mobile
            FROM students
            WHERE school_id = %s
        """

        params = [school_id]

        if cls:
            query += " AND `class` = %s"
            params.append(cls)

        if month:
            query += " AND MONTH(admission_date) = %s"
            params.append(month)

        if year:
            query += " AND YEAR(admission_date) = %s"
            params.append(year)

        query += " ORDER BY id DESC"

        cursor.execute(query, tuple(params))
        students = cursor.fetchall()

        if not students:
            return "No data found for selected filters ❌"

        def safe_date(value):
            if not value:
                return ""

            try:
                return value.strftime("%d-%m-%Y")
            except Exception:
                return str(value)

        def sanitize_excel(value):
            if value is None:
                return ""

            if isinstance(value, str):
                value = value.strip()

                if value.startswith(("=", "+", "-", "@")):
                    return "'" + value

            return value

        for student in students:
            student["dob"] = safe_date(student.get("dob"))
            student["admission_date"] = safe_date(student.get("admission_date"))

            for key in student:
                student[key] = sanitize_excel(student[key])

        df = pd.DataFrame(students)

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(
                writer,
                index=False,
                sheet_name="Students"
            )

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"students_export_{school_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("❌ EXPORT ERROR:", e)
        return "Export failed ❌"

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            
                
 
@app.route("/coming-soon/<feature>")
@login_required
@subscription_required
def coming_soon(feature):

    try:
        school_id = session.get("clerk_school_id")

        if not school_id:
            return "School session missing ❌"

        feature = (feature or "").lower().strip()

        feature_names = {
            "attendance": "Attendance Management",
            "marks": "Marks & Results",
            "teachers": "Teacher Management",
            "fees": "Fee Management",
            "timetable": "Timetable",
            "notice-board": "Notice Board"
        }

        if feature not in feature_names:
            return "Invalid feature ❌"

        school = get_school_details(school_id)

        if not school:
            return "School not found ❌"

        return render_template(
            "features/coming_soon.html",
            feature=feature_names[feature],
            role="clerk",
            school_name=school["school_name"],
            school_udise=school["school_udise"],
            active_page=feature
        )

    except Exception as e:
        print("❌ COMING SOON ERROR:", e)
        return "Something went wrong ❌"
    
    
# =========================================================
# 📊 ATTENDANCE MODULE
# =========================================================

@app.route("/clerk/attendance")
@login_required
@subscription_required
@feature_required("enable_attendance")
def attendance_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="attendance"
        )
    )
# =========================================================
#  MARKS & RESULTS MODULE
# =========================================================
@app.route("/clerk/results")
@login_required
@subscription_required
@feature_required("enable_results")
def results_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="marks"
        )
    )
# =========================================================
# TEACHERS MODULE
# =========================================================
@app.route("/clerk/teachers")
@login_required
@subscription_required
@feature_required("enable_teacher_management")
def teachers_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="teachers"
        )
    )

# =========================================================
# 💰 FEE MANAGEMENT
# =========================================================

@app.route("/clerk/fees")
@login_required
@subscription_required
@feature_required("enable_fee_management")
def fee_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="fees"
        )
    )
# =========================================================
# TIMETABLE MODULE
# =========================================================
@app.route("/clerk/timetable")
@login_required
@subscription_required
@feature_required("enable_timetable")
def timetable_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="timetable"
        )
    )
# =========================================================
# NOTICE BOARD MODULE
# =========================================================
@app.route("/clerk/notice-board")
@login_required
@subscription_required
@feature_required("enable_notice_board")
def notice_board_dashboard():

    return redirect(
        url_for(
            "coming_soon",
            feature="notice-board"
        )
    )


# =========================================================
# 🚀 RUN APP
# =========================================================
if __name__ == '__main__':
    app.run(debug=True)